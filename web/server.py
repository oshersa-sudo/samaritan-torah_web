# -*- coding: utf-8 -*-
"""
Web edition of the Samaritan Torah app — a Flask backend that REUSES the existing
query layer (app/services/database.py) and serves a single-page browser UI with
full feature parity. It is fully isolated: it never writes to the database (only
SELECTs run), and it touches nothing under app/, main.py or buildozer.spec. The
original Kivy app keeps working unchanged.

Run:  py -3 web/server.py     →  http://127.0.0.1:5000
"""
import os
import sys
import re
import sqlite3
import difflib

# make the project root importable so we reuse the app's own service layer
_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _ROOT not in sys.path:
    sys.path.insert(0, _ROOT)
# also make this directory importable for local sibling modules (analytics.py) —
# needed under gunicorn, which imports us as the package module "web.server"
# rather than running this file as a script (which would add it automatically)
_WEB_DIR = os.path.dirname(os.path.abspath(__file__))
if _WEB_DIR not in sys.path:
    sys.path.insert(0, _WEB_DIR)

from flask import Flask, jsonify, request, render_template, send_from_directory, send_file, redirect

from app.services import database as db
from app.services.interpreter import get_chapter_interpretations
import analytics

app = Flask(__name__, static_folder='static', template_folder='templates')
app.config['TEMPLATES_AUTO_RELOAD'] = True

# Running behind nginx (reverse proxy) in production: honour the X-Forwarded-*
# headers so the app sees the real client IP, host and https scheme.
try:
    from werkzeug.middleware.proxy_fix import ProxyFix
    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)
except Exception:
    pass

APP_VERSION = '3.3'
_VER_UPDATES = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'VER_UPDATES.txt')
_SYSTEM_DOC = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'SYSTEM_DOC.txt')


# ── admin text editing (LOCAL only; secured entirely server-side) ─────────────
# Secrets come from a gitignored .env that exists only on the maintainer's
# machine; the public deployment has no .env, so ADMIN_PASSWORD is empty and
# admin login / editing is disabled there. The password is NEVER sent to the
# client. Edits are restricted to a whitelist of (table, column) pairs and
# require a valid session token, so nothing else can be written.
def _load_dotenv():
    p = os.path.join(_ROOT, '.env')
    if os.path.exists(p):
        for line in open(p, encoding='utf-8'):
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                k, v = line.split('=', 1)
                os.environ.setdefault(k.strip(), v.strip().strip('"\''))


_load_dotenv()
import secrets, hmac, hashlib, time
ADMIN_USER = os.environ.get('ADMIN_USER', 'oshersa')
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', '')
_TOKEN_TTL = 12 * 3600                 # admin session token lifetime (seconds)
_LOGIN_FAILS = {}                      # ip -> [failure timestamps] (best-effort throttle)
_EDITABLE = {'verses': {'text', 'masoretic_text', 'interpretation', 'sam_aramaic',
                        'sam_hebrew', 'simple_hebrew', 'english', 'arabic_trans',
                        # per-version comparison texts, editable from the compare-view pencils
                        'lxx_text', 'onkelos_text', 'qumran_text',
                        # override for the Masoretic-comparison chapter number (e.g. "20:1")
                        'mas_chapter'},
             # per-line edits inside a saved private composition (חיבורים פרטיים)
             'private_composition_lines': {'text'}}


def _make_token():
    """Stateless signed token (works across gunicorn workers; carries its own expiry).
    Signature is keyed by the secret password, so it can't be forged without it."""
    ts = str(int(time.time()))
    sig = hmac.new(ADMIN_PASSWORD.encode(), ts.encode(), hashlib.sha256).hexdigest()
    return ts + '.' + sig


def _valid_token(tok):
    if not ADMIN_PASSWORD or not tok or '.' not in str(tok):
        return False
    ts, _, sig = str(tok).partition('.')
    if not ts.isdigit() or time.time() - int(ts) > _TOKEN_TTL:
        return False
    good = hmac.new(ADMIN_PASSWORD.encode(), ts.encode(), hashlib.sha256).hexdigest()
    return hmac.compare_digest(sig, good)


def _throttled(ip):
    """True if this IP has too many recent failed logins (speed-bump against guessing)."""
    now = time.time()
    fails = [t for t in _LOGIN_FAILS.get(ip, []) if now - t < 600]   # 10-min window
    _LOGIN_FAILS[ip] = fails
    return len(fails) >= 8


@app.route('/api/admin/status')
def admin_status():
    return jsonify({'enabled': bool(ADMIN_PASSWORD),
                     'webauthn': bool(ADMIN_PASSWORD) and analytics.wa_has_credential()})


@app.route('/api/admin/login', methods=['POST'])
def admin_login():
    if not ADMIN_PASSWORD:
        return jsonify({'ok': False, 'disabled': True})
    ip = request.headers.get('X-Forwarded-For', request.remote_addr or '').split(',')[0].strip()
    if _throttled(ip):
        return jsonify({'ok': False, 'error': 'too many attempts'}), 429
    d = request.get_json(silent=True) or {}
    u, p = str(d.get('user', '')), str(d.get('password', ''))
    if hmac.compare_digest(u, ADMIN_USER) and hmac.compare_digest(p, ADMIN_PASSWORD):
        return jsonify({'ok': True, 'token': _make_token()})
    _LOGIN_FAILS.setdefault(ip, []).append(time.time())
    return jsonify({'ok': False})


# ── WebAuthn (phone fingerprint / Face ID) as a second factor for the admin
# login: register once with the password, then sign in with just the platform
# authenticator. The challenge is carried in a signed, stateless "state" string
# (same HMAC trick as the session token) rather than server-side session memory,
# so it survives being routed to a different gunicorn worker between the two
# steps of the ceremony.
import json as _json
import webauthn
from webauthn.helpers import bytes_to_base64url, base64url_to_bytes
from webauthn.helpers.structs import (
    PublicKeyCredentialDescriptor, AuthenticatorSelectionCriteria,
    AuthenticatorAttachment, ResidentKeyRequirement, UserVerificationRequirement,
)


def _wa_rp_id():
    return request.host.split(':')[0]


def _wa_origin():
    return request.host_url.rstrip('/')


def _wa_state(challenge):
    b64 = bytes_to_base64url(challenge)
    exp = str(int(time.time()) + 300)
    sig = hmac.new(ADMIN_PASSWORD.encode(), (b64 + '.' + exp).encode(), hashlib.sha256).hexdigest()
    return b64 + '.' + exp + '.' + sig


def _wa_challenge(state):
    try:
        b64, exp, sig = str(state).split('.')
    except Exception:
        return None
    if not exp.isdigit() or time.time() > int(exp):
        return None
    good = hmac.new(ADMIN_PASSWORD.encode(), (b64 + '.' + exp).encode(), hashlib.sha256).hexdigest()
    if not hmac.compare_digest(sig, good):
        return None
    return base64url_to_bytes(b64)


def _wa_exclude_or_allow():
    return [PublicKeyCredentialDescriptor(id=base64url_to_bytes(cid)) for cid in analytics.wa_credential_ids()]


@app.route('/api/admin/webauthn/register_options', methods=['POST'])
def wa_register_options():
    d = request.get_json(silent=True) or {}
    if not _valid_token(d.get('token')):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    challenge = secrets.token_bytes(32)
    opts = webauthn.generate_registration_options(
        rp_id=_wa_rp_id(), rp_name='התורה השומרונית הישראלית',
        user_name=ADMIN_USER, user_display_name=ADMIN_USER, user_id=ADMIN_USER.encode(),
        challenge=challenge,
        authenticator_selection=AuthenticatorSelectionCriteria(
            authenticator_attachment=AuthenticatorAttachment.PLATFORM,
            resident_key=ResidentKeyRequirement.PREFERRED,
            user_verification=UserVerificationRequirement.REQUIRED),
        exclude_credentials=_wa_exclude_or_allow(),
    )
    return jsonify({'ok': True, 'options': _json.loads(webauthn.options_to_json(opts)), 'state': _wa_state(challenge)})


@app.route('/api/admin/webauthn/register_verify', methods=['POST'])
def wa_register_verify():
    d = request.get_json(silent=True) or {}
    if not _valid_token(d.get('token')):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    challenge = _wa_challenge(d.get('state'))
    if not challenge:
        return jsonify({'ok': False, 'error': 'expired, try again'}), 400
    try:
        verified = webauthn.verify_registration_response(
            credential=d.get('credential'), expected_challenge=challenge,
            expected_rp_id=_wa_rp_id(), expected_origin=_wa_origin())
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400
    analytics.wa_add_credential(bytes_to_base64url(verified.credential_id),
                                 verified.credential_public_key, verified.sign_count)
    return jsonify({'ok': True})


@app.route('/api/admin/webauthn/login_options')
def wa_login_options():
    if not ADMIN_PASSWORD or not analytics.wa_has_credential():
        return jsonify({'ok': False, 'error': 'not available'}), 400
    challenge = secrets.token_bytes(32)
    opts = webauthn.generate_authentication_options(
        rp_id=_wa_rp_id(), challenge=challenge, allow_credentials=_wa_exclude_or_allow(),
        user_verification=UserVerificationRequirement.REQUIRED)
    return jsonify({'ok': True, 'options': _json.loads(webauthn.options_to_json(opts)), 'state': _wa_state(challenge)})


@app.route('/api/admin/webauthn/login_verify', methods=['POST'])
def wa_login_verify():
    if not ADMIN_PASSWORD:
        return jsonify({'ok': False}), 400
    d = request.get_json(silent=True) or {}
    challenge = _wa_challenge(d.get('state'))
    if not challenge:
        return jsonify({'ok': False, 'error': 'expired, try again'}), 400
    cred = d.get('credential') or {}
    cred_id = cred.get('id')
    stored = analytics.wa_get_credential(cred_id) if cred_id else None
    if not stored:
        return jsonify({'ok': False, 'error': 'unknown credential'}), 400
    try:
        verified = webauthn.verify_authentication_response(
            credential=cred, expected_challenge=challenge,
            expected_rp_id=_wa_rp_id(), expected_origin=_wa_origin(),
            credential_public_key=stored['public_key'], credential_current_sign_count=stored['sign_count'])
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400
    analytics.wa_update_sign_count(cred_id, verified.new_sign_count)
    return jsonify({'ok': True, 'token': _make_token()})


@app.route('/api/track', methods=['POST'])
def track_visit():
    """Public visit beacon (called by every visitor's browser, not just admins) —
    feeds the admin analytics dashboard. IP/User-Agent are read from the request
    itself (never trusted from the client body)."""
    d = request.get_json(silent=True) or {}
    ip = request.headers.get('X-Forwarded-For', request.remote_addr or '').split(',')[0].strip()
    ua = request.headers.get('User-Agent', '')
    sid = str(d.get('sid', ''))[:64]
    path = str(d.get('path', '') or '')[:200]
    title = str(d.get('title', '') or '')[:200]
    try:
        analytics.track(sid, ip, ua, path, title)
    except Exception:
        pass
    return jsonify({'ok': True})


@app.route('/api/admin/analytics')
def admin_analytics():
    if not _valid_token(request.args.get('token')):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    return jsonify({'ok': True, 'sessions': analytics.recent_sessions(300)})


@app.route('/api/admin/download_db')
def admin_download_db():
    """Download the live DB (with online edits) so it can be committed back to git.
    Token passed as a query param since this is a direct download link."""
    if not _valid_token(request.args.get('token')):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    from flask import send_file
    return send_file(db.DB_PATH, as_attachment=True, download_name='torah.db')


_RESEED_CONTENT_TABLES = [
    ('verse_dictionary', 'verse_id'), ('word_gloss', 'verse_id'), ('word_jewish', 'verse_id'),
    ('word_samaritan', 'verse_id'), ('word_align', 'verse_id'), ('verse_translit', 'verse_id'),
    ('vongall_apparatus', 'verse_id'), ('binyamim_verse_links', 'verse_id'),
    ('eyalk_verse_links', 'verse_id'), ('shyt_verse_links', 'verse_id'), ('sir_verse_links', 'verse_id'),
    ('tm_verse_links', 'verse_id'), ('tradart_verse_links', 'verse_id'), ('tzdaka_verse_links', 'verse_id'),
    ('bhuq_verse_links', 'verse_id'), ('asatir_verse_links', 'verse_id'),
]


def _reseed_state_hash(bundled, live):
    """Cheap fingerprint of both DB files' current on-disk state (mtime+size) —
    used to bind a diff report to the exact state it was computed from, so an
    approval can't silently apply to a live DB that changed after the report
    was shown."""
    parts = []
    for p in (bundled, live):
        try:
            st = os.stat(p)
            parts.append('%d:%d' % (int(st.st_mtime), st.st_size))
        except OSError:
            parts.append('missing')
    return hashlib.sha256('|'.join(parts).encode()).hexdigest()[:24]


def _reseed_diff_report():
    """Compare the BUNDLED (git) DB about to be copied over the LIVE (persistent-
    disk) DB — sam_chapters/verses added-removed-changed per book, verse text
    changes, and two specific risk checks the project owner asked to always see
    before a reseed: (1) commentary/dictionary content that exists on the LIVE
    verse but would vanish because the bundled verse lacks it, and (2) reading/
    witness audio-manifest entries (shipped with the bundled code) that would
    point at a Samaritan chapter id/number the bundled DB doesn't actually have —
    both are exactly the kind of silent disconnection a blind file copy can cause."""
    bundled = getattr(db, '_BUNDLED_DB', None)
    live = db.DB_PATH
    if not bundled or not os.path.exists(bundled):
        return {'noop': True, 'reason': 'no bundled DB found'}
    if live == bundled:
        return {'noop': True, 'reason': 'not running off a separate disk — local no-op'}
    if not os.path.exists(live):
        return {'noop': False, 'first_seed': True}

    bconn = sqlite3.connect(bundled); bconn.row_factory = sqlite3.Row
    lconn = sqlite3.connect(live); lconn.row_factory = sqlite3.Row

    books = [dict(r) for r in lconn.execute('SELECT id, name FROM books ORDER BY id')]
    book_reports = []
    for bk in books:
        bid = bk['id']
        b_sc = {r['id']: r['number'] for r in bconn.execute('SELECT id, number FROM sam_chapters WHERE book_id=?', (bid,))}
        l_sc = {r['id']: r['number'] for r in lconn.execute('SELECT id, number FROM sam_chapters WHERE book_id=?', (bid,))}
        added = sorted(set(b_sc) - set(l_sc))
        removed = sorted(set(l_sc) - set(b_sc))
        renumbered = [{'id': i, 'live': l_sc[i], 'bundled': b_sc[i]}
                      for i in sorted(set(b_sc) & set(l_sc)) if b_sc[i] != l_sc[i]]
        if added or removed or renumbered or len(b_sc) != len(l_sc):
            book_reports.append({
                'book_id': bid, 'name': bk['name'],
                'sam_count_live': len(l_sc), 'sam_count_bundled': len(b_sc),
                'added': len(added), 'removed': len(removed), 'renumbered': len(renumbered),
                'renumbered_sample': renumbered[:10],
            })

    b_v = {r['id']: (r['sam_ch_id'], r['text']) for r in bconn.execute('SELECT id, sam_ch_id, text FROM verses')}
    l_v = {r['id']: (r['sam_ch_id'], r['text']) for r in lconn.execute('SELECT id, sam_ch_id, text FROM verses')}
    v_added = sorted(set(b_v) - set(l_v))
    v_removed = sorted(set(l_v) - set(b_v))
    common = set(b_v) & set(l_v)
    text_changed = [vid for vid in common if b_v[vid][1] != l_v[vid][1]]
    sam_ch_changed = [vid for vid in common if b_v[vid][0] != l_v[vid][0]]

    def _loc(conn, vid):
        r = conn.execute("""SELECT bk.name book, c.number cn, v.number vn FROM verses v
            JOIN chapters c ON c.id=v.chapter_id JOIN books bk ON bk.id=c.book_id WHERE v.id=?""", (vid,)).fetchone()
        return '%s %s:%s' % (r['book'], r['cn'], r['vn']) if r else str(vid)

    text_sample = [{'verse_id': vid, 'ref': _loc(lconn, vid),
                     'live': (l_v[vid][1] or '')[:80], 'bundled': (b_v[vid][1] or '')[:80]}
                    for vid in text_changed[:25]]

    # ── content-loss risk: LIVE verse has rows in a content table that the
    #    BUNDLED verse (same id) lacks — would be silently dropped by the copy.
    content_loss = []
    for vid in (v_removed + text_changed):
        lost_in = []
        for table, col in _RESEED_CONTENT_TABLES:
            try:
                l_n = lconn.execute('SELECT COUNT(*) FROM %s WHERE %s=?' % (table, col), (vid,)).fetchone()[0]
                if not l_n:
                    continue
                b_n = bconn.execute('SELECT COUNT(*) FROM %s WHERE %s=?' % (table, col), (vid,)).fetchone()[0]
                if b_n < l_n:
                    lost_in.append(table)
            except sqlite3.OperationalError:
                continue
        if lost_in:
            content_loss.append({'verse_id': vid, 'ref': _loc(lconn, vid), 'tables': lost_in})
        if len(content_loss) >= 30:
            break

    # ── audio manifest self-consistency vs the BUNDLED DB (what will actually
    #    be live right after the copy). readingFor() in app.js matches recordings
    #    by (book_id, sam_ch NUMBER) — not by sam_ch_id, which is just informational
    #    and can legitimately be stale across DB copies — so the real question is
    #    whether that (book, number) pair still names an actual chapter after the
    #    reseed; a manifest entry pointing at a number the bundled DB no longer has
    #    would silently misplay or vanish.
    audio_issues = []
    base = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'audio')
    b_sam_all = {r['id']: (r['book_id'], r['number']) for r in bconn.execute('SELECT id, book_id, number FROM sam_chapters')}
    b_sam_by_num = set(b_sam_all.values())
    try:
        rd = _json.load(open(os.path.join(base, 'readings', 'readings.json'), encoding='utf-8'))
        for bk in rd.get('books', []):
            bid = bk.get('book_id')
            for ch in bk.get('chapters', []):
                if (bid, ch.get('sam_ch_number')) not in b_sam_by_num:
                    audio_issues.append({'file': 'readings.json', 'issue': 'orphaned sam_ch_number',
                                          'detail': 'book %s, chapter %s (%s)'
                                                    % (bid, ch.get('sam_ch_number'), ch.get('name', ''))})
                if len(audio_issues) >= 20:
                    break
    except Exception:
        pass
    try:
        wj = _json.load(open(os.path.join(base, 'witnesses.json'), encoding='utf-8'))
        for it in wj.get('items', []):
            key = (it.get('book_id'), it.get('sam_ch_number'))
            if key not in b_sam_by_num:
                audio_issues.append({'file': 'witnesses.json', 'issue': 'orphaned sam_ch_number',
                                      'detail': 'book %s, chapter %s (%s)'
                                                % (it.get('book_id'), it.get('sam_ch_number'), it.get('reader', ''))})
            if len(audio_issues) >= 30:
                break
    except Exception:
        pass

    bconn.close(); lconn.close()
    return {
        'noop': False,
        'books': book_reports,
        'verses': {
            'added': len(v_added), 'removed': len(v_removed),
            'text_changed': len(text_changed), 'sam_ch_changed': len(sam_ch_changed),
            'text_sample': text_sample,
        },
        'content_loss': content_loss,
        'audio_issues': audio_issues,
        'state_hash': _reseed_state_hash(bundled, live),
    }


@app.route('/api/admin/reseed_diff')
def admin_reseed_diff():
    """The report the project owner requires be reviewed BEFORE any 'טען DB
    מהמאגר' reseed is allowed to run — see _reseed_diff_report()."""
    if not _valid_token(request.args.get('token')):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    try:
        return jsonify({'ok': True, 'report': _reseed_diff_report()})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/admin/reseed_db', methods=['POST'])
def admin_reseed_db():
    """Overwrite the live (persistent-disk) DB with the one bundled in the repo —
    used to apply a DB update pushed via git to the live site. DESTROYS unsynced
    online edits, so download_db first. No-op when not running off a separate disk.
    Requires the diff report's state_hash (from /api/admin/reseed_diff) to still
    match the CURRENT file state — if the live/bundled files changed since the
    report was fetched, the approval no longer corresponds to reality and this
    is rejected; the admin must re-fetch the report and approve again."""
    d = request.get_json(silent=True) or {}
    if not _valid_token(d.get('token')):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    if d.get('confirm') != 'REPLACE':
        return jsonify({'ok': False, 'error': "send confirm:'REPLACE'"}), 400
    bundled = getattr(db, '_BUNDLED_DB', None)
    if not bundled or db.DB_PATH == bundled or not os.path.exists(bundled):
        return jsonify({'ok': False, 'error': 'no separate disk / bundled DB'}), 400
    if os.path.exists(db.DB_PATH):
        current_hash = _reseed_state_hash(bundled, db.DB_PATH)
        if d.get('state_hash') != current_hash:
            return jsonify({'ok': False, 'error':
                'המצב השתנה מאז שהופק דוח ההשוואה — יש להפיק דוח מחדש ולאשר שוב.'}), 409
    try:
        _backup_db()
        shutil.copy2(bundled, db.DB_PATH)
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    return jsonify({'ok': True})


# ── putting back a DB from a copy on the administrator's own device ──────
# The sister of download_db. When the live site turns out to hold newer work than
# the copy on the machine — or when a sync has just written an older DB over it —
# the administrator opens the admin panel on that same phone or computer, picks
# the torah.db that was downloaded earlier, and puts it back. No deploy, no git.
#
# The file runs to some hundred and twenty megabytes, far too much for one
# request from a telephone, so it arrives in pieces: begin clears the staging
# file (and refuses if the disk has no room for it), chunk appends to it — each
# piece naming the offset it belongs at, so two tabs can never interleave — and
# commit opens the finished file and satisfies itself that it really is the Torah
# database before anything at all is replaced. Only then is the live DB backed up
# and the new one moved into its place, in one step.
_UPLOAD_SUFFIX = '.upload_admin'


def _upload_path():
    src = getattr(db, 'DB_PATH', None)
    return (src + _UPLOAD_SUFFIX) if src else None


def _db_counts(path):
    """The few numbers that tell a Torah database from any other file.

    Raises ValueError with a plain reason when the file is not one."""
    import sqlite3
    try:
        conn = sqlite3.connect(path)
    except Exception:
        raise ValueError('not a database file')
    try:
        try:
            names = {r[0] for r in conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table'")}
        except Exception:
            raise ValueError('not a database file')
        missing = {'books', 'chapters', 'verses', 'portions'} - names
        if missing:
            raise ValueError('missing tables: ' + ', '.join(sorted(missing)))
        if conn.execute('PRAGMA quick_check').fetchone()[0] != 'ok':
            raise ValueError('the file is damaged')
        out = {}
        for t in ('books', 'chapters', 'verses', 'portions', 'sam_chapters'):
            if t in names:
                out[t] = conn.execute('SELECT COUNT(*) FROM %s' % t).fetchone()[0]
        return out
    finally:
        conn.close()


def _counts_or_empty(path):
    try:
        return _db_counts(path)
    except Exception:
        return {}


@app.route('/api/admin/restore_db/begin', methods=['POST'])
def admin_restore_begin():
    d = request.get_json(silent=True) or {}
    if not _valid_token(d.get('token')):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    up = _upload_path()
    if not up:
        return jsonify({'ok': False, 'error': 'no db path'}), 400
    try:
        want = int(d.get('size') or 0)
    except (TypeError, ValueError):
        want = 0
    if want < 100000:
        return jsonify({'ok': False, 'error': 'file too small to be the database'}), 400
    live = os.path.getsize(db.DB_PATH) if os.path.exists(db.DB_PATH) else 0
    try:
        os.remove(up)
    except OSError:
        pass
    # room for the upload itself AND for the backup taken just before the swap
    free = shutil.disk_usage(os.path.dirname(up) or '.').free
    if free < want + live + 20 * 1024 * 1024:
        return jsonify({'ok': False, 'free': free, 'needed': want + live,
                        'error': 'not enough room on the disk — clean the old backups first'}), 400
    try:
        with open(up, 'wb'):
            pass
    except OSError as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    return jsonify({'ok': True, 'offset': 0, 'current': _counts_or_empty(db.DB_PATH)})


@app.route('/api/admin/restore_db/chunk', methods=['POST'])
def admin_restore_chunk():
    if not _valid_token(request.args.get('token')):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    up = _upload_path()
    if not up or not os.path.exists(up):
        return jsonify({'ok': False, 'error': 'call begin first'}), 400
    try:
        offset = int(request.args.get('offset', '-1'))
    except ValueError:
        offset = -1
    have = os.path.getsize(up)
    if offset != have:                       # a piece out of turn, or a second tab
        return jsonify({'ok': False, 'error': 'chunk out of order', 'offset': have}), 409
    data = request.get_data(cache=False)
    if not data:
        return jsonify({'ok': False, 'error': 'empty chunk'}), 400
    try:
        with open(up, 'ab') as f:
            f.write(data)
    except OSError as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    return jsonify({'ok': True, 'offset': os.path.getsize(up)})


@app.route('/api/admin/restore_db/commit', methods=['POST'])
def admin_restore_commit():
    d = request.get_json(silent=True) or {}
    if not _valid_token(d.get('token')):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    if d.get('confirm') != 'RESTORE':
        return jsonify({'ok': False, 'error': "send confirm:'RESTORE'"}), 400
    up = _upload_path()
    if not up or not os.path.exists(up):
        return jsonify({'ok': False, 'error': 'nothing was uploaded'}), 400
    size = os.path.getsize(up)
    try:
        want = int(d.get('size') or 0)
    except (TypeError, ValueError):
        want = 0
    if want and want != size:
        return jsonify({'ok': False,
                        'error': 'the upload is incomplete (%d of %d bytes)' % (size, want)}), 400
    try:
        after = _db_counts(up)
    except Exception as e:
        try:
            os.remove(up)
        except OSError:
            pass
        return jsonify({'ok': False, 'error': str(e)}), 400
    if after.get('books', 0) != 5 or after.get('verses', 0) < 5000:
        try:
            os.remove(up)
        except OSError:
            pass
        return jsonify({'ok': False, 'error': 'this is not the Torah database'}), 400
    before = _counts_or_empty(db.DB_PATH)
    try:
        _backup_db()
        try:
            os.replace(up, db.DB_PATH)       # one step: the site never sees half a file
        except OSError:
            # Windows will not rename over a file that something still has open;
            # write through the existing one instead. Not a single step, but the
            # backup taken a moment ago is the way back if it is interrupted.
            shutil.copyfile(up, db.DB_PATH)
            try:
                os.remove(up)
            except OSError:
                pass
        for sfx in ('-wal', '-shm'):         # stale journals of the file just replaced
            try:
                os.remove(db.DB_PATH + sfx)
            except OSError:
                pass
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    return jsonify({'ok': True, 'bytes': size, 'before': before, 'after': after})


@app.route('/api/admin/restore_db/abort', methods=['POST'])
def admin_restore_abort():
    d = request.get_json(silent=True) or {}
    if not _valid_token(d.get('token')):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    up = _upload_path()
    try:
        if up:
            os.remove(up)
    except OSError:
        pass
    return jsonify({'ok': True})


@app.route('/api/admin/edit', methods=['POST'])
def admin_edit():
    d = request.get_json(silent=True) or {}
    if not _valid_token(d.get('token')):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    table, col = d.get('table'), d.get('column')
    if table not in _EDITABLE or col not in _EDITABLE[table]:
        return jsonify({'ok': False, 'error': 'field not editable'}), 400
    try:
        vid = int(d.get('id'))
    except Exception:
        return jsonify({'ok': False, 'error': 'bad id'}), 400
    val = d.get('value', '')
    if not isinstance(val, str):
        return jsonify({'ok': False, 'error': 'bad value'}), 400
    conn = db.get_connection()
    try:
        conn.execute('UPDATE %s SET %s = ? WHERE id = ?' % (table, col), (val, vid))
        conn.commit()
    finally:
        conn.close()
    return jsonify({'ok': True})


# ── "חיבורים פרטיים": AI-assisted draft (admin only, own ANTHROPIC_API_KEY) ────
# read (list/get) requires only a valid token since the content is private;
# generate/save/delete/duplicate are POST-only admin actions like the rest of
# the admin_* endpoints above.
@app.route('/api/private_comp/list')
def private_comp_list():
    if not _valid_token(request.args.get('token')):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    return jsonify({'ok': True, 'items': db.list_private_compositions()})


@app.route('/api/private_comp/get')
def private_comp_get():
    if not _valid_token(request.args.get('token')):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    try:
        cid = int(request.args.get('id'))
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'error': 'bad id'}), 400
    comp = db.get_private_composition(cid)
    if not comp:
        return jsonify({'ok': False, 'error': 'not found'}), 404
    return jsonify({'ok': True, 'comp': comp})


@app.route('/api/admin/private_comp/generate', methods=['POST'])
def private_comp_generate():
    d = request.get_json(silent=True) or {}
    if not _valid_token(d.get('token')):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    prompt = (d.get('prompt') or '').strip()
    if not prompt:
        return jsonify({'ok': False, 'error': 'empty prompt'}), 400
    import claude_composer
    try:
        text = claude_composer.generate_composition_draft(prompt)
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    return jsonify({'ok': True, 'text': text})


@app.route('/api/admin/private_comp/save', methods=['POST'])
def private_comp_save():
    d = request.get_json(silent=True) or {}
    if not _valid_token(d.get('token')):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    title = (d.get('title') or '').strip() or 'חיבור ללא כותרת'
    prompt = d.get('prompt') or ''
    text = d.get('text') or ''
    lines = text.split('\n')
    cid = db.create_private_composition(title, prompt, lines)
    return jsonify({'ok': True, 'id': cid})


@app.route('/api/admin/private_comp/delete', methods=['POST'])
def private_comp_delete():
    d = request.get_json(silent=True) or {}
    if not _valid_token(d.get('token')):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    try:
        cid = int(d.get('id'))
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'error': 'bad id'}), 400
    db.delete_private_composition(cid)
    return jsonify({'ok': True})


@app.route('/api/admin/private_comp/duplicate', methods=['POST'])
def private_comp_duplicate():
    d = request.get_json(silent=True) or {}
    if not _valid_token(d.get('token')):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    try:
        cid = int(d.get('id'))
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'error': 'bad id'}), 400
    new_id = db.duplicate_private_composition(cid)
    if new_id is None:
        return jsonify({'ok': False, 'error': 'not found'}), 404
    return jsonify({'ok': True, 'id': new_id})


# ── admin chapter restructuring (merge / split) — local only, gated + backed up ─
import shutil
from datetime import datetime as _dt


_BACKUP_KEEP = 3  # rotate: an unbounded number of ~90MB copies fills the 1GB disk


def _backup_glob(src):
    import glob
    return sorted(glob.glob('%s.bak_admin_*' % src))


def _backup_db():
    _mem_clear()          # the text is about to change; the small cache must go
    src = getattr(db, 'DB_PATH', None)
    if src and os.path.exists(src):
        shutil.copy2(src, '%s.bak_admin_%s' % (src, _dt.now().strftime('%Y%m%d_%H%M%S')))
        stale = _backup_glob(src)[:-_BACKUP_KEEP]
        for f in stale:
            try:
                os.remove(f)
            except OSError:
                pass


@app.route('/api/admin/disk_usage')
def admin_disk_usage():
    if not _valid_token(request.args.get('token')):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    src = getattr(db, 'DB_PATH', None)
    backups = _backup_glob(src) if src else []
    disk_dir = os.path.dirname(src) if src else '.'
    total, used, free = shutil.disk_usage(disk_dir)
    return jsonify({'ok': True,
                     'db_bytes': os.path.getsize(src) if src and os.path.exists(src) else 0,
                     'backups': [{'name': os.path.basename(f), 'bytes': os.path.getsize(f)} for f in backups],
                     'backups_bytes': sum(os.path.getsize(f) for f in backups),
                     'disk_total': total, 'disk_used': used, 'disk_free': free})


@app.route('/api/admin/clean_backups', methods=['POST'])
def admin_clean_backups():
    """Delete old .bak_admin_* copies to reclaim disk space. Keeps the most recent
    _BACKUP_KEEP by default; pass keep:0 to wipe all of them (git history is the
    real backup — these are just a same-disk safety net for the last few edits)."""
    d = request.get_json(silent=True) or {}
    if not _valid_token(d.get('token')):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    src = getattr(db, 'DB_PATH', None)
    if not src:
        return jsonify({'ok': False, 'error': 'no db path'}), 400
    try:
        keep = max(0, int(d.get('keep', _BACKUP_KEEP)))
    except (TypeError, ValueError):
        keep = _BACKUP_KEEP
    stale = _backup_glob(src)[:-keep] if keep else _backup_glob(src)
    freed = 0
    for f in stale:
        try:
            freed += os.path.getsize(f)
            os.remove(f)
        except OSError:
            pass
    return jsonify({'ok': True, 'deleted': len(stale), 'freed_bytes': freed})


def _portion_spans(conn, book_id):
    """For every portion of the book, the first/last verse_id it currently covers
    (by standard chapter:verse), so its boundaries can be recomputed after a
    re-chaptering. Returns {portion_id: (first_vid, last_vid, end_was_sentinel)}."""
    out = {}
    for p in conn.execute('SELECT id,start_ch,start_v,end_ch,end_v FROM portions WHERE book_id=?',
                          (book_id,)).fetchall():
        rows = conn.execute(
            """SELECT v.id FROM verses v JOIN chapters c ON c.id=v.chapter_id
               WHERE c.book_id=?
                 AND (c.number>? OR (c.number=? AND CAST(v.number AS INTEGER)>=?))
                 AND (c.number<? OR (c.number=? AND CAST(v.number AS INTEGER)<=?))
               ORDER BY c.number, CAST(v.number AS INTEGER), v.id""",
            (book_id, p['start_ch'], p['start_ch'], p['start_v'],
             p['end_ch'], p['end_ch'], p['end_v'])).fetchall()
        if rows:
            out[p['id']] = (rows[0]['id'], rows[-1]['id'], p['end_v'] >= 9999)
    return out


def _pos(conn, vid):
    r = conn.execute("""SELECT c.number ch, CAST(v.number AS INTEGER) vn
                        FROM verses v JOIN chapters c ON c.id=v.chapter_id WHERE v.id=?""", (vid,)).fetchone()
    return (r['ch'], r['vn']) if r else (None, None)


def _portion_id_for_sam_chapter(conn, sam_id):
    """Which Samaritan-mode portion a sam_chapter's first verse falls in (by the
    same chapter:verse range logic as get_sam_chapters_in_portion), for the
    per-portion canon guard."""
    r = conn.execute("""
        SELECT p.id FROM sam_chapters sc
        JOIN (SELECT sam_ch_id, MIN(id) AS first_v_id FROM verses GROUP BY sam_ch_id) fv
          ON fv.sam_ch_id = sc.id
        JOIN verses v ON v.id = fv.first_v_id
        JOIN chapters c ON c.id = v.chapter_id
        JOIN portions p ON p.book_id = c.book_id AND p.mode = 'samaritan'
        WHERE sc.id = ?
          AND ( sc.portion_id = p.id
             OR (sc.portion_id IS NULL
                 AND (c.number > p.start_ch OR (c.number = p.start_ch AND CAST(v.number AS INTEGER) >= p.start_v))
                 AND (c.number < p.end_ch   OR (c.number = p.end_ch   AND CAST(v.number AS INTEGER) <= p.end_v))) )
        """, (sam_id,)).fetchone()
    return r['id'] if r else None


def _portion_chapter_counts(conn, book_id):
    """Every Samaritan-mode portion of a book with how many Samaritan chapters it
    holds and the number of its last one. A chapter belongs to the portion its
    FIRST verse falls in — one that opens in a portion and runs past its end is
    still that portion's chapter, and counting it by where it ends would file it
    under the next one."""
    por = [dict(r) for r in conn.execute(
        "SELECT id, name, order_n, start_ch, start_v, end_ch, end_v FROM portions "
        "WHERE book_id=? AND mode='samaritan' ORDER BY order_n", (book_id,))]
    out = {p['id']: {'portion_id': p['id'], 'name': p['name'], 'order_n': p['order_n'],
                     'count': 0, 'last_number': None} for p in por}
    rows = conn.execute("""
        SELECT sc.id, sc.number, sc.portion_id AS pin, c.number ch,
               CAST(v.number AS INTEGER) vn
        FROM sam_chapters sc
        JOIN (SELECT sam_ch_id, MIN(id) fid FROM verses GROUP BY sam_ch_id) f ON f.sam_ch_id = sc.id
        JOIN verses v ON v.id = f.fid
        JOIN chapters c ON c.id = v.chapter_id
        WHERE sc.book_id = ? ORDER BY sc.number""", (book_id,)).fetchall()
    for r in rows:
        pid = r['pin']
        if pid not in out:
            pid = None
            k = r['ch'] * 10000 + (r['vn'] or 0)
            for p in por:
                if (p['start_ch'] * 10000 + p['start_v']) <= k <= (p['end_ch'] * 10000 + min(p['end_v'], 9999)):
                    pid = p['id']
                    break
        if pid in out:
            out[pid]['count'] += 1
            out[pid]['last_number'] = r['number']
    return {k: v for k, v in out.items() if v['count']}


# The canon WARNS, it does not block (the project owner's instruction, 2026-08-14):
# a split or merge that moves a count off its canon is still the owner's to make,
# so it is answered with a confirmation request carrying the exact numbers, and it
# goes through once the agreed phrase is typed. Settable per deployment; the
# default is the phrase the owner chose.
CANON_PHRASE = os.environ.get('CANON_OVERRIDE_PHRASE', 'קאנון מאושר')


def _canon_phrase_ok(given):
    """Whitespace-insensitive, so a stray double space or a trailing blank from a
    phone keyboard does not read as the wrong phrase."""
    norm = lambda s: ' '.join(str(s or '').split())
    return bool(norm(given)) and norm(given) == norm(CANON_PHRASE)


def _canon_check(conn, book_id, sam_id, delta):
    """How far a split (delta=+1) or merge (delta=-1) moves the book's and the
    affected portion's Samaritan chapter count from its engraved canon.
    Returns None when both stay exactly on canon, otherwise a report the caller
    turns into a confirmation request."""
    out = {'book': None, 'portion': None}
    canon = conn.execute('SELECT canonical_count FROM canon_chapter_counts WHERE book_id=?',
                         (book_id,)).fetchone()
    if canon:
        n_now = conn.execute('SELECT COUNT(*) FROM sam_chapters WHERE book_id=?', (book_id,)).fetchone()[0]
        if n_now + delta != canon['canonical_count']:
            out['book'] = {'now': n_now, 'after': n_now + delta, 'canon': canon['canonical_count']}
    pid = _portion_id_for_sam_chapter(conn, sam_id)
    if pid:
        pc = conn.execute('SELECT canonical_count, portion_name FROM canon_portion_counts WHERE portion_id=?',
                          (pid,)).fetchone()
        if pc:
            p_now = conn.execute('SELECT COUNT(DISTINCT sc.id) FROM sam_chapters sc '
                                 'JOIN (SELECT sam_ch_id, MIN(id) fid FROM verses GROUP BY sam_ch_id) fv '
                                 'ON fv.sam_ch_id=sc.id JOIN verses v ON v.id=fv.fid '
                                 'JOIN chapters c ON c.id=v.chapter_id JOIN portions p ON p.id=? '
                                 'WHERE c.book_id=p.book_id AND (sc.portion_id=p.id OR '
                                 '(sc.portion_id IS NULL AND '
                                 '(c.number>p.start_ch OR (c.number=p.start_ch AND CAST(v.number AS INTEGER)>=p.start_v)) AND '
                                 '(c.number<p.end_ch OR (c.number=p.end_ch AND CAST(v.number AS INTEGER)<=p.end_v))))',
                                 (pid,)).fetchone()[0]
            if p_now + delta != pc['canonical_count']:
                out['portion'] = {'name': pc['portion_name'], 'now': p_now,
                                  'after': p_now + delta, 'canon': pc['canonical_count']}
    if not out['book'] and not out['portion']:
        return None
    # whether each count walks toward its canon or away from it — the reviewer's
    # first question, and the two can point in opposite directions at once
    for k in ('book', 'portion'):
        if out[k]:
            x = out[k]
            x['closer'] = abs(x['after'] - x['canon']) < abs(x['now'] - x['canon'])
    parts = []
    if out['book']:
        b = out['book']
        parts.append('בספר: %d ← %d (הקאנון %d, %s)'
                     % (b['now'], b['after'], b['canon'], 'מתקרב' if b['closer'] else 'מתרחק'))
    if out['portion']:
        p = out['portion']
        parts.append('בפרשת "%s": %d ← %d (הקאנון %d, %s)'
                     % (p['name'], p['now'], p['after'], p['canon'], 'מתקרב' if p['closer'] else 'מתרחק'))
    out['message'] = 'הפעולה מוציאה את מניין הפרקים השומרוניים מן הקאנון — ' + ' · '.join(parts)
    return out


def _canon_gate(conn, book_id, sam_id, delta, given_phrase):
    """None to proceed, or the (response, status) that asks for the phrase."""
    dev = _canon_check(conn, book_id, sam_id, delta)
    if not dev or _canon_phrase_ok(given_phrase):
        return None
    return jsonify({'ok': False, 'canon_confirm': True,
                    'error': dev['message'], 'details': dev}), 409


def _fix_portions(conn, spans):
    for pid, (fv, lv, sentinel) in spans.items():
        sch, svn = _pos(conn, fv)
        ech, evn = _pos(conn, lv)
        if sch is None or ech is None:
            continue
        conn.execute('UPDATE portions SET start_ch=?,start_v=?,end_ch=?,end_v=? WHERE id=?',
                     (sch, svn, ech, 9999 if sentinel else evn, pid))


def _fix_root_index(conn, book_id):
    for r in conn.execute("""SELECT v.id vid, c.number ch, v.number vn FROM verses v
                             JOIN chapters c ON c.id=v.chapter_id WHERE c.book_id=?""", (book_id,)).fetchall():
        conn.execute('UPDATE root_index SET chapter=?, verse=? WHERE verse_id=?', (r['ch'], r['vn'], r['vid']))


def _shift_reading_manifests(book_id, pivot, delta):
    """Keep the reading manifests aligned after a Samaritan-chapter split/merge.
    readings.json + witnesses.json key recordings by (book_id, sam chapter NUMBER);
    a split (delta=+1) / merge (delta=-1) renumbers every chapter after `pivot`,
    so entries past the pivot must shift too or every later chapter plays the
    previous/next chapter's audio. On a merge the swallowed chapter's own entry
    (number pivot+1) is dropped from readings.json; in witnesses.json its segments
    are appended to the surviving chapter's entry when the reader matches."""
    base = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'audio')

    def _rewrite(path, fn):
        if not os.path.exists(path):
            return
        try:
            with open(path, encoding='utf-8') as f:
                data = _json.load(f)
            fn(data)
            tmp = path + '.tmp'
            with open(tmp, 'w', encoding='utf-8') as f:
                _json.dump(data, f, ensure_ascii=False)
            os.replace(tmp, path)
        except Exception:
            pass   # a manifest problem must never fail the structural edit itself

    def _readings(data):
        books = data.get('books') if isinstance(data, dict) else None
        if not isinstance(books, list):
            return
        for b in books:
            if b.get('book_id') != book_id or not isinstance(b.get('chapters'), list):
                continue
            if delta < 0:   # merge: chapter pivot+1 no longer exists
                b['chapters'] = [c for c in b['chapters'] if c.get('sam_ch_number') != pivot + 1]
            for c in b['chapters']:
                n = c.get('sam_ch_number')
                if isinstance(n, int) and n > pivot:
                    c['sam_ch_number'] = n + delta

    def _witnesses(data):
        items = data.get('items') if isinstance(data, dict) else None
        if not isinstance(items, list):
            return
        if delta < 0:
            keep = []
            absorbed = [it for it in items
                        if it.get('book_id') == book_id and it.get('sam_ch_number') == pivot + 1]
            for it in items:
                if it.get('book_id') == book_id and it.get('sam_ch_number') == pivot + 1:
                    continue
                if (it.get('book_id') == book_id and it.get('sam_ch_number') == pivot
                        and it.get('segs')):
                    ext = next((a for a in absorbed
                                if a.get('reader') == it.get('reader') and a.get('segs')), None)
                    if ext:
                        it['segs'] = list(it['segs']) + list(ext['segs'])
                        it['duration'] = (it.get('duration') or 0) + (ext.get('duration') or 0)
                keep.append(it)
            items[:] = keep
        for it in items:
            n = it.get('sam_ch_number')
            if it.get('book_id') == book_id and isinstance(n, int) and n > pivot:
                it['sam_ch_number'] = n + delta

    _rewrite(os.path.join(base, 'readings', 'readings.json'), _readings)
    _rewrite(os.path.join(base, 'witnesses.json'), _witnesses)


@app.route('/api/admin/merge_next', methods=['POST'])
def admin_merge_next():
    d = request.get_json(silent=True) or {}
    if not _valid_token(d.get('token')):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    try:
        chapter_id = int(d.get('chapter_id'))
    except Exception:
        return jsonify({'ok': False, 'error': 'bad chapter'}), 400
    _backup_db()
    conn = db.get_connection()
    try:
        cur = conn.execute('SELECT id,book_id,number FROM chapters WHERE id=?', (chapter_id,)).fetchone()
        if not cur:
            return jsonify({'ok': False, 'error': 'chapter not found'}), 404
        nxt = conn.execute('SELECT id,number FROM chapters WHERE book_id=? AND number=?',
                           (cur['book_id'], cur['number'] + 1)).fetchone()
        if not nxt:
            return jsonify({'ok': False, 'error': 'אין פרק הבא לאיחוד'}), 400
        book_id, N = cur['book_id'], cur['number']
        spans = _portion_spans(conn, book_id)
        k = conn.execute('SELECT COALESCE(MAX(CAST(number AS INTEGER)),0) FROM verses WHERE chapter_id=?',
                        (cur['id'],)).fetchone()[0]
        moved = conn.execute('SELECT id FROM verses WHERE chapter_id=? ORDER BY CAST(number AS INTEGER), id',
                            (nxt['id'],)).fetchall()
        for i, r in enumerate(moved, 1):
            conn.execute('UPDATE verses SET chapter_id=?, number=? WHERE id=?', (cur['id'], str(k + i), r['id']))
        conn.execute('DELETE FROM chapters WHERE id=?', (nxt['id'],))
        conn.execute('UPDATE chapters SET number=number-1 WHERE book_id=? AND number>?', (book_id, N + 1))
        _fix_portions(conn, spans)
        _fix_root_index(conn, book_id)
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()
    return jsonify({'ok': True})


@app.route('/api/admin/split', methods=['POST'])
def admin_split():
    d = request.get_json(silent=True) or {}
    if not _valid_token(d.get('token')):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    try:
        chapter_id = int(d.get('chapter_id')); after_vid = int(d.get('after_verse_id'))
    except Exception:
        return jsonify({'ok': False, 'error': 'bad params'}), 400
    _backup_db()
    conn = db.get_connection()
    try:
        cur = conn.execute('SELECT id,book_id,number FROM chapters WHERE id=?', (chapter_id,)).fetchone()
        if not cur:
            return jsonify({'ok': False, 'error': 'chapter not found'}), 404
        book_id, N = cur['book_id'], cur['number']
        ids = [r['id'] for r in conn.execute(
            'SELECT id FROM verses WHERE chapter_id=? ORDER BY CAST(number AS INTEGER), id', (cur['id'],)).fetchall()]
        if after_vid not in ids:
            return jsonify({'ok': False, 'error': 'הפסוק אינו בפרק זה'}), 400
        pos = ids.index(after_vid)
        if pos >= len(ids) - 1:
            return jsonify({'ok': False, 'error': 'לא ניתן לפצל אחרי הפסוק האחרון'}), 400
        moved = ids[pos + 1:]
        spans = _portion_spans(conn, book_id)
        conn.execute('UPDATE chapters SET number=number+1 WHERE book_id=? AND number>?', (book_id, N))
        c2 = conn.cursor()
        c2.execute('INSERT INTO chapters (book_id, number) VALUES (?,?)', (book_id, N + 1))
        new_id = c2.lastrowid
        for i, vid in enumerate(moved, 1):
            conn.execute('UPDATE verses SET chapter_id=?, number=? WHERE id=?', (new_id, str(i), vid))
        _fix_portions(conn, spans)
        _fix_root_index(conn, book_id)
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()
    return jsonify({'ok': True})


# Samaritan-division chapter merge/split: only sam_ch_id + sam_chapters change, so
# the Jewish division, parashot (standard chapter:verse) and root_index are untouched.
@app.route('/api/admin/merge_next_sam', methods=['POST'])
def admin_merge_next_sam():
    d = request.get_json(silent=True) or {}
    if not _valid_token(d.get('token')):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    try:
        sam_id = int(d.get('chapter_id'))
    except Exception:
        return jsonify({'ok': False, 'error': 'bad chapter'}), 400
    _backup_db()
    conn = db.get_connection()
    try:
        cur = conn.execute('SELECT id,book_id,number FROM sam_chapters WHERE id=?', (sam_id,)).fetchone()
        if not cur:
            return jsonify({'ok': False, 'error': 'chapter not found'}), 404
        nxt = conn.execute('SELECT id FROM sam_chapters WHERE book_id=? AND number=?',
                           (cur['book_id'], cur['number'] + 1)).fetchone()
        if not nxt:
            return jsonify({'ok': False, 'error': 'אין פרק הבא לאיחוד'}), 400
        gate = _canon_gate(conn, cur['book_id'], cur['id'], -1, d.get('canon_phrase'))
        if gate:
            return gate
        conn.execute('UPDATE verses SET sam_ch_id=? WHERE sam_ch_id=?', (cur['id'], nxt['id']))
        conn.execute('DELETE FROM sam_chapters WHERE id=?', (nxt['id'],))
        conn.execute('UPDATE sam_chapters SET number=number-1 WHERE book_id=? AND number>?',
                     (cur['book_id'], cur['number'] + 1))
        conn.commit()
        _shift_reading_manifests(cur['book_id'], cur['number'], -1)
    except Exception as e:
        conn.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()
    return jsonify({'ok': True})


@app.route('/api/admin/split_sam', methods=['POST'])
def admin_split_sam():
    d = request.get_json(silent=True) or {}
    if not _valid_token(d.get('token')):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    try:
        sam_id = int(d.get('chapter_id')); after_vid = int(d.get('after_verse_id'))
    except Exception:
        return jsonify({'ok': False, 'error': 'bad params'}), 400
    _backup_db()
    conn = db.get_connection()
    try:
        cur = conn.execute('SELECT id,book_id,number,portion_id FROM sam_chapters WHERE id=?', (sam_id,)).fetchone()
        if not cur:
            return jsonify({'ok': False, 'error': 'chapter not found'}), 404
        gate = _canon_gate(conn, cur['book_id'], cur['id'], +1, d.get('canon_phrase'))
        if gate:
            return gate
        ids = [r['id'] for r in conn.execute(
            """SELECT v.id FROM verses v JOIN chapters c ON c.id=v.chapter_id
               WHERE v.sam_ch_id=? ORDER BY c.number, CAST(v.number AS INTEGER), v.id""", (cur['id'],)).fetchall()]
        if after_vid not in ids:
            return jsonify({'ok': False, 'error': 'הפסוק אינו בפרק זה'}), 400
        pos = ids.index(after_vid)
        if pos >= len(ids) - 1:
            return jsonify({'ok': False, 'error': 'לא ניתן לפצל אחרי הפסוק האחרון'}), 400
        moved = ids[pos + 1:]
        conn.execute('UPDATE sam_chapters SET number=number+1 WHERE book_id=? AND number>?',
                     (cur['book_id'], cur['number']))
        c2 = conn.cursor()
        # the new chapter inherits a manual portion pin, if the split one had any
        c2.execute('INSERT INTO sam_chapters (book_id, number, portion_id) VALUES (?,?,?)',
                   (cur['book_id'], cur['number'] + 1, cur['portion_id']))
        new_id = c2.lastrowid
        conn.executemany('UPDATE verses SET sam_ch_id=? WHERE id=?', [(new_id, vid) for vid in moved])
        conn.commit()
        _shift_reading_manifests(cur['book_id'], cur['number'], +1)
    except Exception as e:
        conn.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()
    return jsonify({'ok': True})


# Split one verse into two in the Samaritan division. The new verse keeps the
# integer base of the original and gets the next free maqaf sub-number
# (10 -> 10-1, 10-1 -> 10-2, 11 -> 11-1). Because get_verses() filters
# typeof(number)='integer', the maqaf verse shows ONLY in the Samaritan division
# (same Jewish chapter and same Samaritan chapter as the original).
@app.route('/api/admin/split_verse', methods=['POST'])
def admin_split_verse():
    d = request.get_json(silent=True) or {}
    if not _valid_token(d.get('token')):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    try:
        verse_id = int(d.get('verse_id'))
    except Exception:
        return jsonify({'ok': False, 'error': 'bad params'}), 400
    text1, text2 = d.get('text1'), d.get('text2')
    if not isinstance(text1, str) or not isinstance(text2, str) or not text1.strip() or not text2.strip():
        return jsonify({'ok': False, 'error': 'שני החלקים חייבים להכיל טקסט'}), 400
    _backup_db()
    conn = db.get_connection()
    try:
        v = conn.execute('SELECT id, chapter_id, number, sam_ch_id FROM verses WHERE id=?', (verse_id,)).fetchone()
        if not v:
            return jsonify({'ok': False, 'error': 'verse not found'}), 404
        base = str(v['number']).split('-')[0]
        if not base.isdigit():
            return jsonify({'ok': False, 'error': 'מספר פסוק לא תקין'}), 400
        # highest existing maqaf sub-number for this base in the same Jewish chapter
        mx = 0
        for r in conn.execute('SELECT number FROM verses WHERE chapter_id=?', (v['chapter_id'],)):
            s = str(r['number'])
            if s.startswith(base + '-'):
                tail = s[len(base) + 1:]
                if tail.isdigit():
                    mx = max(mx, int(tail))
        new_number = '%s-%d' % (base, mx + 1)
        conn.execute('UPDATE verses SET text=? WHERE id=?', (text1.strip(), verse_id))
        conn.execute('INSERT INTO verses (chapter_id, number, text, sam_ch_id) VALUES (?,?,?,?)',
                     (v['chapter_id'], new_number, text2.strip(), v['sam_ch_id']))
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()
    return jsonify({'ok': True, 'new_number': new_number})


# ── merging two verses into one ────────────────────────────────────────────────
# The inverse of admin_split_verse, and the answer to a scan that carries as one
# verse what we carry as two. The survivor is always the EARLIER of the pair and
# it keeps its own number — merge upwards and the lower verse joins the one above
# under that number; merge downwards and the pair is filed under this verse's
# number. The later row is then deleted, so everything hanging off it (word
# glosses, dictionary rows, source links, transliteration, the root index) is
# first carried over to the survivor rather than orphaned.
_MERGE_PROSE_COLS = ('text', 'masoretic_text', 'lxx_text', 'onkelos_text', 'qumran_text',
                     'sam_aramaic', 'english', 'site_english', 'simple_hebrew', 'sam_hebrew',
                     'arabic_trans', 'interpretation', 'interpretation_ar', 'old_text',
                     'rashi', 'ramban', 'cassuto', 'baal_haturim')


def _carry_verse_refs(conn, keep_id, drop_id):
    """Move every row that points at the swallowed verse onto the survivor.
    Word-level tables are keyed UNIQUE(verse_id, pos) — the merged verse's words
    are the two lists one after the other, so those positions are shifted past
    the survivor's own rather than colliding with them."""
    for tbl in ('word_gloss', 'word_jewish', 'word_samaritan', 'word_align'):
        base = conn.execute('SELECT COALESCE(MAX(pos), -1) FROM %s WHERE verse_id=?' % tbl,
                            (keep_id,)).fetchone()[0]
        conn.execute('UPDATE %s SET verse_id=?, pos=pos+? WHERE verse_id=?' % tbl,
                     (keep_id, base + 1, drop_id))
    # one row per verse: the two readings become one, in order
    for tbl in ('verse_translit', 'verse_translit_fix'):
        a = conn.execute('SELECT text FROM %s WHERE verse_id=?' % tbl, (keep_id,)).fetchone()
        b = conn.execute('SELECT text FROM %s WHERE verse_id=?' % tbl, (drop_id,)).fetchone()
        if b and (b['text'] or '').strip():
            if a:
                conn.execute('UPDATE %s SET text=? WHERE verse_id=?' % tbl,
                             (' '.join(x for x in ((a['text'] or '').strip(), b['text'].strip()) if x), keep_id))
                conn.execute('DELETE FROM %s WHERE verse_id=?' % tbl, (drop_id,))
            else:
                conn.execute('UPDATE %s SET verse_id=? WHERE verse_id=?' % tbl, (keep_id, drop_id))
        else:
            conn.execute('DELETE FROM %s WHERE verse_id=?' % tbl, (drop_id,))
    # UNIQUE(verse_id, root_norm): a root both verses share is already recorded
    conn.execute('UPDATE OR IGNORE dict_torah_sense SET verse_id=? WHERE verse_id=?', (keep_id, drop_id))
    conn.execute('DELETE FROM dict_torah_sense WHERE verse_id=?', (drop_id,))
    # the root index also carries the reference in words
    pos = conn.execute('''SELECT c.number ch, v.number vn FROM verses v
                          JOIN chapters c ON c.id=v.chapter_id WHERE v.id=?''', (keep_id,)).fetchone()
    if pos:
        conn.execute('UPDATE root_index SET verse_id=?, chapter=?, verse=? WHERE verse_id=?',
                     (keep_id, pos['ch'], pos['vn'], drop_id))
    # everything else that names a verse, discovered from the schema so a table
    # added later is carried too
    done = {'word_gloss', 'word_jewish', 'word_samaritan', 'word_align', 'verse_translit',
            'verse_translit_fix', 'dict_torah_sense', 'root_index', 'verses'}
    for (tbl,) in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall():
        if tbl in done:
            continue
        if any(c[1] == 'verse_id' for c in conn.execute('PRAGMA table_info(%s)' % tbl)):
            conn.execute('UPDATE OR IGNORE %s SET verse_id=? WHERE verse_id=?' % tbl, (keep_id, drop_id))
            conn.execute('DELETE FROM %s WHERE verse_id=?' % tbl, (drop_id,))


@app.route('/api/admin/merge_verse', methods=['POST'])
def admin_merge_verse():
    d = request.get_json(silent=True) or {}
    if not _valid_token(d.get('token')):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    try:
        verse_id = int(d.get('verse_id')); other_id = int(d.get('other_verse_id'))
    except Exception:
        return jsonify({'ok': False, 'error': 'bad params'}), 400
    direction = d.get('direction')          # 'prev' | 'next', for the message only
    _backup_db()
    conn = db.get_connection()
    try:
        a = conn.execute('SELECT * FROM verses WHERE id=?', (verse_id,)).fetchone()
        b = conn.execute('SELECT * FROM verses WHERE id=?', (other_id,)).fetchone()
        if not a or not b:
            return jsonify({'ok': False, 'error': 'verse not found'}), 404
        if a['chapter_id'] != b['chapter_id']:
            return jsonify({'ok': False, 'error': 'הפסוקים אינם באותו פרק'}), 400
        if a['sam_ch_id'] != b['sam_ch_id']:
            return jsonify({'ok': False, 'error': 'הפסוקים אינם באותו פרק שומרוני'}), 400
        keep, drop = (a, b) if direction == 'next' else (b, a)
        cols = [c[1] for c in conn.execute('PRAGMA table_info(verses)')]
        sets, vals = [], []
        for col in cols:
            if col in ('id', 'chapter_id', 'number', 'sam_ch_id'):
                continue
            if col in _MERGE_PROSE_COLS:
                parts = [str(keep[col] or '').strip(), str(drop[col] or '').strip()]
                sets.append('%s=?' % col); vals.append(' '.join(p for p in parts if p))
            elif not str(keep[col] or '').strip() and str(drop[col] or '').strip():
                sets.append('%s=?' % col); vals.append(drop[col])
        if sets:
            conn.execute('UPDATE verses SET %s WHERE id=?' % ', '.join(sets), vals + [keep['id']])
        _carry_verse_refs(conn, keep['id'], drop['id'])
        conn.execute('DELETE FROM verses WHERE id=?', (drop['id'],))
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()
    return jsonify({'ok': True, 'number': keep['number'], 'dropped': drop['number']})


# ── engraving the canon from the app itself ───────────────────────────────────
@app.route('/api/admin/set_canon', methods=['POST'])
def admin_set_canon():
    """Stamp the count as it stands: the portion's, from its last Samaritan
    chapter — and from the last portion of a book, every portion of that book and
    the book's own total. What is stamped is signed and dated, and from then on
    the canon gate asks for the phrase before anything moves it."""
    d = request.get_json(silent=True) or {}
    if not _valid_token(d.get('token')):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    try:
        sam_id = int(d.get('sam_ch_id'))
    except Exception:
        return jsonify({'ok': False, 'error': 'bad params'}), 400
    _backup_db()
    conn = db.get_connection()
    try:
        sc = conn.execute('SELECT id, book_id, number FROM sam_chapters WHERE id=?', (sam_id,)).fetchone()
        if not sc:
            return jsonify({'ok': False, 'error': 'chapter not found'}), 404
        pid = _portion_id_for_sam_chapter(conn, sam_id)
        if not pid:
            return jsonify({'ok': False, 'error': 'לא נמצאה פרשה לפרק זה'}), 400
        counts = _portion_chapter_counts(conn, sc['book_id'])
        if not counts.get(pid) or counts[pid]['last_number'] != sc['number']:
            return jsonify({'ok': False,
                            'error': 'אפשר לקבוע קאנון רק מן הפרק השומרוני האחרון בפרשה'}), 400
        stamp = time.strftime('%d/%m/%Y')
        note = ('נחתם ע"י בעל הפרויקט ב-%s: %%d פרקים בחלוקה השומרונית. '
                'החתימה נעולה — כל שינוי במניין מחייב את מילת האישור.') % stamp
        last_portion = max(counts.values(), key=lambda x: x['order_n'])['portion_id'] == pid
        stamped = []
        targets = list(counts.items()) if last_portion else [(pid, counts[pid])]
        for p_id, info in targets:
            conn.execute('INSERT OR REPLACE INTO canon_portion_counts '
                         '(portion_id, book_id, portion_name, canonical_count, note) VALUES (?,?,?,?,?)',
                         (p_id, sc['book_id'], info['name'], info['count'], note % info['count']))
            stamped.append({'portion': info['name'], 'count': info['count']})
        book_total = None
        if last_portion:
            book_total = conn.execute('SELECT COUNT(*) FROM sam_chapters WHERE book_id=?',
                                      (sc['book_id'],)).fetchone()[0]
            bname = conn.execute('SELECT name FROM books WHERE id=?', (sc['book_id'],)).fetchone()['name']
            conn.execute('INSERT OR REPLACE INTO canon_chapter_counts '
                         '(book_id, book_name, canonical_count, note) VALUES (?,?,?,?)',
                         (sc['book_id'], bname, book_total, note % book_total))
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()
    return jsonify({'ok': True, 'stamped': stamped, 'book_total': book_total,
                    'whole_book': bool(book_total)})


# ── comparison-text "reflow" ops (split / merge across ADJACENT rows) ──────────
# The Masoretic/LXX/Onkelos/Qumran/Aramaic columns have no independent verse
# numbering of their own — each is just a string field living on the same
# Samaritan verse row. So unlike admin_split_verse (which creates a genuine new
# Samaritan maqaf verse), "splitting"/"merging" these can only move a chunk of
# text between THIS row and the adjacent one — no new row, no renumbering.
# next_verse_id is supplied by the client from its already-correctly-ordered
# S.verses array, rather than re-derived here from the mixed-format `number`
# column (which is not safe to ORDER BY numerically as-is).
_CMP_REFLOW_COLS = {'masoretic_text', 'lxx_text', 'onkelos_text', 'qumran_text', 'sam_aramaic'}


def _cmp_reflow_rows(conn, column, verse_id, next_id):
    v = conn.execute('SELECT chapter_id, %s AS val FROM verses WHERE id=?' % column, (verse_id,)).fetchone()
    nv = conn.execute('SELECT chapter_id, %s AS val FROM verses WHERE id=?' % column, (next_id,)).fetchone()
    if not v or not nv:
        return None, None, 'verse not found'
    if v['chapter_id'] != nv['chapter_id']:
        return None, None, 'הפסוקים אינם באותו פרק'
    return v, nv, None


@app.route('/api/admin/cmp_split_next', methods=['POST'])
def admin_cmp_split_next():
    d = request.get_json(silent=True) or {}
    if not _valid_token(d.get('token')):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    column = d.get('column')
    if column not in _CMP_REFLOW_COLS:
        return jsonify({'ok': False, 'error': 'field not splittable'}), 400
    try:
        verse_id = int(d.get('verse_id'))
        next_id = int(d.get('next_verse_id'))
    except Exception:
        return jsonify({'ok': False, 'error': 'bad params'}), 400
    text1, text2 = d.get('text1'), d.get('text2')
    if not isinstance(text1, str) or not isinstance(text2, str) or not text1.strip() or not text2.strip():
        return jsonify({'ok': False, 'error': 'שני החלקים חייבים להכיל טקסט'}), 400
    _backup_db()
    conn = db.get_connection()
    try:
        v, nv, err = _cmp_reflow_rows(conn, column, verse_id, next_id)
        if err:
            return jsonify({'ok': False, 'error': err}), 400
        existing_next = (nv['val'] or '').strip()
        merged_next = (text2.strip() + (' ' + existing_next if existing_next else '')).strip()
        conn.execute('UPDATE verses SET %s=? WHERE id=?' % column, (text1.strip(), verse_id))
        conn.execute('UPDATE verses SET %s=? WHERE id=?' % column, (merged_next, next_id))
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()
    return jsonify({'ok': True})


@app.route('/api/admin/cmp_merge_next', methods=['POST'])
def admin_cmp_merge_next():
    d = request.get_json(silent=True) or {}
    if not _valid_token(d.get('token')):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    column = d.get('column')
    if column not in _CMP_REFLOW_COLS:
        return jsonify({'ok': False, 'error': 'field not mergeable'}), 400
    try:
        verse_id = int(d.get('verse_id'))
        next_id = int(d.get('next_verse_id'))
    except Exception:
        return jsonify({'ok': False, 'error': 'bad params'}), 400
    _backup_db()
    conn = db.get_connection()
    try:
        v, nv, err = _cmp_reflow_rows(conn, column, verse_id, next_id)
        if err:
            return jsonify({'ok': False, 'error': err}), 400
        cur = (v['val'] or '').strip()
        nxt = (nv['val'] or '').strip()
        merged = (cur + (' ' + nxt if nxt else '')).strip()
        conn.execute('UPDATE verses SET %s=? WHERE id=?' % column, (merged, verse_id))
        conn.execute('UPDATE verses SET %s=? WHERE id=?' % column, ('', next_id))
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()
    return jsonify({'ok': True})


# Change a verse's SAMARITAN-division number (verses.sam_number) — the Jewish
# `number` is never touched. With cascade=True, every following verse in the same
# Jewish chapter (real integer base >= the target's) has its effective Samaritan
# number shifted by the same delta (maqaf suffix preserved); otherwise only the
# target verse changes. sam_number does not affect root_index (which keys on the
# Jewish number), so no reindex is needed.
@app.route('/api/admin/renumber_verse', methods=['POST'])
def admin_renumber_verse():
    d = request.get_json(silent=True) or {}
    if not _valid_token(d.get('token')):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    try:
        verse_id = int(d.get('verse_id'))
    except Exception:
        return jsonify({'ok': False, 'error': 'bad params'}), 400
    new_number = str(d.get('new_number') or '').strip()
    cascade = bool(d.get('cascade'))
    if not new_number:
        return jsonify({'ok': False, 'error': 'מספר חדש חסר'}), 400
    _backup_db()
    conn = db.get_connection()
    try:
        v = conn.execute('SELECT id, chapter_id, number, sam_number FROM verses WHERE id=?', (verse_id,)).fetchone()
        if not v:
            return jsonify({'ok': False, 'error': 'verse not found'}), 404
        cid = v['chapter_id']
        if not cascade:
            conn.execute('UPDATE verses SET sam_number=? WHERE id=?', (new_number, verse_id))
        else:
            eff_old = str(v['sam_number'] or v['number']).split('-')[0]
            new_base_s = new_number.split('-')[0]
            real_base_s = str(v['number']).split('-')[0]
            if not (eff_old.isdigit() and new_base_s.isdigit() and real_base_s.isdigit()):
                return jsonify({'ok': False, 'error': 'נדרש מספר שלם לשינוי מדורג'}), 400
            delta = int(new_base_s) - int(eff_old); real_base = int(real_base_s)
            for r in conn.execute('SELECT id, number, sam_number FROM verses WHERE chapter_id=?', (cid,)).fetchall():
                rb = str(r['number']).split('-')[0]
                if rb.isdigit() and int(rb) >= real_base:
                    eff = str(r['sam_number'] or r['number']); eb = eff.split('-')[0]
                    if eb.isdigit():
                        conn.execute('UPDATE verses SET sam_number=? WHERE id=?',
                                     (str(int(eb) + delta) + eff[len(eb):], r['id']))
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()
    return jsonify({'ok': True})

# columns returned for a verse (everything the UI's content modes need)
_VERSE_COLS = ('id', 'number', 'text', 'english', 'masoretic_text', 'lxx_text',
               'sam_aramaic', 'onkelos_text', 'qumran_text', 'qumran_scroll',
               'arabic_trans', 'interpretation', 'rashi', 'ramban',
               'cassuto', 'baal_haturim')
# The four Jewish commentaries are the heaviest thing a verse carries — on the
# first chapter of Genesis they are 90% of the whole chapter's JSON (Cassuto
# alone 33,600 characters against 260 for the Samaritan text) — and the reader
# sees them only inside "פרשנות יהודית". So a chapter goes out without them, and
# the panel asks for them by ?full=1 when it is opened.
_HEAVY_COLS = ('rashi', 'ramban', 'cassuto', 'baal_haturim')
_LIGHT_COLS = tuple(c for c in _VERSE_COLS if c not in _HEAVY_COLS)
_NIKUD_RE = re.compile(u'[֑-ׇ]')
# everything that is NOT a Hebrew consonant (incl. niqqud, te'amim, U+034F and
# punctuation); used to reduce a word to bare consonants for the compare diff.
_HEB_LETTERS_RE = re.compile(u'[^א-ת]')


def _verse_dict(row, cols=None):
    keys = row.keys()
    return {k: (row[k] if k in keys else None) for k in (cols or _VERSE_COLS)}


def _verse_cols():
    """Which columns this request wants — everything only when it says ?full=1."""
    return _VERSE_COLS if request.args.get('full') == '1' else _LIGHT_COLS


def _ids_arg(name='verse_ids'):
    raw = request.args.get(name, '')
    return [int(x) for x in raw.split(',') if x.strip().isdigit()]


_FIN = {'ך': 'כ', 'ם': 'מ', 'ן': 'נ', 'ף': 'פ', 'ץ': 'צ'}


def _heb_fold(s):
    """Hebrew letters only, final forms folded to their base — for matching a
    searched word against the verse_dictionary entries."""
    return ''.join(_FIN.get(c, c) for c in (s or '') if ('א' <= c <= 'ת') or c in _FIN)


def _first_match_word(text, query, exact):
    """The actual word IN THE VERSE that the (plain) query matched — i.e. the word
    highlighted in the result — so the meaning reflects it, not the typed query."""
    qf = _heb_fold(query)
    if not qf:
        return ''
    for w in re.findall('[א-ת]+', text or ''):
        wf = _heb_fold(w)
        if (wf == qf if exact else qf in wf):
            return w
    return ''


_TAL_GLOSS_CACHE = {}


def _tal_gloss(aramaic_word):
    """Short HEBREW meaning of an Aramaic word, read off the authoritative Tal
    extraction (the root's primary sense). Cached — the dictionary is static."""
    if not aramaic_word:
        return ''
    if aramaic_word in _TAL_GLOSS_CACHE:
        return _TAL_GLOSS_CACHE[aramaic_word]
    g = ''
    try:
        d = db.tal_full_lookup(aramaic_word)
        for rt in d.get('roots', []):
            for s in rt.get('senses', []):
                gl = (s.get('gloss') or '').strip()
                if gl:
                    m = re.search('[A-Za-z]', gl)      # keep the Hebrew part, drop English tail
                    if m and m.start() > 2:
                        gl = gl[:m.start()].strip(' ,;·—-')
                    if re.search('[א-ת]', gl):
                        g = gl[:90]
                        break
            if g:
                break
    except Exception:
        g = ''
    _TAL_GLOSS_CACHE[aramaic_word] = g
    return g


_SAM_OPENING = {}


def _sam_opening(sam_ch_id):
    """First three words of a Samaritan chapter — shown next to a search result's
    Samaritan-division path to identify the chapter. Cached (static data)."""
    if sam_ch_id in _SAM_OPENING:
        return _SAM_OPENING[sam_ch_id]
    try:
        rows = db.get_verses_by_sam_ch(sam_ch_id)
        words = re.findall('[א-ת]+', (rows[0]['text'] if rows else '') or '')
        txt = ' '.join(words[:3])
    except Exception:
        txt = ''
    _SAM_OPENING[sam_ch_id] = txt
    return txt


# ── pages ──────────────────────────────────────────────────────────────────
def _asset_build():
    """A stamp that changes whenever app.js or style.css does, so the page can ask
    for them by version and they can then be cached for good."""
    try:
        newest = max(os.path.getmtime(os.path.join(app.static_folder, n))
                     for n in ('app.js', 'style.css'))
        return str(int(newest))
    except OSError:
        return APP_VERSION


ASSET_BUILD = _asset_build()


# the weekly post goes out on its own; see web/social.py for what "on its own"
# is allowed to mean (nothing, until an account is connected and armed)
try:
    import social as _social
    _social.start_scheduler()
except Exception as _e:                      # never let it keep the app from starting
    print('social scheduler off:', _e)


@app.route('/')
# Every screen of the app has an address of its own — /t/sam/1/8/10 is the tenth
# Samaritan chapter of the eighth portion of Genesis — so a chapter can be linked
# to, a refresh returns to where the reader was, and the phone's Back button
# steps back through the app instead of leaving it. The app itself reads the path
# and opens it; the server's part is simply to hand the same page to all of them.
@app.route('/t/<path:_rest>')
def index(_rest=None):
    return render_template('index.html', version=APP_VERSION, asset_build=ASSET_BUILD)


# Google Play will not publish an app without a reachable privacy policy, and it
# has to stay reachable for as long as the listing is up. Served as its own page,
# standalone (no app.js, no service worker shell), so a reviewer — or anyone —
# can read it without loading the whole reader. ?lang=en switches to English.
@app.route('/privacy')
def privacy():
    return render_template('privacy.html')


# ── ציר הזמן ההיסטורי השומרוני ───────────────────────────────────────────────
# A self-contained page of its own (its own HTML/CSS/JS and its own generated
# data files), served straight out of History_timeline/ rather than copied into
# web/static — one copy on disk, so its build script (History_timeline/scripts/
# build_data.py) keeps regenerating the very files the app serves. The app opens
# it in a full-screen frame from the menu.
_TIMELINE_DIR = os.path.join(_ROOT, 'History_timeline')


@app.route('/timeline')
def timeline_root():
    """The page asks for css/style.css and js/app.js by RELATIVE path. Without the
    trailing slash those resolve against the site root and 404, so the timeline
    would arrive unstyled and unscripted — hence the redirect rather than serving
    index.html here."""
    qs = request.query_string.decode()
    return redirect('/timeline/' + (('?' + qs) if qs else ''))


@app.route('/timeline/')
@app.route('/timeline/<path:sub>')
def timeline_page(sub='index.html'):
    return send_from_directory(_TIMELINE_DIR, sub)


# ── אוצר השירה השומרונית ──────────────────────────────────────────────────────
# The recordings archive, another page of its own, opened from the menu in the
# same way. Its routes live in web/shira.py because it brings a catalog and a
# media server with it, where the timeline is only files. See that module.
from shira import shira as _shira_bp                                # noqa: E402
app.register_blueprint(_shira_bp)

# אוצר כתבי היד השומרוניים בתבל — the manuscript treasury, a unit of its own
# (see manuscripts.py). Registered the same way; its scans live on the Contabo
# media host and never pass through this server.
from manuscripts import manuscripts as _mss_unit                    # noqa: E402
app.register_blueprint(_mss_unit)


@app.route('/favicon.ico')
def favicon():
    """index.html names its own icon, but a page that doesn't — the timeline —
    makes the browser ask for /favicon.ico and log a 404 in the console."""
    return send_from_directory(app.static_folder, 'img/app_icon.png')


@app.route('/api/whats_new')
def api_whats_new():
    """The current version's changelog, read from VER_UPDATES.txt on the server."""
    try:
        with open(_VER_UPDATES, encoding='utf-8') as f:
            text = f.read()
    except Exception:
        text = ''
    return jsonify({'version': APP_VERSION, 'text': text})


@app.route('/api/admin/system_doc')
def api_system_doc():
    """The system's own documentation (SYSTEM_DOC.txt) — what the system is, what
    it is built of, where each body of text came from and what every version did.
    Maintainer-facing, so it is behind the admin token like the other tools."""
    if not _valid_token(request.args.get('token')):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    try:
        with open(_SYSTEM_DOC, encoding='utf-8') as f:
            text = f.read()
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    return jsonify({'ok': True, 'version': APP_VERSION, 'text': text})


@app.route('/fonts/<path:fn>')
def fonts(fn):
    return send_from_directory(os.path.join(app.static_folder, 'fonts'), fn)


# ── how long each thing may be kept ──────────────────────────────────────────
# Everything went out as `no-cache` (Flask's default for static files), which
# meant the CDN in front of the app could not hold a single byte: every visit
# fetched app.js, the five fonts and every image from Frankfurt again, and the
# service worker then fetched the same files a SECOND time to fill its own
# cache — a cold load of 5.8MB where 2.9MB of it was the same files twice.
# The rules below are ordered by how a file actually changes:
#   fonts                 never change → a year
#   ?v=<build> assets     the URL itself changes on deploy → a year, immutable
#   images / audio / json → a week, and served from cache while refreshed behind
#   anything else static  → an hour
# The service worker and the page itself are never cached: they carry the state
# that decides everything else.
_STATIC_MEDIA = ('.png', '.jpg', '.jpeg', '.webp', '.svg', '.ico', '.mp3', '.m4a', '.json')

# Rewritten by every boundary edit, so they must never be served from cache
# without revalidating — everything else about a recording is read from them.
_LIVE_MANIFESTS = ('/static/audio/readings/readings.json',
                   '/static/audio/witnesses.json')


@app.after_request
def _cache_rules(resp):
    p = request.path
    if p in ('/sw.js', '/manifest.json') or p.startswith('/api/'):
        resp.headers['Cache-Control'] = 'no-cache'
    elif p.startswith('/fonts/'):
        resp.headers['Cache-Control'] = 'public, max-age=31536000, immutable'
    elif p.startswith('/static/'):
        if p in _LIVE_MANIFESTS:
            # These two index the recordings and are rewritten by every boundary
            # edit. A week of caching meant a listener kept the old split — and
            # kept it even after the audio itself was fixed, because the stale
            # manifest is what names the files and their cut points.
            resp.headers['Cache-Control'] = 'no-cache'
        elif request.args.get('v'):
            resp.headers['Cache-Control'] = 'public, max-age=31536000, immutable'
        elif p.endswith(_STATIC_MEDIA):
            resp.headers['Cache-Control'] = 'public, max-age=604800, stale-while-revalidate=86400'
        else:
            resp.headers['Cache-Control'] = 'public, max-age=3600'
    elif (resp.mimetype or '').startswith('text/html'):
        resp.headers['Cache-Control'] = 'no-cache'      # the page names the build
    return resp


# PWA: service worker + manifest must be served from the root scope ('/').
@app.route('/sw.js')
def sw():
    resp = send_from_directory(app.static_folder, 'sw.js')
    resp.headers['Cache-Control'] = 'no-cache'
    resp.headers['Service-Worker-Allowed'] = '/'
    return resp


@app.route('/manifest.json')
def manifest():
    return send_from_directory(app.static_folder, 'manifest.json')


# TWA (the Android app): Digital Asset Links proves that the signed package is
# allowed to act for this domain — without it the app opens with a URL bar on
# top, like a plain browser tab. The fingerprints are read from the environment
# (TWA_FINGERPRINTS, comma-separated SHA-256) so the signing key can be added or
# rotated on Render without a commit; web/assetlinks.json is the fallback.
@app.route('/.well-known/assetlinks.json')
def assetlinks():
    fps = [f.strip().upper()
           for f in os.environ.get('TWA_FINGERPRINTS', '').split(',') if f.strip()]
    if not fps:
        return send_from_directory(_WEB_DIR, 'assetlinks.json',
                                   mimetype='application/json')
    return jsonify([{
        'relation': ['delegate_permission/common.handle_all_urls'],
        'target': {
            'namespace': 'android_app',
            'package_name': os.environ.get('TWA_PACKAGE', 'net.thesamaritans.torah'),
            'sha256_cert_fingerprints': fps,
        },
    }])


# The signed Android app (the TWA), offered in the install menu alongside the
# PWA install. It lives in web/downloads/ rather than web/static/ so the service
# worker — which caches everything under /static/ — never pulls a 1.7 MB APK
# into the page cache.
_APK_DIR = os.path.join(_WEB_DIR, 'downloads')
_APK_NAME = 'samaritan-torah.apk'


@app.route('/api/apk_info')
def api_apk_info():
    """Size and version for the install card, and 'available': False before the
    APK is ever published — so the card simply omits the option instead of
    offering a download that 404s."""
    path = os.path.join(_APK_DIR, _APK_NAME)
    if not os.path.exists(path):
        return jsonify({'available': False})
    version = ''
    try:                                    # single source of truth: the TWA manifest
        with open(os.path.join(_ROOT, 'twa', 'twa-manifest.json'), encoding='utf-8') as f:
            version = _json.load(f).get('appVersion', '')
    except Exception:
        pass
    out = {'available': True, 'version': version,
           'size_mb': round(os.path.getsize(path) / 1048576.0, 1)}
    if _valid_token(request.args.get('token')):     # the tally is the admin's, not the public's
        n, last = analytics.counter('apk_download')
        out['downloads'] = n
        out['last_download'] = time.strftime('%d/%m/%Y %H:%M', time.localtime(last)) if last else None
    return jsonify(out)


@app.route('/download/samaritan-torah.apk')
def download_apk():
    if not os.path.exists(os.path.join(_APK_DIR, _APK_NAME)):
        return jsonify({'error': 'apk not published'}), 404
    analytics.bump_counter('apk_download')   # counted where the file actually leaves
    return send_from_directory(_APK_DIR, _APK_NAME, as_attachment=True,
                               mimetype='application/vnd.android.package-archive')


# The install guide, one rendering per interface language. Kept out of static/
# for the same reason the APK is: the service worker caches everything under
# /static/, and three 600 KB videos have no business sitting in the page cache
# of someone who already installed. conditional=True keeps range requests
# working, without which a phone cannot seek inside the clip.
_GUIDE_DIR = os.path.join(_WEB_DIR, 'media')
_GUIDE_LANGS = ('he', 'en', 'ar')


@app.route('/guide/<lang>.mp4')
def install_guide(lang):
    if lang not in _GUIDE_LANGS:
        lang = 'he'
    name = 'install-guide-%s.mp4' % lang
    if not os.path.exists(os.path.join(_GUIDE_DIR, name)):
        return jsonify({'error': 'guide not published'}), 404
    return send_from_directory(_GUIDE_DIR, name, mimetype='video/mp4',
                               conditional=True)


# ── navigation API ─────────────────────────────────────────────────────────
# Render waits for this to answer before it sends anyone to a new instance. It
# deliberately does real work — opens the database and reads from it — so that
# the warming up (importing, opening the file, filling the page cache) is paid
# by the deploy and not by the first reader, who used to wait about four seconds.
@app.route('/healthz')
def healthz():
    t0 = time.time()
    try:
        books = db.get_books()
        _books_payload('samaritan')                      # fills the small cache too
        return jsonify({'ok': True, 'books': len(books),
                        'ms': round((time.time() - t0) * 1000)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ── a small cache for the answers that never change between deploys ──────────
# books, portions and the chapter lists are read-mostly and tiny; they were being
# rebuilt from SQLite on every navigation (~300ms each on the live server). They
# are held in memory instead and dropped whenever an admin writes to the text.
_MEM = {}


def _mem(key, build):
    if key not in _MEM:
        _MEM[key] = build()
    return _MEM[key]


def _mem_clear():
    _MEM.clear()


def _books_payload(mode):
    def build():
        out = []
        for b in db.get_books():
            item = {'id': b['id'], 'name': b['name']}
            if mode == 'samaritan':
                item['n_portions'] = len(db.get_portions(b['id'], mode='samaritan'))
                item['n_chapters'] = len(db.get_sam_chapters(b['id']))
            out.append(item)
        return out
    return _mem('books:' + mode, build)


@app.route('/api/books')
def api_books():
    mode = request.args.get('mode', 'samaritan')
    out = _books_payload(mode)
    return jsonify(out)


@app.route('/api/portions')
def api_portions():
    book_id = int(request.args['book_id'])
    mode = request.args.get('mode', 'samaritan')          # 'samaritan' | 'standard'
    key = 'portions:%s:%s' % (book_id, mode)
    if key in _MEM:
        return jsonify(_MEM[key])
    pmode = 'jewish' if mode == 'standard' else 'samaritan'
    out = []
    for p in db.get_portions(book_id, mode=pmode):
        item = {'id': p['id'], 'name': p['name'],
                'start_ch': p['start_ch'], 'end_ch': p['end_ch']}
        if mode == 'samaritan':
            item['n_chapters'] = db.count_sam_chapters_in_portion(p['id'])
        out.append(item)
    _MEM[key] = out
    return jsonify(out)


@app.route('/api/chapters')
def api_chapters():
    """Standard (Jewish) chapters — by portion, or all of a book (spread)."""
    pid = request.args.get('portion_id')
    bid = request.args.get('book_id')
    rows = db.get_chapters(portion_id=int(pid)) if pid else db.get_chapters(book_id=int(bid))
    return jsonify([{'id': r['id'], 'number': r['number']} for r in rows])


@app.route('/api/sam_chapters')
def api_sam_chapters():
    """Samaritan chapters whose first verse falls in the portion; or all of a book."""
    pid = request.args.get('portion_id')
    bid = request.args.get('book_id')
    key = 'samch:%s:%s' % (pid, bid)
    if key in _MEM:
        return jsonify(_MEM[key])
    if pid and pid.isdigit():   # the SPA can send a literal 'null' before its state settles
        rows = db.get_sam_chapters_in_portion(int(pid))
    elif bid and bid.isdigit():
        rows = db.get_sam_chapters(int(bid))
    else:
        return jsonify([])
    texts = [r['first_text'] if 'first_text' in r.keys() else '' for r in rows]
    openings = _dedupe_openings(texts)
    out = [{'id': r['id'], 'number': r['number'], 'opening': o}
           for r, o in zip(rows, openings)]
    _MEM[key] = out
    return jsonify(out)


def _opening_words(text, n=2):
    """The first n words of a verse, for the chapter-grid incipit: drop a leading
    brace/angle gloss (e.g. '{ובקצריכם}') and Samaritan stop-marks/punctuation."""
    t = re.sub(r'^[\s]*[{<\[][^}>\]]*[}>\]][\s]*', '', text or '')   # strip a leading {..}/<..> gloss
    words = [re.sub(r'[^֐-תࠀ-࠿]', '', w) for w in t.split()]
    words = [w for w in words if w]
    return ' '.join(words[:n])


def _dedupe_openings(texts, start_n=2, max_n=8):
    """Chapter-grid incipits, but a two-word opening like 'ויהי כאשר' or 'ויאמר
    אלהים' recurs often enough in the Torah that many chapters in the SAME
    portion end up with identical-looking tiles. Only the colliding entries get
    progressively more words (3, 4, ...) until they're distinct from their
    siblings in this result set — chapters that were already unique at 2 words
    stay short. Genuinely identical formulaic openings (e.g. 5 chapters all
    starting 'ויהי אחר הדברים האלה') stop growing at max_n rather than bloating
    the tile forever."""
    n = start_n
    result = [None] * len(texts)
    pending = list(range(len(texts)))
    while pending and n <= max_n:
        groups = {}
        for i in pending:
            groups.setdefault(_opening_words(texts[i], n), []).append(i)
        next_pending = []
        for key, idxs in groups.items():
            if len(idxs) == 1:
                result[idxs[0]] = key
            else:
                next_pending.extend(idxs)
        pending = next_pending
        n += 1
    for i in pending:
        result[i] = _opening_words(texts[i], max_n)
    return result


@app.route('/api/verses')
def api_verses():
    """Standard-division verses of a Jewish chapter."""
    cid = int(request.args['chapter_id'])
    pid = request.args.get('portion_id')
    rows = db.get_verses(cid, portion_id=int(pid) if pid else None)
    cols = _verse_cols()
    out = []
    for r in rows:
        dd = _verse_dict(r, cols)
        mc = r['mas_chapter'] if 'mas_chapter' in r.keys() else None
        dd['jchapter'] = mc if mc else (r['jchapter'] if 'jchapter' in r.keys() else None)
        mn = r['mas_number'] if 'mas_number' in r.keys() else None
        dd['masnum'] = mn if mn else dd['number']     # Masoretic-comparison number
        out.append(dd)
    return jsonify(out)


@app.route('/api/sam_verses')
def api_sam_verses():
    """Samaritan-division verses of a Samaritan chapter."""
    sid = int(request.args['sam_ch_id'])
    rows = db.get_verses_by_sam_ch(sid)
    cols = _verse_cols()
    out = []
    for r in rows:
        dd = _verse_dict(r, cols)
        mc = r['mas_chapter'] if 'mas_chapter' in r.keys() else None
        dd['jchapter'] = mc if mc else (r['jchapter'] if 'jchapter' in r.keys() else None)
        mn = r['mas_number'] if 'mas_number' in r.keys() else None
        dd['masnum'] = mn if mn else dd['number']     # Masoretic-comparison number (real)
        sn = r['sam_number'] if 'sam_number' in r.keys() else None
        if sn:                          # Samaritan division shows the Samaritan number
            dd['number'] = sn
        out.append(dd)
    return jsonify(out)


@app.route('/api/sam_chapter_marks')
def api_sam_chapter_marks():
    """What a chapter's landing needs and cannot see in its own verses: the last
    word of the PREVIOUS Samaritan chapter of the same book. Genesis blesses the
    reader who has just come through a chapter that ended in death."""
    sid = int(request.args['sam_ch_id'])
    conn = db.get_connection()
    out = {'prev_last_word': None, 'anim': _anim_overrides(conn, sid)}
    sc = conn.execute('SELECT book_id, number FROM sam_chapters WHERE id=?', (sid,)).fetchone()
    if sc and sc['number'] > 1:
        prev = conn.execute('SELECT id FROM sam_chapters WHERE book_id=? AND number=?',
                            (sc['book_id'], sc['number'] - 1)).fetchone()
        if prev:
            r = conn.execute("""SELECT v.text FROM verses v JOIN chapters c ON c.id = v.chapter_id
                                WHERE v.sam_ch_id = ?
                                ORDER BY c.number DESC, CAST(v.number AS INTEGER) DESC LIMIT 1""",
                             (prev['id'],)).fetchone()
            if r:
                words = re.findall(r'[א-ת]+', r['text'] or '')
                out['prev_last_word'] = words[-1] if words else None
    conn.close()
    return jsonify(out)


# The admin's animation flags have to judge every chapter of a book from the
# outside, and the rules turn on things a chapter list does not carry: the words
# a chapter opens with (beyond its two-word incipit), the word the chapter before
# it ended on, and whether any verse inside it opens or closes with a given
# phrase. The PHRASES come from the client, so the rules themselves stay in one
# place (the blessing tables in app.js) and only the searching happens here.
# ── the administrator's changes to a chapter's animations ────────────────────
# The rules in app.js say what a chapter carries by nature; this table is where
# the project owner overrules them — turning one off, moving it from the opening
# of a chapter to its end, or giving it other words. One row per animation, kept
# by the chapter and a stable slot ('entry', 'chapter-end', 'verse-end:5'), and
# read by EVERY reader, not only by an admin: a blessing switched off here must
# stop appearing for everyone.
def _anim_table(conn):
    conn.execute("""CREATE TABLE IF NOT EXISTS anim_overrides (
        sam_ch_id  INTEGER NOT NULL,
        slot       TEXT    NOT NULL,
        enabled    INTEGER NOT NULL DEFAULT 1,
        text       TEXT,
        timing     TEXT,
        updated_at TEXT,
        PRIMARY KEY (sam_ch_id, slot))""")


def _anim_overrides(conn, sam_ch_id):
    _anim_table(conn)
    out = {}
    for r in conn.execute('SELECT slot, enabled, text, timing FROM anim_overrides WHERE sam_ch_id=?',
                          (sam_ch_id,)):
        out[r['slot']] = {'enabled': bool(r['enabled']),
                          'text': r['text'], 'timing': r['timing']}
    return out


@app.route('/api/admin/anim_override', methods=['POST'])
def admin_anim_override():
    d = request.get_json(silent=True) or {}
    if not _valid_token(d.get('token')):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    try:
        sid = int(d.get('sam_ch_id'))
    except Exception:
        return jsonify({'ok': False, 'error': 'bad params'}), 400
    slot = (d.get('slot') or '').strip()
    if not slot:
        return jsonify({'ok': False, 'error': 'no slot'}), 400
    timing = d.get('timing')
    if timing not in (None, '', 'entry', 'chapter-end', 'verse-end', 'verse-start'):
        return jsonify({'ok': False, 'error': 'bad timing'}), 400
    text = d.get('text')
    if text is not None:
        text = str(text)[:80].strip() or None      # as long as the field the admin types in
    enabled = 1 if d.get('enabled', True) else 0
    conn = db.get_connection()
    try:
        _anim_table(conn)
        if d.get('delete'):                    # a blessing the owner added, taken back
            conn.execute('DELETE FROM anim_overrides WHERE sam_ch_id=? AND slot=?', (sid, slot))
            conn.commit()
            return jsonify({'ok': True, 'anim': _anim_overrides(conn, sid)})
        conn.execute("""INSERT INTO anim_overrides (sam_ch_id, slot, enabled, text, timing, updated_at)
                        VALUES (?,?,?,?,?,?)
                        ON CONFLICT(sam_ch_id, slot) DO UPDATE SET
                          enabled=excluded.enabled, text=excluded.text,
                          timing=excluded.timing, updated_at=excluded.updated_at""",
                     (sid, slot, enabled, text, (timing or None),
                      time.strftime('%Y-%m-%d %H:%M:%S')))
        conn.commit()
        out = _anim_overrides(conn, sid)
    finally:
        conn.close()
    return jsonify({'ok': True, 'anim': out})


# ── the weekly post about the portion of the week ────────────────────────────
# The mechanism itself is web/social.py. Here are only the three doors the admin
# panel knocks on: what is connected, what the coming week's post looks like, and
# "send it". Nothing publishes without an account being connected AND armed by
# the owner, and the weekly hand (social.start_scheduler) does the same work on
# its own — a person is never in the loop.
@app.route('/api/admin/social', methods=['GET', 'POST'])
def admin_social():
    d = request.get_json(silent=True) or {}
    tok = d.get('token') or request.args.get('token')
    if not _valid_token(tok):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    import social
    if request.method == 'POST':
        net = d.get('network')
        try:
            accounts = social.set_account(net, d.get('config') or {}, bool(d.get('armed')))
        except ValueError as e:
            return jsonify({'ok': False, 'error': str(e)}), 400
    else:
        accounts = social.accounts()
    week, entry = social.coming_sabbath()
    return jsonify({'ok': True, 'accounts': accounts, 'week': week,
                    'portion': (entry or {}).get('name'), 'post': social.post_row(week),
                    'history': social.recent_posts()})


@app.route('/api/admin/social/preview', methods=['POST'])
def admin_social_preview():
    d = request.get_json(silent=True) or {}
    if not _valid_token(d.get('token')):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    import social
    week, entry = social.coming_sabbath()
    post = social.build(week, entry)
    return jsonify({'ok': bool(post), 'week': week, 'post': post})


@app.route('/api/admin/social/publish', methods=['POST'])
def admin_social_publish():
    d = request.get_json(silent=True) or {}
    if not _valid_token(d.get('token')):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    import social
    return jsonify(social.publish(dry=bool(d.get('dry')), force=bool(d.get('force'))))


@app.route('/api/anim_marks')
def api_anim_marks():
    bid = int(request.args['book_id'])
    ends = [p for p in request.args.getlist('ends') if p.strip()]
    starts = [p for p in request.args.getlist('starts') if p.strip()]
    conn = db.get_connection()
    rows = conn.execute("""SELECT sc.number num, v.text
                           FROM verses v
                           JOIN chapters c ON c.id = v.chapter_id
                           JOIN sam_chapters sc ON sc.id = v.sam_ch_id
                           WHERE c.book_id = ? AND v.sam_ch_id IS NOT NULL
                           ORDER BY sc.number, c.number, CAST(v.number AS INTEGER)""", (bid,)).fetchall()
    conn.close()

    def fold(s):
        s = _HEB_LETTERS_RE.sub('', _NIKUD_RE.sub('', s or ''))
        for a, b in (('ך', 'כ'), ('ם', 'מ'), ('ן', 'נ'), ('ף', 'פ'), ('ץ', 'צ')):
            s = s.replace(a, b)
        return s

    fends = [fold(p) for p in ends]
    fstarts = [fold(p) for p in starts]
    per = {}                    # Samaritan chapter number → what the rules ask about
    for r in rows:
        d = per.setdefault(str(r['num']), {'first': '', 'last_word': None,
                                           'ends': False, 'starts': False})
        if not d['first']:
            d['first'] = (r['text'] or '')[:200]
        words = re.findall(r'[א-ת]+', r['text'] or '')
        if words:
            d['last_word'] = words[-1]
        f = fold(r['text'])
        if f:
            if any(f.endswith(p) for p in fends):
                d['ends'] = True
            if any(f.startswith(p) for p in fstarts):
                d['starts'] = True
    return jsonify(per)


@app.route('/api/canon_note')
def api_canon_note():
    """Canon note for a book, if this sam_ch_id is that book's LAST Samaritan chapter —
    powers the fixed 'N chapters, engraved' marker shown after the book's last verse."""
    sid = int(request.args['sam_ch_id'])
    conn = db.get_connection()
    sc = conn.execute('SELECT book_id, number FROM sam_chapters WHERE id=?', (sid,)).fetchone()
    if not sc:
        return jsonify(None)
    out = {}
    canon = conn.execute('SELECT canonical_count, note FROM canon_chapter_counts WHERE book_id=?',
                          (sc['book_id'],)).fetchone()
    if canon and sc['number'] == canon['canonical_count']:
        out.update({'count': canon['canonical_count'], 'note': canon['note']})
    # the same, one level down: the portion's own signature, after its last chapter
    counts = _portion_chapter_counts(conn, sc['book_id'])
    pid = _portion_id_for_sam_chapter(conn, sid)
    info = counts.get(pid)
    if info and info['last_number'] == sc['number']:
        pc = conn.execute('SELECT canonical_count, portion_name, note FROM canon_portion_counts '
                          'WHERE portion_id=?', (pid,)).fetchone()
        out['portion'] = {'name': info['name'], 'live': info['count'], 'last': True,
                          'sam_ch_id': sid,
                          'count': pc['canonical_count'] if pc else None,
                          'note': pc['note'] if pc else None}
    return jsonify(out or None)


# ── content-mode API ───────────────────────────────────────────────────────
@app.route('/api/interpretations')
def api_interpretations():
    ids = _ids_arg()
    rows = [{'id': i} for i in ids]
    lang = request.args.get('lang', 'he')
    m = get_chapter_interpretations(rows, lang) if ids else {}
    return jsonify({str(k): v for k, v in m.items()})


@app.route('/api/dictionary')
def api_dictionary():
    m = db.get_verse_dictionary(_ids_arg())
    return jsonify({str(k): v for k, v in m.items()})


@app.route('/api/word_table')
def api_word_table():
    m = db.get_word_table(_ids_arg())
    return jsonify({str(k): v for k, v in m.items()})


@app.route('/api/dict_select')
def api_dict_select():
    """Per verse, which word-index maps to which dictionary row — powers the
    word-by-word picker of the 'מילון מילים' mode (underline a word, tap to open it)."""
    m = db.get_dict_select(_ids_arg())
    return jsonify({str(k): v for k, v in m.items()})


@app.route('/api/tibat_marqe')
def api_tibat_marqe():
    return jsonify(db.get_tibat_marqe(_ids_arg()))


@app.route('/api/translit')
def api_translit():
    return jsonify(db.get_translit(_ids_arg()))


# ── server-side read-aloud: Azure Neural TTS driven by IPA phonemes, so the correct
#    (penultimate) stress and Ben-Ḥayyim phonetics are honoured. Audio is cached per
#    verse. Returns 503 until AZURE_SPEECH_KEY/REGION are set in the environment. ──
import urllib.request as _urlreq
from xml.sax.saxutils import escape as _xesc
from app.services.ipa import ipa_words as _ipa_words
from app.services.heb import translit_to_heb as _translit_to_heb
_TTS_CACHE = os.path.join(_ROOT, 'data', 'tts_cache')
_AZ_KEY = os.environ.get('AZURE_SPEECH_KEY', '')
_AZ_REGION = os.environ.get('AZURE_SPEECH_REGION', '')
_AZ_VOICE = os.environ.get('AZURE_SPEECH_VOICE', 'he-IL-HilaNeural')
# 'ipa' = IPA phoneme tags — he-IL honours them (verified: invalid phones → HTTP 400), giving
# authentic Samaritan phonetics: /w/ (no /v/), hard /b/,/k/, pharyngeal /ħ/, and mil'el stress.
# 'heb' = pointed Hebrew fallback (natural but forces /v/ and modern mil'ra stress).
_AZ_MODE = os.environ.get('AZURE_SPEECH_MODE', 'ipa')


@app.route('/api/tts_status')
def api_tts_status():
    return jsonify({'enabled': bool(_AZ_KEY and _AZ_REGION), 'voice': _AZ_VOICE, 'mode': _AZ_MODE})


@app.route('/api/tts_probe')
def api_tts_probe():
    """Debug: synthesize an arbitrary IPA string on any voice, to calibrate phones/glides."""
    if not (_AZ_KEY and _AZ_REGION):
        return ('TTS not configured', 503)
    ph = request.args.get('ph', '')
    voice = request.args.get('voice', _AZ_VOICE)
    if not ph:
        return ('missing ph', 400)
    lang = '-'.join(voice.split('-')[:2]) if voice.count('-') >= 2 else 'he-IL'
    inner = '<phoneme alphabet="ipa" ph="%s">x</phoneme>' % _xesc(ph, {'"': '&quot;'})
    ssml = ('<speak version="1.0" xmlns="http://www.w3.org/2001/10/synthesis" xml:lang="%s">'
            '<voice name="%s">%s</voice></speak>' % (lang, voice, inner))
    req = _urlreq.Request(
        'https://%s.tts.speech.microsoft.com/cognitiveservices/v1' % _AZ_REGION,
        data=ssml.encode('utf-8'), method='POST',
        headers={'Ocp-Apim-Subscription-Key': _AZ_KEY, 'Content-Type': 'application/ssml+xml',
                 'X-Microsoft-OutputFormat': 'audio-24khz-48kbitrate-mono-mp3', 'User-Agent': 'avnei-shoham'})
    try:
        with _urlreq.urlopen(req, timeout=30) as resp:
            audio = resp.read()
    except Exception as e:
        return ('tts error: %s' % e, 502)
    from flask import Response
    return Response(audio, mimetype='audio/mpeg')


@app.route('/api/tts')
def api_tts():
    if not (_AZ_KEY and _AZ_REGION):
        return ('TTS not configured', 503)
    try:
        vid = int(request.args.get('verse_id', ''))
    except (TypeError, ValueError):
        return ('bad verse_id', 400)
    mode = request.args.get('mode', _AZ_MODE)
    voice = request.args.get('voice', _AZ_VOICE)
    text = db.get_translit([vid]).get(vid)
    if not text or not text.strip():
        return ('no transcription', 404)
    if mode == 'ipa':
        pairs = _ipa_words(text)
        if not pairs:
            return ('no transcription', 404)
        inner = ' '.join('<phoneme alphabet="ipa" ph="%s">%s</phoneme>'
                         % (_xesc(ip, {'"': '&quot;'}), _xesc(w)) for w, ip in pairs)
        sig = ' '.join(ip for _, ip in pairs)
    else:                                            # 'heb' — natural neural reading of pointed Hebrew
        heb = _translit_to_heb(text)
        if not heb.strip():
            return ('no transcription', 404)
        inner = _xesc(heb)
        sig = heb
    # xml:lang follows the voice's locale (an Arabic voice renders a true [w], [ħ], [q] — closer
    # to Samaritan than modern Hebrew, whose voice realises /w/ as [v]).
    lang = '-'.join(voice.split('-')[:2]) if voice.count('-') >= 2 else 'he-IL'
    ssml = ('<speak version="1.0" xmlns="http://www.w3.org/2001/10/synthesis" xml:lang="%s">'
            '<voice name="%s"><prosody rate="-8%%">%s</prosody></voice></speak>' % (lang, voice, inner))
    key = hashlib.sha1(('%s|%s|%s' % (voice, mode, sig)).encode('utf-8')).hexdigest()
    path = os.path.join(_TTS_CACHE, key + '.mp3')
    if not os.path.exists(path):
        req = _urlreq.Request(
            'https://%s.tts.speech.microsoft.com/cognitiveservices/v1' % _AZ_REGION,
            data=ssml.encode('utf-8'), method='POST',
            headers={'Ocp-Apim-Subscription-Key': _AZ_KEY,
                     'Content-Type': 'application/ssml+xml',
                     'X-Microsoft-OutputFormat': 'audio-24khz-48kbitrate-mono-mp3',
                     'User-Agent': 'avnei-shoham'})
        try:
            with _urlreq.urlopen(req, timeout=30) as resp:
                audio = resp.read()
        except Exception as e:
            return ('tts error: %s' % e, 502)
        os.makedirs(_TTS_CACHE, exist_ok=True)
        with open(path, 'wb') as f:
            f.write(audio)
    return send_file(path, mimetype='audio/mpeg', conditional=True)


# ── Tibåt Mårqe full-book reader (Samaritan Library) ──
@app.route('/api/tm_toc')
def api_tm_toc():
    return jsonify(db.get_tm_toc())


@app.route('/api/tm_chapter')
def api_tm_chapter():
    return jsonify(db.get_tm_chapter(request.args.get('book', '')))


@app.route('/api/tm_search')
def api_tm_search():
    return jsonify(db.search_tm(request.args.get('q', '')))


@app.route('/api/tm_words')
def api_tm_words():
    return jsonify(db.get_tm_words(request.args.get('book', '')))


@app.route('/api/locate_verse')
def api_locate_verse():
    try:
        vid = int(request.args.get('verse_id', ''))
    except ValueError:
        return jsonify(None)
    return jsonify(db.locate_verse(vid))


# ── Ṣadaqah al-Ḥakīm full-book reader (Genesis) ──
@app.route('/api/tz_toc')
def api_tz_toc():
    return jsonify(db.get_tz_toc())


@app.route('/api/tz_chapter')
def api_tz_chapter():
    return jsonify(db.get_tz_chapter(request.args.get('chap', '')))


@app.route('/api/tz_search')
def api_tz_search():
    return jsonify(db.search_tz(request.args.get('q', '')))


@app.route('/api/eyalk')
def api_eyalk():
    return jsonify(db.get_eyalk_commentary(_ids_arg()))


@app.route('/api/tzdaka')
def api_tzdaka():
    return jsonify(db.get_tzdaka_commentary(_ids_arg()))


@app.route('/api/bhuq')
def api_bhuq():
    return jsonify(db.get_bhuq_commentary(_ids_arg()))


@app.route('/api/sir')
def api_sir():
    return jsonify(db.get_sir_commentary(_ids_arg()))


@app.route('/api/sir_toc')
def api_sir_toc():
    return jsonify(db.get_sir_toc())


@app.route('/api/sir_chapter')
def api_sir_chapter():
    return jsonify(db.get_sir_chapter(request.args.get('sec', '')))


@app.route('/api/sir_search')
def api_sir_search():
    return jsonify(db.search_sir(request.args.get('q', '')))


@app.route('/api/bhuq_by_verse')
def api_bhuq_by_verse():
    """פירוש אם בחקותי keyed by verse — feeds the verse-commentary panel, which
    sets each passage under the verse it argues about."""
    lang = 'ar' if request.args.get('lang') == 'ar' else 'he'
    return jsonify(db.get_bhuq_by_verse(_ids_arg(), lang))


# ── פירוש אם בחקותי: full-book reader (the verse source is /api/bhuq above) ──
@app.route('/api/bhuq_toc')
def api_bhuq_toc():
    return jsonify(db.get_bhuq_toc())


@app.route('/api/bhuq_chapter')
def api_bhuq_chapter():
    return jsonify(db.get_bhuq_chapter(request.args.get('part', '')))


@app.route('/api/bhuq_search')
def api_bhuq_search():
    return jsonify(db.search_bhuq(request.args.get('q', '')))


# ── ספר האסאטיר: verse source + full-book reader ──
@app.route('/api/asatir')
def api_asatir():
    return jsonify(db.get_asatir_commentary(_ids_arg()))


@app.route('/api/asatir_by_verse')
def api_asatir_by_verse():
    """Asatir passages keyed by verse — feeds the verse-commentary panel, which
    sets each passage under the verse it retells."""
    return jsonify(db.get_asatir_by_verse(_ids_arg()))


@app.route('/api/asatir_toc')
def api_asatir_toc():
    return jsonify(db.get_asatir_toc())


@app.route('/api/asatir_chapter')
def api_asatir_chapter():
    return jsonify(db.get_asatir_chapter(request.args.get('chap', '')))


@app.route('/api/asatir_search')
def api_asatir_search():
    return jsonify(db.search_asatir(request.args.get('q', '')))


# ── Samaritan piyyutim reader + rhyme finder ────────────────────────────────
@app.route('/api/piyutim_toc')
def api_piyutim_toc():
    return jsonify(db.get_piyutim_toc())


@app.route('/api/piyutim_chapter')
def api_piyutim_chapter():
    r = db.get_piyutim_chapter(request.args.get('id', ''))
    if not r:
        return jsonify({'error': 'not found'}), 404
    return jsonify(r)


@app.route('/api/piyutim_search')
def api_piyutim_search():
    return jsonify(db.search_piyutim(request.args.get('q', '')))


@app.route('/api/piyutim_dict')
def api_piyutim_dict():
    return jsonify(db.get_piyutim_dict())


@app.route('/api/piyutim_rhyme_sounds')
def api_piyutim_rhyme_sounds():
    return jsonify(db.piyutim_rhyme_sounds())


@app.route('/api/piyutim_rhyme')
def api_piyutim_rhyme():
    mode = request.args.get('mode', 'sound')
    q = request.args.get('q', '')
    clean_only = request.args.get('clean_only') == '1'
    start_letter = request.args.get('start_letter', '')
    group = request.args.get('group', '')
    return jsonify(db.piyutim_rhyme_search(mode, q, clean_only, start_letter, group))


# ── אישי השומרונים (who's-who library unit) ─────────────────────────────────
@app.route('/api/people_toc')
def api_people_toc():
    return jsonify(db.get_people_toc())


@app.route('/api/person')
def api_person():
    r = db.get_person(request.args.get('id', ''))
    if not r:
        return jsonify({'error': 'not found'}), 404
    return jsonify(r)


@app.route('/api/people_search')
def api_people_search():
    return jsonify(db.search_people(request.args.get('q', '')))


@app.route('/api/shyt')
def api_shyt():
    return jsonify(db.get_shyt_commentary(_ids_arg()))


@app.route('/api/shyt_toc')
def api_shyt_toc():
    return jsonify(db.get_shyt_toc())


@app.route('/api/shyt_chapter')
def api_shyt_chapter():
    return jsonify(db.get_shyt_chapter(request.args.get('q', '')))


@app.route('/api/shyt_search')
def api_shyt_search():
    return jsonify(db.search_shyt(request.args.get('q', '')))


@app.route('/api/apparatus')
def api_apparatus():
    return jsonify(db.get_apparatus(_ids_arg()))


@app.route('/api/tal')
def api_tal():
    word = request.args.get('word', '')
    return jsonify(db.lookup_tal_dictionary(word))


@app.route('/api/tal_lookup')
def api_tal_lookup():
    """Authoritative Tal-dictionary entry for an Aramaic word — root(s), the Hebrew
    senses read off the dictionary, the word's Torah occurrences, and related forms.
    Powers the in-app dictionary, the word-table 'more results', and search."""
    word = request.args.get('word', '')
    return jsonify(db.tal_full_lookup(word))


# ── dictionary app: page-browse · index-browse · direct word search · form locations ──
@app.route('/api/dict_page')
def api_dict_page():
    return jsonify(db.get_dict_page(request.args.get('page', '1')))


@app.route('/api/dict_index')
def api_dict_index():
    return jsonify(db.get_dict_index(request.args.get('start', '0'),
                                     prefix=request.args.get('prefix', '')))


@app.route('/api/dict_word')
def api_dict_word():
    return jsonify(db.dict_word_search(request.args.get('word', '')))


@app.route('/api/dict_locations')
def api_dict_locations():
    return jsonify(db.dict_form_locations(request.args.get('word', '')))


# ── comprehensive word index: browse every dictionary word · drill into one word ──
@app.route('/api/dict_words')
def api_dict_words():
    return jsonify(db.dict_words_browse(request.args.get('start', '0'),
                                        prefix=request.args.get('prefix', '')))


@app.route('/api/dict_phrases')
def api_dict_phrases():
    """Browse the Aramaic set phrases — epithets and idioms harvested from the
    piyyutim and Memar Marqe. `q` filters by any word in the phrase."""
    return jsonify(db.dict_phrases_browse(request.args.get('q', '')))


@app.route('/api/dict_he')
def api_dict_he():
    """Hebrew-side word index — Hebrew words leading to their Aramaic root(s)."""
    return jsonify(db.dict_he_browse(request.args.get('start', '0'),
                                     prefix=request.args.get('prefix', '')))


@app.route('/api/dict_he_search')
def api_dict_he_search():
    """Search a Hebrew word among the results → the Aramaic root(s) it renders."""
    return jsonify(db.dict_he_search(request.args.get('word', '')))


@app.route('/api/dict_word_detail')
def api_dict_word_detail():
    """A clicked index word, grouped by meaning (root): Tal sense(s), the Torah
    verses where that root occurs, and the Tibåt Mårqe passages with the same root —
    every result guaranteed to share the word's meaning via its root."""
    return jsonify(db.dict_word_detail(request.args.get('word', ''),
                                       root=request.args.get('root') or None))


@app.route('/api/root_box')
def api_root_box():
    """Index-extracted root for the editable root box (runs as the user types)."""
    return jsonify({'root': db.root_from_index(request.args.get('word', ''))})


@app.route('/api/sefaria')
def api_sefaria():
    """Live, free, key-less extra Jewish commentators from Sefaria for one verse
    (the 'פרשנים נוספים (ספריא)' option). Resolves the verse's Jewish ref first."""
    from app.services import sefaria_live
    vid = request.args.get('verse_id')
    if not vid or not vid.isdigit():
        return jsonify({'ok': False, 'items': [], 'error': 'bad verse'})
    ref = db.get_verse_ref(int(vid))
    if ref is None:
        return jsonify({'ok': False, 'items': [], 'error': 'no ref'})
    try:
        items = sefaria_live.fetch_live_commentaries(ref['book'], ref['chapter'], ref['verse'])
    except Exception:
        return jsonify({'ok': False, 'items': [], 'error': 'fetch failed'})
    return jsonify({'ok': True, 'items': [{'name': n, 'text': t} for n, t in items]})


@app.route('/api/online_dict')
def api_online_dict():
    """Free, key-less Hebrew-Hebrew definitions (Wiktionary + Wikipedia) for the
    given words, looked up in bulk."""
    from app.services import hebrew_dict
    words = [w for w in request.args.get('words', '').split(',') if w.strip()]
    if not words:
        return jsonify({})
    try:
        res = hebrew_dict.lookup_many(words)
    except Exception:
        res = {}
    out = {}
    for w, payload in res.items():
        if payload and payload[0]:
            out[w] = {'summary': payload[0],
                      'sources': [[name, site] for name, site in payload[1]]}
    return jsonify(out)


# ── compare (Masoretic vs Samaritan) diff, computed server-side ─────────────
def _diff_tokens(sam_num, mas_num, sam_raw, mas_raw):
    """Returns (sam_tokens, mas_tokens); each token is [word, is_diff]. Comparison is
    consonant-only: niqqud, cantillation, the combining grapheme joiner (U+034F) and
    punctuation (periods, dashes, colons, maqaf) are all ignored, so only genuine
    letter differences between the versions are highlighted — the displayed words keep
    their original spelling and marks. Each column carries its own leading number token
    (Samaritan number on the Samaritan side, Masoretic number on the Masoretic side —
    they can differ)."""
    MAQAF = u'־'
    sam_words = sam_raw.split() if sam_raw else []
    mas_words = mas_raw.split() if mas_raw else []
    sam_numtok = [str(sam_num), False]
    mas_numtok = [str(mas_num), False]
    if not sam_words and not mas_words:
        return [], []
    if not sam_words:
        return [], [mas_numtok] + [[w, False] for w in mas_words]
    if not mas_words:
        return [sam_numtok] + [[w, False] for w in sam_words], []

    def tokenize(words):
        # token -> list of consonant-only atoms (maqaf-separated). Atoms that hold no
        # Hebrew letter (pure punctuation: '.', '--', ':--', …) are dropped, so such
        # tokens never count as a difference.
        tokens = []
        for w in words:
            atoms = [_HEB_LETTERS_RE.sub(u'', a) for a in w.split(MAQAF)]
            atoms = [a for a in atoms if a]
            tokens.append((w, atoms))
        return tokens

    sam_tok = tokenize(sam_words)
    mas_tok = tokenize(mas_words)
    sam_atoms, sam_a2t = [], {}
    for ti, (_, atoms) in enumerate(sam_tok):
        for a in atoms:
            sam_a2t[len(sam_atoms)] = ti
            sam_atoms.append(a)
    mas_atoms, mas_a2t = [], {}
    for ti, (_, atoms) in enumerate(mas_tok):
        for a in atoms:
            mas_a2t[len(mas_atoms)] = ti
            mas_atoms.append(a)
    sam_diff = [False] * len(sam_tok)
    mas_diff = [False] * len(mas_tok)
    for tag, i1, i2, j1, j2 in difflib.SequenceMatcher(
            None, sam_atoms, mas_atoms, autojunk=False).get_opcodes():
        if tag != 'equal':
            for ai in range(i1, i2):
                sam_diff[sam_a2t[ai]] = True
            for aj in range(j1, j2):
                mas_diff[mas_a2t[aj]] = True
    sam_tokens = [sam_numtok] + [[w, sam_diff[i]] for i, (w, _) in enumerate(sam_tok)]
    mas_tokens = [mas_numtok] + [[w, mas_diff[i]] for i, (w, _) in enumerate(mas_tok)]
    return sam_tokens, mas_tokens


@app.route('/api/compare', methods=['POST'])
def api_compare():
    """Body: {verses:[{number,text,masoretic_text}, ...]}. Returns per-verse diff
    token lists for the Samaritan and Masoretic columns."""
    data = request.get_json(force=True)
    out = []
    for v in data.get('verses', []):
        st, mt = _diff_tokens(v.get('sam_num'), v.get('mas_num'),
                              v.get('text') or '', v.get('masoretic_text') or '')
        out.append({'sam': st, 'mas': mt})
    return jsonify(out)


# ── search ─────────────────────────────────────────────────────────────────
def _do_search(args):
    """Run a verse search and return the JSON-able result dict (rows + meta).
    Shared by /api/search and the Excel export so both are identical."""
    query = args.get('q', '').strip()
    if not query:
        return {'rows': [], 'count': 0}
    exact = args.get('exact') == '1'
    aramaic = args.get('aramaic') == '1'
    root_flag = args.get('root') == '1'
    root = root_flag and len(query.split()) == 1
    root_letters = args.get('root_letters') or None
    ignore_finals = args.get('ignore_finals') == '1'
    if exact and root:
        root = False

    rows, total = db.search_verses(query, exact=exact, root=root, aramaic=aramaic,
                                   root_letters=root_letters if root else None,
                                   ignore_finals=ignore_finals)

    occ_map, searched_root = {}, ''
    if not aramaic and rows:
        from app.services.hebrew_root import normalize
        if root:
            searched_root = (normalize(root_letters) if root_letters
                             else normalize(db.root_from_index(query) or ''))
            occ_map = db.get_root_occurrences(
                searched_root, [(r['id'], r['text']) for r in rows])
            rows = [r for r in rows if r['id'] in occ_map]
            rows = sorted(rows, key=lambda r: occ_map.get(r['id'], {}).get('order', 1 << 30))
        else:
            occ_map = db.get_word_occurrences(query, [r['id'] for r in rows])

    # batch the per-verse Aramaic word-pairs once, for the meaning enrichment
    vdict = db.get_verse_dictionary([r['id'] for r in rows]) if rows else {}

    out = []
    for r in rows:
        sam = db.get_samaritan_location(r['id'])
        info = occ_map.get(r['id'])
        item = {
            'id': r['id'], 'number': r['number'],
            'text': r['text'], 'sam_aramaic': r['sam_aramaic'],
            'book_id': r['book_id'], 'book_name': r['book_name'],
            'chapter_id': r['chapter_id'], 'chapter_num': r['chapter_num'],
            'portion_id': r['portion_id'], 'portion_name': r['portion_name'] or '',
            'sam': None, 'occ': None, 'match_words': None, 'subroot': '',
            'aramaic': '', 'meaning': '', 'matched_word': '',
        }
        if sam and sam['sam_portion_id']:
            item['sam'] = {
                'sam_ch_id': sam['sam_ch_id'], 'sam_ch_num': sam['sam_ch_num'],
                'number': sam['number'], 'sam_portion_id': sam['sam_portion_id'],
                'sam_portion_name': sam['sam_portion_name'],
                'opening': _sam_opening(sam['sam_ch_id']),
            }
        if info:
            item['occ'] = [list(o) for o in info.get('occ', [])]
            item['subroot'] = info.get('subroot') or ''
            if root and not aramaic:
                item['match_words'] = info.get('words') or []

        # meaning of the HIGHLIGHTED word (not the typed query): its Aramaic
        # translation (from the verse's word-pairs) + the gloss from Tal's dict.
        if info and root and info.get('words'):
            mword = info['words'][0]
        else:
            mword = _first_match_word(r['sam_aramaic'] if aramaic else r['text'],
                                      query, exact) or query
        item['matched_word'] = mword
        # pronunciation: for non-root searches the index pron can belong to a
        # different word (esp. multi-term '+' queries), so use the transcription of
        # THIS matched word, gated by relatedness; drop unrelated index prons.
        if not root and not aramaic:
            tpron = db.verse_word_pron(r['id'], mword)
            if tpron:
                binyan = item['occ'][0][1] if item['occ'] else ''
                item['occ'] = [[tpron, binyan, '']]
            elif item['occ']:
                item['occ'] = [o for o in item['occ'] if db.pron_related(mword, o[0])] or None
        pairs = vdict.get(r['id'], [])
        cands = [c for c in [_heb_fold(mword)] if c]
        aramaic_w = ''
        for a, h in pairs:                                  # exact word match first
            side = _heb_fold(a if aramaic else h)
            if side and side in cands:
                aramaic_w = a; break
        if not aramaic_w:                                   # then a substring match
            for a, h in pairs:
                side = _heb_fold(a if aramaic else h)
                if side and any(c in side or side in c for c in cands):
                    aramaic_w = a; break
        item['aramaic'] = aramaic_w
        item['meaning'] = _tal_gloss(aramaic_w)
        out.append(item)

    return {
        'rows': out, 'count': total, 'shown': len(out),
        'aramaic': aramaic, 'root': root, 'exact': exact, 'query': query,
        'searched_root': searched_root,
        'root_requested_multi': (root_flag and not root),
    }


@app.route('/api/search')
def api_search():
    return jsonify(_do_search(request.args))


def _clean_pron(p):
    """Latin transliteration only (drop Hebrew/Arabic and parentheticals that hold
    them) — mirrors cleanPron() in the client so the export matches the screen."""
    p = re.sub(r'\([^)]*[א-ת؀-ۿ][^)]*\)', '', p or '')
    p = re.sub(r'[א-ת؀-ۿ]', '', p)
    return re.sub(r'\s+', ' ', p).strip()


def _uniq(seq):
    out = []
    for x in seq:
        if x and x not in out:
            out.append(x)
    return out


@app.route('/api/search_export')
def api_search_export():
    """Export the current search results to an .xlsx with the columns the user asked
    for: Samaritan path · Jewish path · the verse · the matched word(s) · binyan ·
    Latin transliteration · the word's meaning."""
    import io as _io
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    data = _do_search(request.args)
    rows = data.get('rows', [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'תוצאות חיפוש'
    ws.sheet_view.rightToLeft = True
    headers = ['נתיב שומרוני', 'נתיב יהודי', 'המשפט', 'המילים בתוצאה',
               'בניין', 'הגייה לטינית', 'פירוש המילה']
    ws.append(headers)
    hfill = PatternFill('solid', fgColor='1A3873')
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = hfill
        c.alignment = Alignment(horizontal='center', vertical='center')

    for r in rows:
        sam = r.get('sam') or {}
        sam_path = ''
        if sam:
            op = ('  (%s)' % sam['opening']) if sam.get('opening') else ''
            sam_path = '%s › %s › פרק שומרוני %s פסוק %s%s' % (
                r['book_name'], sam.get('sam_portion_name', ''),
                sam.get('sam_ch_num', ''), sam.get('number', ''), op)
        jew_path = '%s › %s › פרק %s פסוק %s' % (
            r['book_name'], r.get('portion_name', ''), r['chapter_num'], r['number'])
        verse = (r.get('sam_aramaic') if data.get('aramaic') else r.get('text')) or ''
        words = ' '.join(r['match_words']) if r.get('match_words') else (r.get('matched_word') or '')
        occ = r.get('occ') or []
        binyan = ' · '.join(_uniq(o[1] for o in occ if len(o) > 1))
        pron = ' · '.join(_uniq(_clean_pron(o[0]) for o in occ if o))
        meaning = r.get('meaning') or ''
        if r.get('aramaic'):
            meaning = (meaning + '  ·  ' if meaning else '') + 'ארמית: ' + r['aramaic']
        ws.append([sam_path, jew_path, verse, words, binyan, pron, meaning])

    widths = [42, 38, 60, 20, 16, 22, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical='top', wrap_text=True,
                                    horizontal='right')

    buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
    q = re.sub(r'[^\w֐-׿]+', '_', data.get('query', 'search'))[:40] or 'search'
    return send_file(buf, as_attachment=True,
                     download_name='תוצאות_חיפוש_%s.xlsx' % q,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def _snippet(text, word, span=70):
    t = text or ''
    i = t.find(word)
    if i < 0:
        return t[:span] + ('…' if len(t) > span else '')
    s = max(0, i - 22); e = min(len(t), i + len(word) + 48)
    return ('…' if s > 0 else '') + t[s:e] + ('…' if e < len(t) else '')


@app.route('/api/word_sources')
def api_word_sources():
    """For a tapped word: its root/entry from Tal's dictionary (with the citation
    locations), plus where the word also occurs in Tibåt Mårqe and the Samaritan-
    tradition (eyalk) sources. Shown in a popup."""
    word = (request.args.get('word') or '').strip()
    out = {'word': word, 'tal': [], 'tibat_marqe': [], 'eyalk': []}
    if len(_heb_fold(word)) < 2:
        return jsonify(out)
    try:
        for e in db.lookup_tal_dictionary(word, limit=4):
            out['tal'].append({
                'lemma': e.get('lemma'), 'pos': e.get('pos'), 'gloss_en': e.get('gloss_en'),
                'citations': [{'quote': q, 'ref': rf} for q, rf in (e.get('citations') or [])][:5],
            })
    except Exception:
        pass
    like = '%' + word + '%'
    conn = db.get_connection()
    try:
        for r in conn.execute(
                "SELECT book, section, book_title, aramaic, hebrew FROM tm_sections "
                "WHERE aramaic LIKE ? OR hebrew LIKE ? ORDER BY sort_key LIMIT 15",
                (like, like)).fetchall():
            letter = db._TM_HE_LETTER.get(r['book'], r['book'])
            out['tibat_marqe'].append({
                'label': 'ספר %s, §%s' % (letter, r['section']),
                'book_title': r['book_title'] or '',
                'snippet': _snippet(r['aramaic'] or r['hebrew'] or '', word),
            })
    except Exception:
        pass
    try:
        for r in conn.execute(
                "SELECT parsha, text FROM eyalk_sections WHERE text LIKE ? ORDER BY ord LIMIT 12",
                (like,)).fetchall():
            out['eyalk'].append({'parsha': r['parsha'] or '', 'snippet': _snippet(r['text'], word)})
    except Exception:
        pass
    conn.close()
    return jsonify(out)


if __name__ == '__main__':
    port = int(os.environ.get('PORT', '5000'))
    app.run(host='127.0.0.1', port=port, debug=False)
