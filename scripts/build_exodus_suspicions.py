# -*- coding: utf-8 -*-
"""
READ-ONLY. Build data/exodus_suspicions.xlsx — every Samaritan expansion I added
to Exodus, plus the items I SUSPECT may be wrong, for manual verification against
the Aziz book / שמות.txt. Lists each inserted passage with a confidence note.
Changes nothing.
"""
import sqlite3, sys, io, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# per anchor: (confidence, note) — my own assessment
NOTES = {
    (7, 18):  ('תקין כנראה', 'ביצוע מכת דם (וילך משה ואהרן) — פרק שומרוני קלאסי'),
    (8, 19):  ('תקין כנראה', 'ביצוע מכת ערוב'),
    (9, 5):   ('תקין כנראה', 'ביצוע מכת דבר'),
    (9, 19):  ('תקין כנראה', 'ביצוע מכת ברד'),
    (10, 2):  ('נוסף עכשיו', 'ציווי הארבה (ואמרת אל פרעה...) — הושלם לפסוק 2'),
    (10, 6):  ('חשוד', '"ויפן ויצא מעם פרעה" — קצר; ייתכן שהוא זנב של פסוק 6 ולא פרק נפרד. לבדוק מבנה'),
    (11, 3):  ('לבדוק', 'אזהרת בכורות; קיים בטקסט אך לא סומן באבחון כפרק חסר — לוודא מיקום/אורך'),
    (18, 25): ('תקין כנראה', 'קטע השופטים (Deut 1) — נוסף אחרי הפסוק'),
    (20, 13): ('תקין כנראה', 'מצוות מזבח הר גריזים'),
    (20, 17): ('תקין כנראה', 'הנביא כמשה'),
    (27, 19): ('חשוד', '"ועשית בגדי תכלת..." — לא סומן באבחון כפרק חסר; לוודא אם פרק אמיתי או גרסת פסוק קיים'),
    (40, 33): ('חשוד', '"ויכלה משה את כל המלאכה" — גרסה שומרונית של 40:33 המסורתי; ייתכן שכפילות'),
    (6, 9):   ('נוסף עכשיו', 'תוספת "ויאמרו אל משה חדל נא ממנו..." — הושלם לסוף פסוק 9'),
    (23, 19): ('תקין כנראה', 'תוספת "כי עשה זאת כזבח..." — הושלם לפסוק 19'),
    (32, 10): ('תקין כנראה', 'תוספת "ובאהרן התאנף..." — הושלם לפסוק 10'),
    (39, 21): ('תקין כנראה', 'תוספת "ויעשו את האורים..." — הושלם לפסוק 21'),
}


def keyn(s):
    s = str(s)
    return (int(s.split('-')[0]), int(s.split('-')[1])) if '-' in s else (int(s), 0)


def main():
    c = sqlite3.connect('data/torah.db'); c.row_factory = sqlite3.Row
    bid = c.execute('SELECT id FROM books WHERE name=?', ('שמות',)).fetchone()['id']
    # text-numbered (chapter insertions)
    ins = c.execute('''SELECT ch.number cn, v.number vn, v.text t FROM verses v
            JOIN chapters ch ON ch.id=v.chapter_id WHERE ch.book_id=? AND typeof(v.number)='text'
            ''', (bid,)).fetchall()
    rows = []
    for r in ins:
        anchor = (r['cn'], int(str(r['vn']).split('-')[0]))
        conf, note = NOTES.get(anchor, ('לבדוק', ''))
        rows.append((r['cn'], str(r['vn']), r['t'], conf, note))
    # completed-tail verses (integer, modified)
    for cn, vn in [(6, 9), (10, 2), (23, 19), (32, 10), (39, 21)]:
        r = c.execute('''SELECT v.text t FROM verses v JOIN chapters ch ON ch.id=v.chapter_id
                WHERE ch.book_id=? AND ch.number=? AND v.number=?''', (bid, cn, str(vn))).fetchone()
        conf, note = NOTES.get((cn, vn), ('לבדוק', ''))
        rows.append((cn, str(vn) + ' (השלמה)', r['t'], conf, note))
    rows.sort(key=lambda x: (x[0], keyn(x[1].split()[0])))
    # extra non-inserted suspicion
    rows.append((26, '35 — לא הוכנס', 'מזבח הקטורת "ועשית מזבח מקטיר קטרת"',
                 'אזעקת-שווא?', 'האבחון סימן "חסר אחרי 26:35", אך הטקסט קיים ב-DB ובטקסט בפרק 30. כנראה היסט-יישור. לבדוק אם הנוסח השומרוני ממקם אותו ב-26:35'))

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'חשדות שמות'
    ws.sheet_view.rightToLeft = True
    ws.append(['פרק', 'פסוק', 'הטקסט שהוכנס', 'רמת חשד', 'הערה / לבדוק מול עזיז'])
    for cc in ws[1]:
        cc.font = Font(bold=True, color='FFFFFF'); cc.fill = PatternFill('solid', fgColor='C00000')
        cc.alignment = Alignment(horizontal='center', wrap_text=True)
    fills = {'חשוד': 'FFC7CE', 'לבדוק': 'FFEB9C', 'אזעקת-שווא?': 'FFEB9C'}
    for cn, vn, t, conf, note in rows:
        ws.append([cn, vn, (t or '')[:120], conf, note])
        if conf in fills:
            ws.cell(row=ws.max_row, column=4).fill = PatternFill('solid', fgColor=fills[conf])
    for i, w in enumerate([7, 16, 75, 14, 60], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    wb.save('data/exodus_suspicions.xlsx')
    print('wrote data/exodus_suspicions.xlsx with %d rows' % len(rows))
    for cn, vn, t, conf, note in rows:
        if conf in ('חשוד', 'לבדוק', 'אזעקת-שווא?'):
            print('  [%s] %s:%s — %s' % (conf, cn, vn, note[:55]))


if __name__ == '__main__':
    main()
