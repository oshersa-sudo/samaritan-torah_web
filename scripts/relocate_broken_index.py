# -*- coding: utf-8 -*-
"""
Repair OCR verse-number misreads in root_index in one batch.

An entry is "broken" when the form is ABSENT at its stated verse (the same test
get_root_occurrences uses to hide a result: no word shares the root skeleton and
no word matches the transliteration with base >= 0.67). For each broken entry we
look in the SAME book+chapter for a verse that genuinely contains the form and is
not already indexed for this (root,pron). When exactly ONE such verse exists the
relocation is unambiguous; we auto-apply those whose target word actually carries
the root skeleton (morphologically certain), and leave the rest (pron-only unique,
ambiguous, or no-target) for the review workbook.

Usage:  py -3 scripts/relocate_broken_index.py            # dry run + report
        py -3 scripts/relocate_broken_index.py --apply
"""
import sqlite3, sys, os, io, re, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from app.services.database import _lat_cons, _word_pron_score
from app.services.hebrew_root import to_skeleton, word_matches_root

APPLY = '--apply' in sys.argv
KEEP = 0.67        # filter's keep threshold (entry not broken at/above this)
STRONG = 0.85      # confidence bar for a relocation target's pron match
WORD = re.compile('[א-ת]+')


def has_form(text, skel, pron, pc, bar):
    """True if some word in text carries the root skeleton or matches pron >= bar."""
    for w in WORD.findall(text or ''):
        if (skel and word_matches_root(w, skel)) or _word_pron_score(w, pron, pc)[0] >= bar:
            return True
    return False


def target_word(text, skel, pron, pc):
    """How a verse carries the form: 'both' (root skeleton AND a strong pron match —
    the exact inflection), 'skel' or 'pron' (only one signal), or None."""
    skn = prn = False
    for w in WORD.findall(text or ''):
        if skel and word_matches_root(w, skel):
            skn = True
        if _word_pron_score(w, pron, pc)[0] >= STRONG:
            prn = True
    if skn and prn:
        return 'both'
    return 'skel' if skn else ('pron' if prn else None)


def main():
    conn = sqlite3.connect('data/torah.db')
    conn.row_factory = sqlite3.Row
    vtext = {r['id']: r['text'] for r in conn.execute('SELECT id, text FROM verses')}
    # (book_id, chapter) -> [(verse_id, number)]
    chap_verses = {}
    vmeta = {}
    for r in conn.execute(
            '''SELECT v.id, v.number vs, c.number ch, c.book_id bk
               FROM verses v JOIN chapters c ON c.id = v.chapter_id'''):
        chap_verses.setdefault((r['bk'], r['ch']), []).append((r['id'], r['vs']))
        vmeta[r['id']] = (r['bk'], r['ch'], r['vs'])
    bookname = {r['id']: r['name'] for r in conn.execute('SELECT id, name FROM books')}

    rows = conn.execute(
        """SELECT id, root, root_norm, pron, verse_id FROM root_index
           WHERE verse_id IS NOT NULL AND pron IS NOT NULL AND TRIM(pron) <> ''"""
    ).fetchall()
    claimed = {}
    for r in rows:
        claimed.setdefault((r['root_norm'], r['pron']), set()).add(r['verse_id'])

    apply_list, review = [], []          # apply_list: (id, new_vid, new_vs)
    n_broken = n_unique = n_ambig = n_none = 0
    for r in rows:
        if r['verse_id'] not in vmeta:
            continue
        pc = _lat_cons(r['pron'])
        if not pc:
            continue                     # all-vowel pron: can't judge
        skel = to_skeleton(r['root_norm'])
        if has_form(vtext.get(r['verse_id']), skel, r['pron'], pc, KEEP):
            continue                     # not broken
        n_broken += 1
        bk, ch, _ = vmeta[r['verse_id']]
        cands = []
        for vid, vs in chap_verses.get((bk, ch), []):
            if vid == r['verse_id'] or vid in claimed[(r['root_norm'], r['pron'])]:
                continue
            how = target_word(vtext.get(vid), skel, r['pron'], pc)
            if how:
                cands.append((vid, vs, how))
        # auto-apply needs a SINGLE 'both' target (root skeleton + exact transliteration);
        # any other shape (skel-only, pron-only, or several candidates) goes to review
        both = [c for c in cands if c[2] == 'both']
        ov = vmeta[r['verse_id']][2]
        if len(both) == 1 and len(cands) == 1:
            n_unique += 1
            vid, vs, _h = both[0]
            apply_list.append((r['id'], vid, vs))
            review.append([bookname[bk], ch, ov, vs, r['root'], r['pron'], 'both',
                           (vtext.get(vid) or '')[:50], 'AUTO-APPLIED'])
        elif len(cands) == 1:
            n_unique += 1
            vid, vs, how = cands[0]
            review.append([bookname[bk], ch, ov, vs, r['root'], r['pron'], how,
                           (vtext.get(vid) or '')[:50], '%s-only (not auto-applied)' % how])
        elif len(cands) > 1:
            n_ambig += 1
            review.append([bookname[bk], ch, ov, '?', r['root'], r['pron'],
                           '%d candidates' % len(cands), '', 'ambiguous'])
        else:
            n_none += 1

    print('broken entries (form absent at stated verse): %d' % n_broken)
    print('  unique same-chapter target: %d  (skel-confident auto-apply: %d, pron-only review: %d)'
          % (n_unique, len(apply_list), n_unique - len(apply_list)))
    print('  ambiguous (multiple targets): %d' % n_ambig)
    print('  no same-chapter target:       %d' % n_none)
    print('\nsample auto-apply relocations:')
    for rid, vid, vs in apply_list[:15]:
        b, c, ov = vmeta[conn.execute('SELECT verse_id FROM root_index WHERE id=?', (rid,)).fetchone()[0]]
        print('   %s %d:%d -> :%d | %s' % (bookname[b], c, ov, vs, (vtext.get(vid) or '')[:42]))

    if APPLY and apply_list:
        bak = 'data/torah.db.bak13'
        if not os.path.exists(bak):
            shutil.copy2('data/torah.db', bak); print('\nbacked up ->', bak)
        conn.executemany('UPDATE root_index SET verse_id=?, verse=? WHERE id=?',
                         [(vid, vs, rid) for rid, vid, vs in apply_list])
        conn.commit()
        print('applied %d skel-confident relocations.' % len(apply_list))

    # review workbook
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'relocations to review'
        ws.sheet_view.rightToLeft = True
        ws.append(['ספר', 'פרק', 'פסוק נוכחי', 'פסוק מוצע', 'שורש', 'הגייה', 'התאמה', 'טקסט יעד', 'הערה'])
        for c in ws[1]:
            c.font = Font(bold=True, color='FFFFFF'); c.fill = PatternFill('solid', fgColor='C00000')
            c.alignment = Alignment(horizontal='center', wrap_text=True)
        for rec in review:
            ws.append(rec)
        for i, w in enumerate([9, 5, 9, 9, 10, 12, 14, 40, 22], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = 'A2'
        wb.save('data/index_relocations_review.xlsx')
        print('review rows: %d -> data/index_relocations_review.xlsx' % len(review))
    except ImportError:
        print('(openpyxl missing; review list has %d rows)' % len(review))
    if not APPLY:
        print('\n[dry-run] re-run with --apply to write.')
    conn.close()


if __name__ == '__main__':
    main()
