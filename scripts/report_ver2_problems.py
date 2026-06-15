# -*- coding: utf-8 -*-
"""Build a comprehensive Excel of every difference / problem found between the .txt
edition and the torah_aziz_ver2 edition (Genesis, Exodus, Leviticus), plus a sheet of
structural problems. Differences are taken from the pre-ver2 backups so the full set is
captured, each tagged with how it was handled."""
import sys
sys.path.insert(0, 'scripts')
from difflib import SequenceMatcher
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import re
import aziz_lib
from aziz_lib import parse_txt, norm, clean_word, RLM, HEB
from aziz2_lib import extract as extract2
from compare_aziz import load_parsha, parsha_for

PRE = {'Genesis': 'data/בראשית.txt.ver2.bak',
       'Exodus':  'data/שמות.txt.ver2.bak',
       'Leviticus': 'data/ויקרא.txt.ver2.bak'}
HE = {'Genesis': 'בראשית', 'Exodus': 'שמות', 'Leviticus': 'ויקרא',
      'Numbers': 'במדבר', 'Deuteronomy': 'דברים'}
HEBSET = set(chr(c) for c in range(0x05D0, 0x05EB))


def affix(raw):
    idx = [i for i, c in enumerate(raw) if c in HEBSET]
    if not idx:
        return ('', raw, '')
    return raw[:idx[0]], raw[idx[0]:idx[-1] + 1], raw[idx[-1] + 1:]


def core_clean(raw):
    return all(c in HEBSET for c in affix(raw)[1])


def classify(rs, hs, tag):
    """Return (status, reason) mirroring apply_ver2's decision."""
    if tag != 'replace' or len(rs) != len(hs):
        if any(w['table'] for w in hs):
            return 'הוחזק', 'טבלה מעורבבת בוורד'
        if max(len(rs), len(hs)) > 12:
            return 'הוחזק', 'בלוק גדול'
        return 'הוחזק', f'{tag} מרובה-מילים'
    # equal-length replace -> per token, but here block-level summary
    rw, hw = rs[0], hs[0]
    if hw['table']:
        return 'הוחזק', 'טבלה מעורבבת בוורד'
    if not core_clean(rw['raw']):
        return 'הוחזק', 'אסימון עם סימון פנימי'
    if (SequenceMatcher(None, norm(rw['word']), norm(hw['word'])).ratio() < 0.5
            and abs(len(rw['word']) - len(hw['word'])) >= 2):
        return 'הוחזק', 'מילים שונות מדי (יישור מסופק)'
    pre, _, suf = affix(rw['raw'])
    if any(c in pre + suf for c in '[]{}<>'):
        return 'הוחזק', 'סימון מיוחד בקצה'
    return 'תוקן', ''


def parse_deut_only(path):
    """Parse only the 'Deuteronomy'-labelled lines of דברים.txt (skip the stray
    Numbers 23-36 block), returning words like parse_txt with book='Deuteronomy'."""
    words = []
    with open(path, encoding='utf-8') as f:
        lines = f.read().split('\n')
    chap = verse = 0
    in_angle = False
    for li, line in enumerate(lines):
        if RLM not in line:
            continue
        if 'Deuteronomy' not in line and not (words and chap):
            # not yet into Deuteronomy section
            if 'Numbers' in line:
                continue
        if 'Numbers' in line:        # never include Numbers-labelled lines
            continue
        for m in aziz_lib._TOK.finditer(line):
            tok = m.group()
            if in_angle:
                if '>>' in tok:
                    in_angle = False
                continue
            if '<<' in tok:
                if '>>' not in tok:
                    in_angle = True
                continue
            mc = aziz_lib._CHV.match(tok)
            if mc:
                chap, verse = int(mc.group(1)), int(mc.group(2)); continue
            mv = aziz_lib._VMK.match(tok)
            if mv:
                verse = int(mv.group(1)); continue
            if aziz_lib._LAT.search(tok) or aziz_lib._DBL.match(tok):
                continue
            w = clean_word(tok)
            if w:
                words.append({'book': 'Deuteronomy', 'chap': chap, 'verse': verse,
                              'word': w, 'raw': tok, 'line': li,
                              'start': m.start(), 'end': m.end()})
    return words


def diff_rows(book, ref, hyp, ranges, applied):
    """Generic: produce difference rows between ref(txt) and hyp(ver2) word lists."""
    rows = []
    fix_lbl = 'תוקן' if applied else 'מומלץ לתיקון'
    sm = SequenceMatcher(None, [norm(w['word']) for w in ref],
                         [norm(w['word']) for w in hyp], autojunk=False)
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == 'equal':
            continue
        rs, hs = ref[i1:i2], hyp[j1:j2]
        chap = rs[0]['chap'] if rs else (ref[i1 - 1]['chap'] if i1 > 0 else 0)
        verse = rs[0]['verse'] if rs else (ref[i1 - 1]['verse'] if i1 > 0 else 0)
        if tag == 'replace' and (i2 - i1) == (j2 - j1):
            for k in range(i2 - i1):
                if norm(rs[k]['word']) == norm(hs[k]['word']):
                    continue
                status, reason = classify([rs[k]], [hs[k]], 'replace')
                if status == 'תוקן':
                    status = fix_lbl
                rows.append([HE.get(book, book), parsha_for(ranges, book, rs[k]['chap'], rs[k]['verse']),
                             rs[k]['chap'], rs[k]['verse'], 'שינוי מילה',
                             rs[k]['word'], hs[k]['word'], status, reason])
        else:
            typ = {'replace': 'החלפה', 'insert': 'תוספת בוורד', 'delete': 'חסר בוורד'}[tag]
            # summarise huge blocks (e.g. Deut 1-9 missing) into one row
            if max(len(rs), len(hs)) > 30:
                status = 'הוחזק'
                reason = f'בלוק ענק ({len(rs)}→{len(hs)} מילים) – ר\' בעיות מבניות'
                otext = (' '.join(w['word'] for w in rs))[:120] + ('…' if len(rs) > 20 else '')
                ntext = (' '.join(w['word'] for w in hs))[:120] + ('…' if len(hs) > 20 else '')
            else:
                status, reason = classify(rs, hs, tag)
                if status == 'תוקן':
                    status = fix_lbl
                otext = ' '.join(w['word'] for w in rs)
                ntext = ' '.join(w['word'] for w in hs)
            rows.append([HE.get(book, book), parsha_for(ranges, book, chap, verse),
                         chap, verse, typ, otext, ntext, status, reason])
    return rows


def text_diffs():
    ranges = load_parsha()
    docx = extract2()
    rows = []
    for book in PRE:
        ref = parse_txt(PRE[book], book)
        hyp = [w for w in docx if w['book'] == book]
        rows += diff_rows(book, ref, hyp, ranges, applied=True)
    return rows


STRUCTURAL = [
    ['קורפציה', 'דברים', 'כל הספר',
     'דברים.txt מכיל כפילות של במדבר 23-36 ואז דברים 10-34; דברים פרקים 1-9 חסרים מכל קבצי הטקסט',
     'חסום – דורש שחזור'],
    ['מספור', 'שמות', 'פרק כב',
     'המספור בקובץ קופץ 21→23 (פרק כב באותה שורה כמו כא); הטקסט קיים ופוענח נכון',
     'טופל'],
    ['טבלה מעורבבת', 'שמות', 'ב:א-י, ג:יג-יז, לד:א-ט',
     'קטעים בוורד מפוצלים לשני טורים בסדר משובש; אי אפשר ליישר אוטומטית',
     'דולג – לבדיקה ידנית'],
    ['טבלה מעורבבת', 'במדבר', 'נבואות בלעם',
     'טבלאות 2-טורים בסדר משובש (לא נכלל – במדבר לא תוקן)',
     'מחוץ להיקף'],
    ['טבלה מעורבבת', 'דברים', 'האזינו (לב)',
     'שירת האזינו בטבלת 2 טורים בסדר משובש',
     'דולג'],
    ['אי-התאמת DB', 'בראשית', 'מפרק ~4 והלאה',
     'חלוקת הפסוקים ב-torah.db (1514) שונה מ-txt (1533) – המספור מוסט',
     'דווח'],
    ['אי-התאמת DB', 'ויקרא', 'כל הספר',
     'torah.db 859 פסוקים מול txt 845 – חלוקה שונה',
     'דווח'],
    ['אי-התאמת DB', 'שמות', '75 פסוקים',
     '75 פסוקי-DB ללא מקבילה ב-txt (חלוקה ישנה)',
     'דווח'],
    ['DB לא עודכן', 'בראשית/שמות/ויקרא', '~182 פסוקים',
     '107 פסוקים "סטייה" + 75 "DB-בלבד" לא עודכנו ב-torah.db עקב אי-התאמת חלוקה',
     'דורש התאמה לפי תוכן'],
    ['סימנים שהושמטו', 'כל הספרים', ') ו- | ו- ..',
     'לפי בקשתך – הסימנים סוגריים, קו-אנכי, ונקודה כפולה לא הוטמעו',
     'לפי בקשה'],
]


def book_sheet_rows(book, ref, hyp):
    return diff_rows(book, ref, hyp, load_parsha(), applied=False)


def main():
    diffs = text_diffs()
    ranges = load_parsha()
    docx = extract2()
    # Numbers: full alignment vs במדבר.txt (never corrected)
    num_rows = diff_rows('Numbers', parse_txt('data/במדבר.txt', 'Numbers'),
                         [w for w in docx if w['book'] == 'Numbers'], ranges, applied=False)
    # Deuteronomy: only the Deut 10-34 portion exists in דברים.txt
    deut_rows = diff_rows('Deuteronomy', parse_deut_only('data/דברים.txt'),
                          [w for w in docx if w['book'] == 'Deuteronomy'], ranges, applied=False)

    wb = openpyxl.Workbook()
    thin = Side(style='thin', color='BBBBBB')
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, color):
        for c in ws[1]:
            c.font = Font(bold=True, color='FFFFFF', size=11)
            c.fill = PatternFill('solid', fgColor=color)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = bd

    fill_fix = PatternFill('solid', fgColor='E2EFDA')
    fill_hold = PatternFill('solid', fgColor='FCE4D6')

    def diff_sheet(ws, rows, color):
        ws.sheet_view.rightToLeft = True
        ws.append(['ספר', 'פרשה', 'פרק', 'פסוק', 'סוג', 'גרסת txt', 'גרסת ver2', 'טיפול', 'הערה'])
        style_header(ws, color)
        for r in rows:
            ws.append(r)
            i = ws.max_row
            f = fill_fix if r[7] in ('תוקן', 'מומלץ לתיקון') else fill_hold
            for c in ws[i]:
                c.border = bd
                c.alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
                c.fill = f
        for idx, wd in enumerate([9, 15, 5, 5, 16, 22, 22, 12, 26], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = wd
        ws.freeze_panes = 'A2'

    # sheet 1: Genesis/Exodus/Leviticus (applied)
    ws = wb.active
    ws.title = 'בראשית-שמות-ויקרא'
    diff_sheet(ws, diffs, '4472C4')
    # separate sheet: Numbers
    diff_sheet(wb.create_sheet('במדבר'), num_rows, '7030A0')
    # separate sheet: Deuteronomy
    diff_sheet(wb.create_sheet('דברים'), deut_rows, '548235')

    # sheet 2: structural problems
    ws2 = wb.create_sheet('בעיות מבניות')
    ws2.sheet_view.rightToLeft = True
    ws2.append(['קטגוריה', 'ספר', 'מיקום', 'תיאור', 'סטטוס'])
    style_header(ws2, 'C00000')
    for r in STRUCTURAL:
        ws2.append(r)
        for c in ws2[ws2.max_row]:
            c.border = bd
            c.alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
    for idx, wd in enumerate([16, 14, 18, 60, 24], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = wd
    ws2.freeze_panes = 'A2'

    # sheet 3: summary
    ws3 = wb.create_sheet('סיכום', 0)
    ws3.sheet_view.rightToLeft = True
    allrows = diffs + num_rows + deut_rows
    by_book = {}
    for r in allrows:
        by_book.setdefault(r[0], [0, 0])
        by_book[r[0]][0 if r[7] in ('תוקן', 'מומלץ לתיקון') else 1] += 1
    ws3.append(['סיכום הבדלים בין נוסח txt לנוסח ver2', ''])
    ws3.append(['', ''])
    ws3.append(['סה"כ הבדלי נוסח (כל הספרים)', len(allrows)])
    ws3.append(['', ''])
    ws3.append(['בראשית/שמות/ויקרא – הוחל בפועל:', len(diffs)])
    ws3.append(['  תוקנו', sum(1 for r in diffs if r[7] == 'תוקן')])
    ws3.append(['  הוחזקו לבדיקה ידנית', sum(1 for r in diffs if r[7] == 'הוחזק')])
    ws3.append(['', ''])
    ws3.append(['במדבר – השוואה בלבד (לא הוחל):', len(num_rows)])
    ws3.append(['דברים – השוואה בלבד (חסום, רק 10-34):', len(deut_rows)])
    ws3.append(['', ''])
    ws3.append(['פירוט לפי ספר (תוקן/מומלץ  |  לבדיקה):', ''])
    for bk, (f, h) in by_book.items():
        ws3.append([f'  {bk}: {f}  |  {h}', ''])
    ws3.append(['', ''])
    ws3.append(['בעיות מבניות (ר\' גיליון נפרד)', len(STRUCTURAL)])
    ws3.append(['', ''])
    ws3.append(['הערה: דברים 1-9 חסרים לגמרי ב-txt; ההשוואה כאן רק ל-10-34.', ''])
    ws3.append(['הערה: סימני העצירה :/. נוספו בנפרד (לא נספרים כאן).', ''])
    ws3['A1'].font = Font(bold=True, size=13)
    ws3.column_dimensions['A'].width = 50
    ws3.column_dimensions['B'].width = 10

    out = 'data/aziz_ver2_differences.xlsx'
    wb.save(out)
    print(f'Gen/Exod/Lev diffs: {len(diffs)} | Numbers: {len(num_rows)} | Deut(10-34): {len(deut_rows)}')
    print(f'structural problems: {len(STRUCTURAL)}')
    print(f'saved -> {out}')


if __name__ == '__main__':
    main()
