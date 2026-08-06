# -*- coding: utf-8 -*-
"""
Generalized version of exod_compare.py — reuses all its book-agnostic logic
(gematria parsing, split-verse merging, word-level chapter reconciliation,
punctuation-merge alignment) for any book. Reconciles a transcribed page-batch
scan against torah.db and writes a comparison xlsx. Read-only against
torah.db — proposes changes, does not apply any.

Usage (as a library — see lev_compare.py / num_compare.py for concrete runs):
  from book_compare import run_comparison
  run_comparison(book_id=3, book_name_he='ספר ויקרא', source_pdf='...',
                  transcript_dir='...', page_order=[...], out_xlsx='...')
"""
import os
import sys
import sqlite3

import openpyxl
from openpyxl.styles import Font

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import exod_compare as X

DB_PATH = X.DB_PATH


def run_comparison(book_id, book_name_he, source_pdf, transcript_dir, page_order, out_xlsx):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute('''SELECT c.number, v.number, v.id, v.text, v.sam_ch_id FROM verses v
                   JOIN chapters c ON c.id=v.chapter_id
                   JOIN books b ON b.id=c.book_id WHERE b.id=?
                   ORDER BY c.number, CAST(v.number AS INTEGER),
                            CASE WHEN instr(v.number,'-')>0
                                 THEN CAST(substr(v.number, instr(v.number,'-')+1) AS INTEGER)
                                 ELSE 0 END''', (book_id,))
    ours = {}
    verse_id_map = {}
    sam_ch_of = {}
    ours_by_chapter = {}
    for chnum, vnum, vid, text, sam_ch_id in cur.fetchall():
        ours[(chnum, vnum)] = text or ''
        verse_id_map[(chnum, vnum)] = vid
        sam_ch_of[vid] = sam_ch_id
        ours_by_chapter.setdefault(chnum, []).append((vnum, text or ''))
    mt_chapter_count = len(ours_by_chapter)

    entries = X.load_batches(transcript_dir=transcript_dir, page_order=page_order)
    entries = X.merge_split_verses(entries)
    canonical, expansions = X.reconcile(entries, ours_by_chapter)

    our_ordered = sorted(ours.keys(), key=X.sort_key)
    our_sam_end_after = set()
    for i, key in enumerate(our_ordered):
        vid = verse_id_map[key]
        this_sc = sam_ch_of.get(vid)
        nxt_sc = sam_ch_of.get(verse_id_map[our_ordered[i + 1]]) if i + 1 < len(our_ordered) else None
        if this_sc != nxt_sc:
            our_sam_end_after.add(key)

    incoming_sam_end_after = {k for k, v in canonical.items() if v['sam_end']}
    chapter_mismatches = sorted(incoming_sam_end_after.symmetric_difference(our_sam_end_after), key=X.sort_key)

    text_diffs = []
    identical = 0
    only_incoming = []
    only_ours = []
    shared = 0
    for key in sorted(canonical.keys(), key=X.sort_key):
        v = canonical[key]
        if key not in ours:
            only_incoming.append(key)
            continue
        shared += 1
        a = X.strip_punct(v['text'])
        b = X.strip_punct(ours[key])
        if a == b:
            identical += 1
        else:
            text_diffs.append({'ch': key[0], 'v': key[1], 'incoming': v['text'], 'ours': ours[key]})
    for key in ours:
        if key not in canonical:
            only_ours.append(key)

    punct_rows, colon_add, period_add, suppressed = X.build_punctuation_merge(canonical, ours)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'סיכום'
    ws.sheet_view.rightToLeft = True
    rows = [
        ('נושא', 'המקור החדש', 'הנוסח שלנו', 'הערה'),
        ('שם/מקור', book_name_he + ' — נוסח התורה השומרונית', 'data/torah.db', source_pdf),
        ('פרקים (חלוקה יהודית)', f'{len(set(k[0] for k in canonical))}', f'{mt_chapter_count}', ''),
        ('פסוקים משותפים', f'{shared}', '', 'מספר הפסוקים המשותפים לשתי המהדורות'),
        ('פסוקים זהים (אותיות בלבד)', f'{identical}', f'{len(text_diffs)} שונים', 'פיסוק לא נספר כאן'),
        ('פסוקים רק במקור החדש', f'{len(only_incoming)}', '', 'לא נמצא פסוק תואם אצלנו'),
        ('פסוקים רק אצלנו', '', f'{len(only_ours)}', 'לא נמצא פסוק תואם במקור החדש'),
        ('תוספות שומרוניות (סטיות ממספור רגיל)', f'{len(expansions)}', '', 'פסוקי הרחבה שומרוניים שאינם ממוספרים בחלוקה היהודית הרגילה'),
        ('אי-התאמות בחלוקה השומרונית', f'{len(chapter_mismatches)}', '', 'מקום בו סימון סוף-פרק שומרוני (.-‏) במקור אינו תואם את גבול sam_chapters אצלנו'),
    ]
    for r in rows:
        ws.append(r)
    for row in ws.iter_rows():
        for c in row:
            c.font = Font(name='Arial', size=11)

    ws2 = wb.create_sheet('חלוקה שומרונית - אי-התאמות')
    ws2.sheet_view.rightToLeft = True
    ws2.append(('פרק (יהודי)', 'פסוק (יהודי)', 'סוף-פרק שומרוני במקור?', 'סוף-פרק שומרוני אצלנו?'))
    for key in chapter_mismatches:
        ws2.append((key[0], key[1], 'כן' if key in incoming_sam_end_after else 'לא',
                    'כן' if key in our_sam_end_after else 'לא'))

    ws3 = wb.create_sheet('הבדלי נוסח (אותיות)')
    ws3.sheet_view.rightToLeft = True
    ws3.append(('פרק', 'פסוק', 'מהמקור החדש', 'אצלנו'))
    for d in text_diffs:
        ws3.append((d['ch'], d['v'], d['incoming'], d['ours']))

    ws4 = wb.create_sheet('הצעת מיזוג פיסוק')
    ws4.sheet_view.rightToLeft = True
    ws4.append(('פרק', 'פסוק', 'אצלנו (לפני)', 'הצעה (אחרי)', 'נקודתיים+', 'נקודה+', 'דוכא'))
    for p in punct_rows:
        ws4.append((p['ch'], p['v'], p['before'], p['after'], p['colon_add'], p['period_add'], p['suppressed']))

    ws5 = wb.create_sheet('תוספות שומרוניות')
    ws5.sheet_view.rightToLeft = True
    ws5.append(('פרק', 'סמוך לפסוק', 'מספר במקור (גולמי)', 'טקסט', 'סוף-פרק שומרוני?'))
    for x in sorted(expansions, key=lambda r: X.sort_key((r['chapter'], r['near_verse']))):
        ws5.append((x['chapter'], x['near_verse'], str(x['verse_num_raw']), x['text'],
                    'כן' if x['sam_end'] else 'לא'))

    ws6 = wb.create_sheet('פסוקים חסרים ועודפים')
    ws6.sheet_view.rightToLeft = True
    ws6.append(('סוג', 'פרק', 'פסוק', 'טקסט'))
    for key in sorted(only_incoming, key=X.sort_key):
        ws6.append(('רק במקור החדש', key[0], key[1], canonical[key]['text']))
    for key in sorted(only_ours, key=X.sort_key):
        ws6.append(('רק אצלנו', key[0], key[1], ours[key]))

    wb.save(out_xlsx)

    print(f'--- {book_name_he} ---')
    print('canonical verses reconciled:', len(canonical))
    print('expansions (samaritan-only / non-standard numbering):', len(expansions))
    print('shared verses:', shared, '| identical:', identical, '| differing:', len(text_diffs))
    print('only in incoming:', len(only_incoming), '| only in ours:', len(only_ours))
    print('chapter-division mismatches:', len(chapter_mismatches))
    print('punctuation-merge candidate verses:', len(punct_rows),
          '(colon+%d period+%d suppressed+%d)' % (colon_add, period_add, suppressed))
    print('saved:', out_xlsx)
    return {
        'canonical': len(canonical), 'expansions': len(expansions), 'shared': shared,
        'identical': identical, 'differing': len(text_diffs), 'only_incoming': len(only_incoming),
        'only_ours': len(only_ours), 'chapter_mismatches': len(chapter_mismatches),
        'punct_rows': len(punct_rows),
    }


def apply_punctuation(book_id, transcript_dir, page_order, apply=False):
    """Apply the punctuation-merge proposal to verses.text for the given book.
    Same treatment as שמות: source commas -> our colon convention, source
    periods -> our period convention, word-aligned. Does NOT touch wording
    differences, Samaritan expansions, or chapter division — punctuation only.
    Each UPDATE is guarded by the verse's CURRENT text matching the computed
    "before" value, so it's safe to re-run and never clobbers a verse that
    changed since this was computed."""
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute('''SELECT c.number, v.number, v.id, v.text, v.sam_ch_id FROM verses v
                   JOIN chapters c ON c.id=v.chapter_id
                   JOIN books b ON b.id=c.book_id WHERE b.id=?
                   ORDER BY c.number, CAST(v.number AS INTEGER),
                            CASE WHEN instr(v.number,'-')>0
                                 THEN CAST(substr(v.number, instr(v.number,'-')+1) AS INTEGER)
                                 ELSE 0 END''', (book_id,))
    ours = {}
    verse_id_map = {}
    ours_by_chapter = {}
    for chnum, vnum, vid, text, sam_ch_id in cur.fetchall():
        ours[(chnum, vnum)] = text or ''
        verse_id_map[(chnum, vnum)] = vid
        ours_by_chapter.setdefault(chnum, []).append((vnum, text or ''))

    entries = X.load_batches(transcript_dir=transcript_dir, page_order=page_order)
    entries = X.merge_split_verses(entries)
    canonical, _expansions = X.reconcile(entries, ours_by_chapter)

    punct_rows, _colon_add, _period_add, _suppressed = X.build_punctuation_merge(canonical, ours)

    changed = 0
    for p in punct_rows:
        key = (p['ch'], p['v'])
        vid = verse_id_map[key]
        if apply:
            cur.execute('UPDATE verses SET text=? WHERE id=? AND text=?',
                        (p['after'], vid, p['before']))
            if cur.rowcount:
                changed += 1
        else:
            changed += 1

    if apply:
        conn.commit()
        print(f'APPLIED: {changed} verses updated in {DB_PATH}')
    else:
        print(f'DRY RUN: {changed} verses would be updated. Re-run with apply=True to write.')
        for p in punct_rows[:5]:
            print(f"  {p['ch']}:{p['v']}  before: {p['before']}")
            print(f"          after : {p['after']}")
    conn.close()
    return changed
