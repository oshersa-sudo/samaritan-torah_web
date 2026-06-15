# -*- coding: utf-8 -*-
"""
Step A — fix the DB from the corrected .txt files, per the user's review of
data/db_vs_txt_diffs.xlsx:
  * sheet '2-marks-only'      -> set verses.text to the TXT value (full text, not
                                 the truncated column E) for those verses.
  * sheet '3-TXT-only verses' -> insert the verses the DB is missing.
Sheets '1-content' and '4-DB-only verses' are left untouched.

Only verses.text is changed / verses are inserted; all other columns and tables
(translations, masoretic, root_index, …) are preserved. Inserted verses inherit
the chapter and the Samaritan chapter of the preceding verse. Full DB backup.

Usage:  py -3 scripts/fix_db_from_txt.py            # dry run
        py -3 scripts/fix_db_from_txt.py --apply
"""
import sys, io, os, shutil, sqlite3
sys.path.insert(0, 'scripts')
import import_torah as IT
import openpyxl
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

APPLY = '--apply' in sys.argv


def load_sheets():
    wb = openpyxl.load_workbook('data/db_vs_txt_diffs.xlsx')
    marks = [(r[0], r[1], r[2]) for r in wb['2-marks-only'].iter_rows(min_row=2, values_only=True) if r[0]]
    txtonly = [(r[0], r[1], r[2]) for r in wb['3-TXT-only verses'].iter_rows(min_row=2, values_only=True) if r[0]]
    return marks, txtonly


def main():
    marks, txtonly = load_sheets()
    # full TXT text per (book_he, chap, verse)
    txt = {}
    for key, (he, fn, _o) in IT.BOOK_MAP.items():
        for cn, vs in IT.parse_book_file('data/' + fn).items():
            for vn, vd in vs.items():
                txt[(he, cn, vn)] = vd['text']

    conn = sqlite3.connect('data/torah.db'); conn.row_factory = sqlite3.Row
    bid = {r['name']: r['id'] for r in conn.execute('SELECT id,name FROM books')}

    # --- A1: marks-only -> verses.text := TXT ---
    upd, miss1 = [], []
    for he, ch, vsn in marks:
        row = conn.execute(
            '''SELECT v.id FROM verses v JOIN chapters c ON c.id=v.chapter_id
               WHERE c.book_id=? AND c.number=? AND v.number=?''', (bid[he], ch, vsn)).fetchone()
        nt = txt.get((he, ch, vsn))
        if row and nt is not None:
            upd.append((nt, row['id']))
        else:
            miss1.append((he, ch, vsn))

    # --- A2: TXT-only -> insert verses ---
    ins, miss2 = [], []
    for he, ch, vsn in txtonly:
        chap = conn.execute('SELECT id FROM chapters WHERE book_id=? AND number=?', (bid[he], ch)).fetchone()
        nt = txt.get((he, ch, vsn))
        if not chap or nt is None:
            miss2.append((he, ch, vsn)); continue
        # Samaritan chapter of the nearest preceding verse in this chapter
        prev = conn.execute(
            '''SELECT sam_ch_id FROM verses WHERE chapter_id=? AND number<? ORDER BY number DESC LIMIT 1''',
            (chap['id'], vsn)).fetchone()
        if not prev:
            prev = conn.execute(
                '''SELECT sam_ch_id FROM verses WHERE chapter_id=? ORDER BY number LIMIT 1''',
                (chap['id'],)).fetchone()
        ins.append((chap['id'], vsn, nt, prev['sam_ch_id'] if prev else None))

    print('A1 marks-only updates: %d   (unmatched: %s)' % (len(upd), miss1 or 'none'))
    print('A2 verses to insert  : %d   (unmatched: %s)' % (len(ins), miss2 or 'none'))
    for cid, vsn, nt, scid in ins:
        print('   INSERT %s -> %s' % (vsn, (nt or '')[:46]))

    if APPLY:
        bak = 'data/torah.db.bak_fixA'
        if not os.path.exists(bak):
            shutil.copy2('data/torah.db', bak); print('backed up ->', bak)
        conn.executemany('UPDATE verses SET text=? WHERE id=?', upd)
        conn.executemany(
            'INSERT INTO verses (chapter_id, number, text, sam_ch_id) VALUES (?,?,?,?)',
            ins)
        conn.commit()
        print('applied: %d text updates, %d inserts' % (len(upd), len(ins)))
        print('verses now:', conn.execute('SELECT COUNT(*) FROM verses').fetchone()[0])
    else:
        print('\n[dry-run] re-run with --apply to write.')
    conn.close()


if __name__ == '__main__':
    main()
