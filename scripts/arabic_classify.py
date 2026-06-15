# -*- coding: utf-8 -*-
"""
READ-ONLY. Add a 'סיווג' (category) column to data/arabic_mismatches.xlsx so the
2913 flagged Arabic verses can be triaged. Categories are deterministic, from the
Arabic-column content itself:
  - עברית בעמודת הערבית      : the field is Hebrew, no Arabic at all
  - ערבוב עברית+ערבית        : both Hebrew and Arabic present
  - ערבית של פסוק אחר (מספור מוטבע): Arabic carrying an embedded foreign verse-number
  - ערבית לא תואמת לפסוק      : real Arabic, wrong content, no embedded number
"""
import openpyxl, sys, io, re
from openpyxl.styles import Font, PatternFill, Alignment
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

HEB = re.compile('[א-ת]')
ARA = re.compile('[؀-ۿ]')
NUM = re.compile(r'\([٠-٩]+\)|\(\s*\d+\s*\)|[٠-٩]')


def classify(ar):
    ar = ar or ''
    h, a = bool(HEB.search(ar)), bool(ARA.search(ar))
    if h and not a:
        return 'עברית בעמודת הערבית'
    if h and a:
        return 'ערבוב עברית+ערבית'
    if NUM.search(ar):
        return 'ערבית של פסוק אחר (מספור מוטבע)'
    return 'ערבית לא תואמת לפסוק'


def main():
    path = 'data/arabic_mismatches.xlsx'
    wb = openpyxl.load_workbook(path); ws = wb.active
    # header
    hcell = ws.cell(row=1, column=7, value='סיווג')
    hcell.font = Font(bold=True, color='FFFFFF'); hcell.fill = PatternFill('solid', fgColor='C00000')
    hcell.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.column_dimensions['G'].width = 30
    from collections import Counter
    tally = Counter()
    for r in range(2, ws.max_row + 1):
        ar = ws.cell(row=r, column=5).value or ''
        cat = classify(ar)
        ws.cell(row=r, column=7, value=cat)
        tally[cat] += 1
    wb.save(path)
    print('classified %d rows -> %s' % (sum(tally.values()), path))
    for cat, n in tally.most_common():
        print('   %-34s %d' % (cat, n))


if __name__ == '__main__':
    main()
