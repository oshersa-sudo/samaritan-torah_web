# -*- coding: utf-8 -*-
"""
READ-ONLY (DB). Check that each verse's Arabic translation (verses.arabic_trans)
faithfully renders the meaning of the Hebrew Torah verse (verses.text). Verses
are sent in small batches to Claude, which flags only real meaning errors
(wrong/missing content, mistranslation, or Arabic that belongs to another
verse) — minor stylistic/dialectal differences are accepted. Flagged verses are
written to data/arabic_mismatches.xlsx. The DB is never modified.

Usage:
  py -3 scripts/arabic_check.py --limit 45          # pilot (first 45 verses)
  py -3 scripts/arabic_check.py                      # full run (all ~2960)
  py -3 scripts/arabic_check.py --model claude-haiku-4-5   # cheaper model
"""
import sqlite3, sys, io, os, json, re, time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import anthropic
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

BATCH = 15
MODEL = 'claude-opus-4-8'
OUT = 'data/arabic_mismatches.xlsx'
LIMIT = None
for i, a in enumerate(sys.argv):
    if a == '--limit':
        LIMIT = int(sys.argv[i + 1])
    elif a == '--model':
        MODEL = sys.argv[i + 1]
    elif a == '--out':
        OUT = sys.argv[i + 1]

SYS = (
    "You verify a Samaritan Arabic translation of the Torah against the Hebrew. "
    "For each numbered pair you get the Hebrew verse and its Arabic translation. "
    "Decide whether the Arabic faithfully conveys the MEANING of the Hebrew verse. "
    "Accept minor stylistic, dialectal, spelling, or word-order differences, and "
    "accept legitimate Samaritan readings. Flag a pair ONLY for a real meaning "
    "problem: wrong or missing key content, an outright mistranslation, opposite "
    "meaning, or Arabic text that clearly belongs to a different verse. "
    "Return ONLY JSON: {\"issues\":[{\"n\":<num>,\"reason\":\"<short reason>\"}]}. "
    "The reason MUST be written in Hebrew. "
    "If every pair is fine, return {\"issues\":[]}. No prose outside the JSON."
)


def save_xlsx(flagged, out):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'תרגום ערבי לא תקין'
    ws.sheet_view.rightToLeft = True
    ws.append(['ספר', 'פרק', 'פסוק', 'משפט מהתורה', 'תרגום ערבי', 'הבעיה'])
    for cc in ws[1]:
        cc.font = Font(bold=True, color='FFFFFF'); cc.fill = PatternFill('solid', fgColor='C00000')
        cc.alignment = Alignment(horizontal='center', wrap_text=True)
    for row in flagged:
        ws.append(list(row))
    for i, w in enumerate([10, 6, 6, 50, 50, 40], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    wb.save(out)


def load_key():
    if os.environ.get('ANTHROPIC_API_KEY'):
        return
    if os.path.exists('.env'):
        for ln in io.open('.env', encoding='utf-8', errors='ignore'):
            ln = ln.strip()
            if ln.startswith('ANTHROPIC_API_KEY=') and not ln.startswith('#'):
                os.environ['ANTHROPIC_API_KEY'] = ln.split('=', 1)[1].strip().strip('"\'')
                return


def main():
    load_key()
    if not os.environ.get('ANTHROPIC_API_KEY'):
        print('No ANTHROPIC_API_KEY found.'); return
    client = anthropic.Anthropic()

    c = sqlite3.connect('data/torah.db'); c.row_factory = sqlite3.Row
    rows = c.execute(
        '''SELECT b.name bk, b.order_n, ch.number cn, v.number vn, v.text he, v.arabic_trans ar
           FROM verses v JOIN chapters ch ON ch.id=v.chapter_id
           JOIN books b ON b.id=ch.book_id
           WHERE v.arabic_trans IS NOT NULL AND TRIM(v.arabic_trans)<>''
           ORDER BY b.order_n, ch.number, v.number''').fetchall()
    c.close()
    if LIMIT:
        rows = rows[:LIMIT]
    print('checking %d verses with model %s, batch %d' % (len(rows), MODEL, BATCH))

    flagged = []
    in_tok = out_tok = 0
    for s in range(0, len(rows), BATCH):
        chunk = rows[s:s + BATCH]
        lines = []
        for j, r in enumerate(chunk, 1):
            lines.append('%d. HE: %s\n   AR: %s' % (j, (r['he'] or '').strip(), (r['ar'] or '').strip()))
        msg = None
        for attempt in range(3):
            try:
                msg = client.messages.create(
                    model=MODEL, max_tokens=1500,
                    system=SYS,
                    messages=[{'role': 'user', 'content': '\n'.join(lines)}],
                )
                break
            except Exception as e:
                print('  [api error batch @%d attempt %d: %s]' % (s, attempt + 1, e))
                time.sleep(5 * (attempt + 1))
        if msg is None:
            print('  [skipping batch @%d after retries]' % s); continue
        in_tok += msg.usage.input_tokens
        out_tok += msg.usage.output_tokens
        txt = ''.join(b.text for b in msg.content if getattr(b, 'type', '') == 'text')
        m = re.search(r'\{.*\}', txt, re.S)
        issues = []
        if m:
            try:
                issues = json.loads(m.group(0)).get('issues', [])
            except Exception as e:
                print('  [parse error batch @%d: %s] %s' % (s, e, txt[:120]))
        for it in issues:
            n = it.get('n')
            if isinstance(n, int) and 1 <= n <= len(chunk):
                r = chunk[n - 1]
                flagged.append((r['bk'], r['cn'], r['vn'], (r['he'] or '').strip(),
                                (r['ar'] or '').strip(), it.get('reason', '')))
        print('  %4d/%d  flagged so far: %d' % (min(s + BATCH, len(rows)), len(rows), len(flagged)))
        if (s // BATCH) % 20 == 0 and flagged:
            save_xlsx(flagged, OUT)        # incremental checkpoint

    # cost (per claude-api skill cached pricing)
    price = {'claude-opus-4-8': (5, 25), 'claude-sonnet-4-6': (3, 15), 'claude-haiku-4-5': (1, 5)}
    pi, po = price.get(MODEL, (5, 25))
    cost = in_tok / 1e6 * pi + out_tok / 1e6 * po
    print('\ntokens  in=%d  out=%d   est. cost $%.4f' % (in_tok, out_tok, cost))

    save_xlsx(flagged, OUT)
    print('flagged %d / %d  -> %s' % (len(flagged), len(rows), OUT))


if __name__ == '__main__':
    main()
