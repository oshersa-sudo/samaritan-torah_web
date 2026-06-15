# -*- coding: utf-8 -*-
"""
Word-by-word comparison of the Torah .txt files against torah_aziz_ver2.docx.
READ-ONLY: writes a review workbook (data/txt_docx_diffs.xlsx); changes nothing.

It uses the global word alignment per book. The RELIABLE signal is the 1-to-1
`replace` blocks (both sides have a word at the same aligned position):
  word   - genuine letter-level difference (the real candidates to apply)
  split  - same letters, different word boundary (usually a docx OCR space, e.g.
           אברהם -> 'א ברהם'); note guesses which side is the OCR error
The add/remove blocks (one side only) are where alignment is unreliable — repeated
phrases make difflib match a word to the wrong copy, so the same words can show as
both add and remove. Those are flagged 'likely alignment artifact' when their words
re-appear in an opposite block in the same chapter; the rest are real one-sided
differences to verify by hand.
"""
import sys, io
sys.path.insert(0, 'scripts')
from difflib import SequenceMatcher
from aziz_lib import parse_txt, norm
from aziz2_lib import extract as extract2
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

FILES = {'Genesis': 'בראשית', 'Exodus': 'שמות', 'Leviticus': 'ויקרא',
         'Numbers': 'במדבר', 'Deuteronomy': 'דברים'}


def loc(ref, i):
    w = ref[min(i, len(ref) - 1)] if ref else None
    return (w['chap'], w['verse']) if w else (0, 0)


def split_note(rs, hs):
    if any(len(norm(w)) == 1 for w in hs.split()):
        return 'docx split (keep TXT)'
    if any(len(norm(w)) == 1 for w in rs.split()):
        return 'TXT split (use docx)'
    return 'boundary'


def main():
    doc = extract2()
    word_rows, split_rows, region_rows = [], [], []
    summary = {}
    for en, he in FILES.items():
        ref = parse_txt('data/%s.txt' % he, en)
        hyp = [w for w in doc if w['book'] == en]
        sm = SequenceMatcher(None, [norm(w['word']) for w in ref],
                             [norm(w['word']) for w in hyp], autojunk=False)
        ops = sm.get_opcodes()
        # collect one-sided blocks per chapter to spot cascade pairs
        by_chap = {}
        for tag, i1, i2, j1, j2 in ops:
            if tag in ('insert', 'delete'):
                ch = loc(ref, i1)[0]
                words = set(norm(w['word']) for w in (ref[i1:i2] if tag == 'delete' else hyp[j1:j2]))
                by_chap.setdefault(ch, []).append((tag, words))
        cnt = {'word': 0, 'split': 0, 'add': 0, 'remove': 0, 'artifact': 0}
        for tag, i1, i2, j1, j2 in ops:
            if tag == 'equal':
                continue
            ch, vs = loc(ref, i1)
            rs = ' '.join(w['word'] for w in ref[i1:i2])
            hs = ' '.join(w['word'] for w in hyp[j1:j2])
            if tag == 'replace' and max(i2 - i1, j2 - j1) <= 3:
                # only SMALL replace blocks are trustworthy word corrections; a large
                # block means difflib mis-paired non-corresponding words (cascade)
                if ''.join(norm(w['word']) for w in ref[i1:i2]) == \
                   ''.join(norm(w['word']) for w in hyp[j1:j2]):
                    cnt['split'] += 1
                    split_rows.append([he, ch, vs, rs, hs, split_note(rs, hs)])
                else:
                    cnt['word'] += 1
                    word_rows.append([he, ch, vs, rs, hs])
            elif tag == 'replace':
                cnt['artifact'] += 1
                region_rows.append([he, ch, vs, 'replace', 'TXT=[%s] / docx=[%s]' % (rs, hs),
                                    'large mismatched block — verify'])
            else:
                mine = set(norm(w['word']) for w in (ref[i1:i2] if tag == 'delete' else hyp[j1:j2]))
                opp = 'insert' if tag == 'delete' else 'delete'
                artifact = any(t == opp and (mine & ws) for t, ws in by_chap.get(ch, []))
                kind = 'remove' if tag == 'delete' else 'add'
                note = 'likely alignment artifact' if artifact else 'one-sided — verify'
                cnt['artifact' if artifact else kind] += 1
                region_rows.append([he, ch, vs, kind, rs or hs, note])
        summary[he] = cnt
        print('%-8s real-word=%-3d boundary=%-3d  add=%-3d remove=%-3d artifact=%-3d'
              % (he, cnt['word'], cnt['split'], cnt['add'], cnt['remove'], cnt['artifact']))

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()

    def sheet(title, header, data, widths):
        ws = wb.create_sheet(title)
        ws.sheet_view.rightToLeft = True
        ws.append(header)
        for c in ws[1]:
            c.font = Font(bold=True, color='FFFFFF'); c.fill = PatternFill('solid', fgColor='C00000')
            c.alignment = Alignment(horizontal='center', wrap_text=True)
        for r in data:
            ws.append(r)
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = 'A2'

    wb.remove(wb.active)
    sheet('1-word-diffs', ['ספר', 'פרק', 'פסוק', 'TXT', 'docx'], word_rows, [9, 5, 5, 28, 28])
    sheet('2-boundary', ['ספר', 'פרק', 'פסוק', 'TXT', 'docx', 'הערה'], split_rows, [9, 5, 5, 28, 28, 20])
    sheet('3-regions', ['ספר', 'פרק', 'פסוק', 'סוג', 'מילים', 'הערה'], region_rows, [9, 5, 5, 8, 40, 22])
    wb.save('data/txt_docx_diffs.xlsx')
    print('\nworkbook -> data/txt_docx_diffs.xlsx  (word=%d, boundary=%d, regions=%d)'
          % (len(word_rows), len(split_rows), len(region_rows)))


if __name__ == '__main__':
    main()
