# -*- coding: utf-8 -*-
"""Correct בראשית/שמות/ויקרא from torah_aziz_ver2.docx and add the Samaritan stop
marks (mid-verse only) plus any missing sof-pasuq.

A. consonant corrections: safe 1-to-1 replaces (same guards as the ver1 pipeline)
B. stop marks: ver2 ':' / '.' inserted after a txt word ONLY mid-verse (option 4 -
   verse ends keep ׃ alone, no doubling)
C. missing sof-pasuq: every verse must end with ׃ - add where the txt lacks it

Marks/corrections are only trusted in 'equal' blocks and clean 1-to-1 replaces;
tables, big blocks, inserts/deletes are skipped to a manual-review workbook.

Usage: py -3 scripts/apply_ver2.py [--apply]
"""
import sys, os, shutil
sys.path.insert(0, 'scripts')
from difflib import SequenceMatcher
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from aziz_lib import parse_txt, norm
from aziz2_lib import extract as extract2

FILES = {'Genesis': 'data/בראשית.txt', 'Exodus': 'data/שמות.txt',
         'Leviticus': 'data/ויקרא.txt'}
BOOK_HE = {'Genesis': 'בראשית', 'Exodus': 'שמות', 'Leviticus': 'ויקרא'}
HEBSET = set(chr(c) for c in range(0x05D0, 0x05EB))


def split_affix(raw):
    idx = [i for i, c in enumerate(raw) if c in HEBSET]
    if not idx:
        return ('', raw, '')
    a, b = idx[0], idx[-1]
    return raw[:a], raw[a:b + 1], raw[b + 1:]


def core_clean(raw):
    _, core, _ = split_affix(raw)
    return all(c in HEBSET for c in core)


def main(apply=False):
    docx = extract2()
    manual = []
    perbook = {}

    for book in FILES:
        ref = parse_txt(FILES[book], book)
        hyp = [w for w in docx if w['book'] == book]
        with open(FILES[book], encoding='utf-8') as f:
            lines = f.read().split('\n')

        n = len(ref)
        def verse_end(i):
            return i == n - 1 or (ref[i + 1]['chap'], ref[i + 1]['verse']) != \
                                 (ref[i]['chap'], ref[i]['verse'])

        edits = []          # (line, start, end, newtext)  [end==start => insert]
        st = dict(corr=0, corr_held=0, marks=0, sof=0)

        sm = SequenceMatcher(None, [norm(w['word']) for w in ref],
                             [norm(w['word']) for w in hyp], autojunk=False)

        def add_mark(i, mark):
            if mark in (':', '.') and not verse_end(i) and not ref[i]['raw'].endswith('׃'):
                edits.append((ref[i]['line'], ref[i]['end'], ref[i]['end'], mark))
                st['marks'] += 1

        for tag, i1, i2, j1, j2 in sm.get_opcodes():
            rs, hs = ref[i1:i2], hyp[j1:j2]
            if tag == 'equal':
                for k in range(i2 - i1):
                    add_mark(i1 + k, hyp[j1 + k]['mark'])
                continue
            if tag == 'replace' and (i2 - i1) == (j2 - j1):
                # token-aligned block: handle each pair as a 1-to-1 case
                for k in range(i2 - i1):
                    rw, hw = ref[i1 + k], hyp[j1 + k]
                    same = norm(rw['word']) == norm(hw['word'])
                    safe = (not hw['table'] and core_clean(rw['raw']) and
                            not (SequenceMatcher(None, norm(rw['word']), norm(hw['word'])).ratio() < 0.5
                                 and abs(len(rw['word']) - len(hw['word'])) >= 2))
                    pre, _, suf = split_affix(rw['raw'])
                    edge_ok = not any(c in pre + suf for c in '[]{}<>')
                    if not same and safe and edge_ok:
                        # combined consonant fix + (mid-verse) mark
                        mk = hw['mark'] if (hw['mark'] in (':', '.') and not verse_end(i1 + k)
                                            and not rw['raw'].endswith('׃')) else ''
                        edits.append((rw['line'], rw['start'], rw['end'], pre + hw['word'] + suf + mk))
                        st['corr'] += 1
                        if mk:
                            st['marks'] += 1
                    elif not same:
                        manual.append([BOOK_HE[book], rw['chap'], rw['verse'], 'שינוי',
                                       rw['word'], hw['word'], 'לא עבר סינון בטיחות'])
                        st['corr_held'] += 1
                    else:
                        add_mark(i1 + k, hw['mark'])
                continue
            # replace(unequal) / insert / delete -> manual review, no marks here
            orig = ' '.join(w['word'] for w in rs)
            new = ' '.join(w['word'] for w in hs)
            chap = rs[0]['chap'] if rs else (ref[i1 - 1]['chap'] if i1 > 0 else 0)
            verse = rs[0]['verse'] if rs else (ref[i1 - 1]['verse'] if i1 > 0 else 0)
            reason = 'טבלה' if any(w['table'] for w in hs) else \
                     ('בלוק גדול' if max(len(rs), len(hs)) > 12 else f'{tag} מרובה')
            manual.append([BOOK_HE[book], chap, verse, tag, orig, new, reason])

        # ---- C: missing sof-pasuq (every verse ends with ׃)
        for i in range(n):
            if verse_end(i) and '׃' not in ref[i]['raw']:
                edits.append((ref[i]['line'], ref[i]['end'], ref[i]['end'], '׃'))
                st['sof'] += 1

        perbook[book] = (lines, edits)
        print(f'{book}: corrections={st["corr"]} (held={st["corr_held"]}) '
              f'marks={st["marks"]} sof-pasuq-added={st["sof"]}')

    if apply:
        for book in FILES:
            path = FILES[book]
            bak = path + '.ver2.bak'
            if not os.path.exists(bak):
                shutil.copy2(path, bak)
            lines, edits = perbook[book]
            per_line = {}
            for (ln, s, e, txt) in edits:
                per_line.setdefault(ln, []).append((s, e, txt))
            for ln, eds in per_line.items():
                # apply right-to-left; for equal start, inserts (s==e) after replaces
                for (s, e, txt) in sorted(eds, key=lambda x: (x[0], x[1] == x[2]), reverse=True):
                    lines[ln] = lines[ln][:s] + txt + lines[ln][e:]
            with open(path, 'w', encoding='utf-8') as f:
                f.write('\n'.join(lines))
            print(f'wrote {path} (backup {bak})')

    # manual-review workbook
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'ver2 לבדיקה ידנית'
    ws.sheet_view.rightToLeft = True
    ws.append(['ספר', 'פרק', 'פסוק', 'סוג', 'מקורי (txt)', 'ver2', 'סיבה'])
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF'); c.fill = PatternFill('solid', fgColor='C00000')
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    for r in manual:
        ws.append(r)
    for i, w in enumerate([9, 6, 6, 12, 24, 24, 22], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    wb.save('data/aziz_ver2_manual_review.xlsx')
    print(f'\nmanual-review rows: {len(manual)} -> data/aziz_ver2_manual_review.xlsx')
    if not apply:
        print('[dry-run] re-run with --apply to write files.')


if __name__ == '__main__':
    main(apply='--apply' in sys.argv)
