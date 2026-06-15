# -*- coding: utf-8 -*-
"""
READ-ONLY. List every Torah verse (Jewish chapter division) that has NO Arabic
translation (verses.arabic_trans empty/NULL). Output: data/arabic_missing.xlsx
with ספר / פרק / פסוק / משפט מהתורה / תרגום ערבי (empty, since none exists).
Changes nothing in the DB.
"""
import sqlite3, sys, io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')


def main():
    c = sqlite3.connect('data/torah.db'); c.row_factory = sqlite3.Row
    rows = c.execute(
        '''SELECT b.name bk, ch.number cn, v.number vn, v.text
           FROM verses v JOIN chapters ch ON ch.id=v.chapter_id
           JOIN books b ON b.id=ch.book_id
           WHERE v.arabic_trans IS NULL OR TRIM(v.arabic_trans)=''
           ORDER BY b.order_n, ch.number, v.number''').fetchall()
    c.close()

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'חסר תרגום ערבי'
    ws.sheet_view.rightToLeft = True
    ws.append(['ספר', 'פרק', 'פסוק', 'משפט מהתורה', 'תרגום ערבי (לא קיים)'])
    for cc in ws[1]:
        cc.font = Font(bold=True, color='FFFFFF'); cc.fill = PatternFill('solid', fgColor='C00000')
        cc.alignment = Alignment(horizontal='center', wrap_text=True)
    for r in rows:
        ws.append([r['bk'], r['cn'], r['vn'], (r['text'] or '').strip(), ''])
    for i, w in enumerate([10, 6, 6, 70, 22], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    wb.save('data/arabic_missing.xlsx')

    print('verses missing Arabic: %d -> data/arabic_missing.xlsx' % len(rows))
    # per-book tally for the report
    from collections import Counter
    by = Counter(r['bk'] for r in rows)
    for bk in ['בראשית', 'שמות', 'ויקרא', 'במדבר', 'דברים']:
        print('   %-8s %d' % (bk, by.get(bk, 0)))


if __name__ == '__main__':
    main()
