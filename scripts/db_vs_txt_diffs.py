# -*- coding: utf-8 -*-
"""
READ-ONLY: report every difference between the DB (verses.text) and the corrected
.txt files. Writes data/db_vs_txt_diffs.xlsx; changes nothing.

Per common verse the difference is classified by comparing the Hebrew-LETTER
stream (after stripping the editorial marks []<>(){}! so old un-cleaned DB text
doesn't count as a difference):
  content - the actual letters differ (real text change: a correction, a genuine
            Samaritan/DB divergence, etc.)
  marks   - same letters, only stop-marks / spacing differ
Plus the structural gaps: verses only in the TXT, and verses only in the DB.
"""
import sys, io, re, sqlite3
sys.path.insert(0, 'scripts')
import import_torah as IT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

EDIT = re.compile(r'[\[\]<>(){}!]')
LET = re.compile('[א-ת]')
WORD = re.compile('[א-ת]+')


def letters(t):
    return ''.join(LET.findall(EDIT.sub('', t or '')))


def word_diff(a, b):
    from difflib import SequenceMatcher
    aw, bw = WORD.findall(EDIT.sub('', a or '')), WORD.findall(EDIT.sub('', b or ''))
    sm = SequenceMatcher(None, aw, bw, autojunk=False)
    out = []
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == 'equal':
            continue
        out.append('%s: [%s] / [%s]' % (tag, ' '.join(aw[i1:i2]), ' '.join(bw[j1:j2])))
    return ' ; '.join(out)


def main():
    c = sqlite3.connect('data/torah.db'); c.row_factory = sqlite3.Row
    content, marks, txt_only, db_only = [], [], [], []
    for key, (he, fn, _o) in IT.BOOK_MAP.items():
        chapters = IT.parse_book_file('data/' + fn)
        bid = c.execute('SELECT id FROM books WHERE name=?', (he,)).fetchone()['id']
        dbv = {(r['cn'], r['vn']): r['text'] for r in c.execute(
            'SELECT ch.number cn,v.number vn,v.text FROM verses v '
            'JOIN chapters ch ON ch.id=v.chapter_id WHERE ch.book_id=?', (bid,))}
        txtv = {(cn, vn): vd['text'] for cn, vs in chapters.items() for vn, vd in vs.items()}
        for k in sorted(set(txtv) - set(dbv)):
            txt_only.append([he, k[0], k[1], (txtv[k] or '')[:80]])
        for k in sorted(set(dbv) - set(txtv)):
            db_only.append([he, k[0], k[1], (dbv[k] or '')[:80]])
        for k in sorted(set(dbv) & set(txtv)):
            d, t = (dbv[k] or '').strip(), (txtv[k] or '').strip()
            if d == t:
                continue
            if letters(d) != letters(t):
                content.append([he, k[0], k[1], d[:60], t[:60], word_diff(d, t)[:80]])
            else:
                marks.append([he, k[0], k[1], d[:55], t[:55]])
    c.close()

    wb = openpyxl.Workbook(); wb.remove(wb.active)

    def sheet(title, header, data, widths):
        ws = wb.create_sheet(title); ws.sheet_view.rightToLeft = True
        ws.append(header)
        for cc in ws[1]:
            cc.font = Font(bold=True, color='FFFFFF'); cc.fill = PatternFill('solid', fgColor='C00000')
            cc.alignment = Alignment(horizontal='center', wrap_text=True)
        for r in data:
            ws.append(r)
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = 'A2'

    sheet('1-content', ['ספר', 'פרק', 'פסוק', 'DB', 'TXT', 'הבדל מילים'], content, [9, 5, 5, 26, 26, 30])
    sheet('2-marks-only', ['ספר', 'פרק', 'פסוק', 'DB', 'TXT'], marks, [9, 5, 5, 30, 30])
    sheet('3-TXT-only verses', ['ספר', 'פרק', 'פסוק', 'TXT'], txt_only, [9, 5, 5, 60])
    sheet('4-DB-only verses', ['ספר', 'פרק', 'פסוק', 'DB'], db_only, [9, 5, 5, 60])
    wb.save('data/db_vs_txt_diffs.xlsx')
    print('content diffs : %d' % len(content))
    print('marks-only    : %d' % len(marks))
    print('TXT-only verses: %d' % len(txt_only))
    print('DB-only verses : %d' % len(db_only))
    print('-> data/db_vs_txt_diffs.xlsx')


if __name__ == '__main__':
    main()
