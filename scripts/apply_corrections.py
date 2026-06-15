# -*- coding: utf-8 -*-
"""Apply the SAFE word-level corrections from torah_aziz.docx back into the .txt files,
strictly by alignment position (no blind search/replace).

Only 1-or-more-word in-place REPLACE blocks that pass every safety guard are applied:
  - the txt tokens of the block are on a single line, contiguous, same verse,
    with no verse marker / sof-pasuq (׃) inside the span,
  - each txt token is "clean" (Hebrew core with only edge punctuation, no brackets/
    braces/angle marks, no interior junk),
  - it is NOT a mere word-spacing artifact (docx split/merged the same letters),
  - it is NOT from a docx table and NOT a large [א]/[ב] doublet/order block.
Inserts, deletes, table blocks, big blocks and anything failing a guard are written to a
manual-review sheet instead.

Usage:  py -3 scripts/apply_corrections.py [--apply]
Without --apply it only reports (dry run) and writes the manual-review workbook.
"""
import sys, io, os, shutil, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.path.insert(0, 'scripts')
from difflib import SequenceMatcher
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from aziz_lib import parse_txt, norm, HEB
from aziz_docx import extract as extract_docx

FILES = {'Genesis': 'data/בראשית.txt', 'Exodus': 'data/שמות.txt'}
BOOK_HE = {'Genesis': 'בראשית', 'Exodus': 'שמות'}
HEBSET = set(chr(c) for c in range(0x05D0, 0x05EB))


def split_affix(raw):
    idx = [i for i, c in enumerate(raw) if c in HEBSET]
    if not idx:
        return ('', raw, '')
    a, b = idx[0], idx[-1]
    return raw[:a], raw[a:b + 1], raw[b + 1:]


def core_is_clean(raw):
    _, core, _ = split_affix(raw)
    return all(c in HEBSET for c in core)


def main(apply=False):
    docx = extract_docx()
    lines_by_book = {}
    edits_by_book = {}        # book -> list of (line, start, end, newtext)
    manual = []               # rows skipped for manual review
    stats = {}

    for book in ('Genesis', 'Exodus'):
        ref = parse_txt(FILES[book], book)
        hyp = [w for w in docx if w['book'] == book]
        with open(FILES[book], encoding='utf-8') as f:
            lines = f.read().split('\n')
        lines_by_book[book] = lines
        edits = []
        st = dict(applied=0, table=0, big=0, insdel=0, multi=0, interior=0,
                  edge=0, dissim=0, artifact=0)

        sm = SequenceMatcher(None, [norm(w['word']) for w in ref],
                             [norm(w['word']) for w in hyp], autojunk=False)
        for tag, i1, i2, j1, j2 in sm.get_opcodes():
            if tag == 'equal':
                continue
            ref_seg = ref[i1:i2]
            hyp_seg = hyp[j1:j2]
            chap = ref_seg[0]['chap'] if ref_seg else (ref[i1 - 1]['chap'] if i1 > 0 else 0)
            verse = ref_seg[0]['verse'] if ref_seg else (ref[i1 - 1]['verse'] if i1 > 0 else 0)
            orig = ' '.join(w['word'] for w in ref_seg)
            new = ' '.join(w['word'] for w in hyp_seg)
            from_table = any(w['table'] for w in hyp_seg)
            big = max(len(ref_seg), len(hyp_seg)) > 12

            reason = None
            if from_table:
                reason = 'טבלה בוורד – סדר מעורבב'; st['table'] += 1
            elif big:
                reason = 'בלוק גדול (כפילות/סדר)'; st['big'] += 1
            elif tag != 'replace':
                reason = 'הוספה/מחיקה – מבני'; st['insdel'] += 1
            elif norm(''.join(w['word'] for w in hyp_seg)) == norm(''.join(w['word'] for w in ref_seg)):
                reason = 'הבדל רווח בלבד (פיצול/איחוד מילה בוורד)'; st['artifact'] += 1
            elif len(ref_seg) != 1 or len(hyp_seg) != 1:
                # multi-word blocks: the docx side is riddled with spurious space-splits
                # (מחרב->מח רב) that would corrupt the txt; leave for manual review.
                reason = 'מרובה-מילים – לבדוק ידנית'; st['multi'] += 1
            elif not core_is_clean(ref_seg[0]['raw']):
                reason = 'אסימון עם סימון פנימי'; st['interior'] += 1
            elif (SequenceMatcher(None, norm(ref_seg[0]['word']), norm(hyp_seg[0]['word'])).ratio() < 0.5
                  and abs(len(ref_seg[0]['word']) - len(hyp_seg[0]['word'])) >= 2):
                reason = 'מילים שונות מדי – ייתכן יישור שגוי'; st['dissim'] += 1
            else:
                w0 = ref_seg[0]
                pre, _, suf = split_affix(w0['raw'])
                if any(ch in pre + suf for ch in '[]{}<>'):
                    reason = 'סימון מיוחד בקצה האסימון'; st['edge'] += 1
                else:
                    ln = w0['line']; s = w0['start']; e = w0['end']
                    newtext = pre + hyp_seg[0]['word'] + suf
                    edits.append((ln, s, e, newtext))
                    st['applied'] += 1

            if reason:
                manual.append([BOOK_HE[book], chap, verse, tag, orig, new, reason])

        edits_by_book[book] = edits
        stats[book] = st
        print(f'{book}: ' + ', '.join(f'{k}={v}' for k, v in st.items()))

    # ---- apply edits to files
    if apply:
        for book in ('Genesis', 'Exodus'):
            path = FILES[book]
            bak = path + '.aziz.bak'
            if not os.path.exists(bak):
                shutil.copy2(path, bak)
            lines = lines_by_book[book]
            per_line = {}
            for (ln, s, e, txt) in edits_by_book[book]:
                per_line.setdefault(ln, []).append((s, e, txt))
            for ln, eds in per_line.items():
                for (s, e, txt) in sorted(eds, key=lambda x: x[0], reverse=True):
                    lines[ln] = lines[ln][:s] + txt + lines[ln][e:]
            with open(path, 'w', encoding='utf-8') as f:
                f.write('\n'.join(lines))
            print(f'  wrote {path}  (backup {bak})')

    # ---- manual-review workbook
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'לבדיקה ידנית'
    ws.sheet_view.rightToLeft = True
    ws.append(['ספר', 'פרק', 'פסוק', 'סוג', 'גרסה מקורית (txt)', 'גרסה בוורד', 'סיבת הדילוג'])
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF'); c.fill = PatternFill('solid', fgColor='C00000')
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    for r in manual:
        ws.append(r)
    for i, w in enumerate([9, 6, 6, 10, 28, 28, 30], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    wb.save('data/aziz_manual_review.xlsx')

    tot_applied = sum(stats[b]['applied'] for b in stats)
    print(f'\n{"APPLIED" if apply else "DRY-RUN"}  total safe replaces = {tot_applied}, '
          f'manual-review = {len(manual)}  -> data/aziz_manual_review.xlsx')


def RLM_in(s):
    return '‏' in s or '‎' in s


if __name__ == '__main__':
    main(apply='--apply' in sys.argv)
