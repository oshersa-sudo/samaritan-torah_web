# -*- coding: utf-8 -*-
"""
Reconcile the 14 transcribed page-batches of the new Exodus CamScanner scan
into a single Masoretic chapter:verse-keyed "incoming" dataset, then compare
against the app's שמות text in torah.db:

  1. Samaritan chapter-division comparison (sam_chapter_end markers vs the
     app's sam_chapters boundaries).
  2. Word-for-word diff (punctuation ignored) per shared verse.
  3. Punctuation-merge proposal (source commas/periods -> app colon/period
     convention), matching the Genesis-precedent report format.
  4. Samaritan-only textual expansions present in the scan but absent from
     the app's verse numbering (extra sub-verses / duplicated numbers).

Writes everything to data/exodus_source_comparison.xlsx. Read-only against
torah.db — proposes changes, does not apply any.
"""
import os
import re
import json
import glob
import sqlite3
import unicodedata
import difflib

import openpyxl
from openpyxl.styles import Font

_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.path.join(_ROOT, 'data', 'torah.db')
TRANSCRIPT_DIR = os.environ.get(
    'EXOD_TRANSCRIPT_DIR',
    r'C:\Users\osher\AppData\Local\Temp\claude\C--Users-osher-Documents-torah\3b9a744a-b8e2-4630-b8a2-116404443172\scratchpad\exod_transcripts')
OUT_XLSX = os.path.join(_ROOT, 'data', 'exodus_source_comparison.xlsx')

PAGE_ORDER = [
    'pages_01-04', 'pages_05-08', 'pages_09-12', 'pages_13-16', 'pages_17-20',
    'pages_21-24', 'pages_25-28', 'pages_29-32', 'pages_33-36', 'pages_37-40',
    'pages_41-44', 'pages_45-48', 'pages_49-52', 'pages_53-57',
]

GEMATRIA = {
    'א': 1, 'ב': 2, 'ג': 3, 'ד': 4, 'ה': 5, 'ו': 6, 'ז': 7, 'ח': 8, 'ט': 9,
    'י': 10, 'כ': 20, 'ל': 30, 'מ': 40, 'נ': 50,
}


def gematria_to_int(label):
    if not label:
        return None
    letters = [c for c in label if c in GEMATRIA]
    return sum(GEMATRIA[c] for c in letters) or None


def load_batches(transcript_dir=None, page_order=None):
    transcript_dir = transcript_dir or TRANSCRIPT_DIR
    page_order = page_order or PAGE_ORDER
    entries = []
    for name in page_order:
        path = os.path.join(transcript_dir, name + '.json')
        with open(path, encoding='utf-8') as f:
            batch = json.load(f)
        for e in batch:
            e['_batch'] = name
            entries.append(e)
    return entries


def merge_split_verses(entries):
    """A verse straddling a page boundary is transcribed as two fragments by
    two independent agents — the batch's last entry (often missing its
    trailing punctuation) and the next batch's first entry (same reported
    verse_num, chapter_label usually null since no new chapter header was on
    that page). Detect these at batch seams and glue them back into one."""
    out = []
    i = 0
    while i < len(entries):
        e = entries[i]
        if (out and i > 0 and e['_batch'] != entries[i - 1]['_batch']
                and str(e['verse_num']) == str(entries[i - 1]['verse_num'])
                and not out[-1]['text'].rstrip().endswith(('.-', ':-', '-'))):
            prev = out[-1]
            prev['text'] = (prev['text'].rstrip() + ' ' + e['text'].lstrip()).strip()
            prev['text'] = re.sub(r' {2,}', ' ', prev['text'])
            prev['sam_chapter_end'] = prev['sam_chapter_end'] or e['sam_chapter_end']
            if not prev.get('chapter_label') and e.get('chapter_label'):
                prev['chapter_label'] = e['chapter_label']
            i += 1
            continue
        out.append(dict(e))
        i += 1
    return out


def bucket_by_chapter(entries):
    """Split entries into per-chapter lists using ONLY explicit chapter_label
    sightings as boundaries (all 40 are reliably present) — verse_num is too
    inconsistently labeled across independently-transcribed batches (Samaritan
    expansions use ad-hoc sub-numbering: '2a'/'2b', '18ב', repeated bare ints,
    etc.) to trust for chapter-boundary detection."""
    buckets = {}
    cur_ch = None
    for e in entries:
        label_ch = gematria_to_int(e.get('chapter_label'))
        if label_ch is not None:
            cur_ch = label_ch
        if cur_ch is None:
            continue
        buckets.setdefault(cur_ch, []).append(e)
    return buckets


def reconcile(entries, ours_by_chapter):
    """Word-level sequence alignment: within each chapter, align the incoming
    entries' words against our verses' words using difflib's exact-match
    SequenceMatcher over WORD tokens (not fuzzy per-verse ratios). Common
    words (proper nouns, ויאמר/אל/את, etc.) anchor the alignment even across
    a genuinely-differing verse, so — unlike a greedy per-verse threshold —
    one divergent verse doesn't permanently desync the rest of the chapter.
    Each incoming entry is assigned to whichever of our verses it shares the
    most aligned words with; an entry with no aligned words at all is a
    Samaritan-only expansion."""
    canonical = {}
    expansions = []
    buckets = bucket_by_chapter(entries)
    for ch, ents in buckets.items():
        ours_list = ours_by_chapter.get(ch, [])  # [(vnum, text), ...] in order
        our_words = []  # (word, vnum)
        for vnum, text in ours_list:
            for w in strip_punct(text).split(' '):
                if w:
                    our_words.append((w, vnum))
        inc_words = []  # (word, entry_index)
        for idx, e in enumerate(ents):
            for w in strip_punct(e['text']).split(' '):
                if w:
                    inc_words.append((w, idx))
        sm = difflib.SequenceMatcher(None, [w for w, _ in our_words], [w for w, _ in inc_words], autojunk=False)
        entry_vote = {}  # entry_index -> Counter(vnum -> matched word count)
        for i, j, n in sm.get_matching_blocks():
            for k in range(n):
                vnum = our_words[i + k][1]
                eidx = inc_words[j + k][1]
                entry_vote.setdefault(eidx, {}).setdefault(vnum, 0)
                entry_vote[eidx][vnum] += 1
        # assign each entry to its majority-overlap verse; concatenate entries
        # in original reading order per verse
        assigned = {}  # vnum -> [entry_idx,...]
        for idx in range(len(ents)):
            votes = entry_vote.get(idx)
            if not votes:
                continue
            best_vnum = max(votes.items(), key=lambda kv: kv[1])[0]
            assigned.setdefault(best_vnum, []).append(idx)
        for vnum, idxs in assigned.items():
            text = ' '.join(ents[i]['text'] for i in idxs)
            sam_end = any(ents[i]['sam_chapter_end'] for i in idxs)
            canonical[(ch, vnum)] = {'text': text, 'sam_end': sam_end, 'batch': ents[idxs[-1]]['_batch']}
        prev_vnum = None
        for idx, e in enumerate(ents):
            if idx not in entry_vote:
                expansions.append({
                    'chapter': ch, 'near_verse': prev_vnum or 0, 'verse_num_raw': e['verse_num'],
                    'text': e['text'], 'sam_end': e['sam_chapter_end'], 'batch': e['_batch'],
                })
            else:
                votes = entry_vote[idx]
                prev_vnum = max(votes.items(), key=lambda kv: kv[1])[0]
    return canonical, expansions


# ---- text normalization for word-diff (punctuation ignored) ----
_STRIP_RE = re.compile(r'[^\u05d0-\u05ea ]')


def strip_punct(text):
    t = _STRIP_RE.sub(' ', text or '')
    t = re.sub(r' {2,}', ' ', t).strip()
    return t


def sort_key(key):
    """(chapter, verse_number_string) -> sortable tuple, matching the app's
    own ORDER BY (base int part, then '-N' suffix if any)."""
    ch, vnum = key
    vnum = str(vnum)
    if '-' in vnum:
        base, suf = vnum.split('-', 1)
        return (ch, int(base), int(suf) if suf.isdigit() else 0)
    return (ch, int(vnum) if vnum.isdigit() else 0, 0)


def build_punctuation_merge(canonical, ours):
    """For every shared verse, map the source's internal commas/periods onto
    our text via word-level alignment (not a naive word-count offset, which
    breaks whenever our text and the source have a different word count for
    the verse — e.g. a genuine content difference): find, for each source
    punctuation mark, which SOURCE word immediately precedes it, then locate
    that same word's position in OUR text via difflib and insert the
    corresponding mark there (comma -> ':', period -> '.') instead. Returns
    (punct_rows, colon_add, period_add, suppressed)."""
    punct_rows = []
    colon_add = period_add = suppressed = 0
    for key in sorted(canonical.keys(), key=sort_key):
        v = canonical[key]
        if key not in ours:
            continue
        our_text = ours[key]
        src_text = v['text']
        src_clean_words = strip_punct(src_text).split(' ')
        our_words_raw = our_text.split(' ')
        our_stripped_words = [strip_punct(w) for w in our_words_raw]
        sm = difflib.SequenceMatcher(None, our_stripped_words, src_clean_words, autojunk=False)
        src_to_our = {}
        for i, j, n in sm.get_matching_blocks():
            for k in range(n):
                src_to_our[j + k] = i + k
        marks = []  # (src_word_idx_before_mark, mark_char)
        word_i = -1
        for tok in re.findall(r'[^\s,\.]+|[,\.]', src_text):
            if tok in (',', '.'):
                if word_i >= 0:
                    marks.append((word_i, ':' if tok == ',' else '.'))
            else:
                word_i += 1
        if not marks:
            continue
        out_words = our_words_raw[:]
        changed = False
        added_colon = added_period = sup = 0
        for src_idx, mark in marks:
            our_idx = src_to_our.get(src_idx)
            if our_idx is None:
                continue
            w = out_words[our_idx]
            if w and w[-1] in ':.׃':
                sup += 1
                continue
            out_words[our_idx] = w + mark
            changed = True
            if mark == ':':
                added_colon += 1
            else:
                added_period += 1
        if changed:
            colon_add += added_colon
            period_add += added_period
            suppressed += sup
            punct_rows.append({
                'ch': key[0], 'v': key[1], 'before': our_text, 'after': ' '.join(out_words),
                'colon_add': added_colon, 'period_add': added_period, 'suppressed': sup,
            })
    return punct_rows, colon_add, period_add, suppressed


def main():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    # exact sort order the app itself uses for Jewish-chapter verse display
    # (base integer part, then the '-N' Samaritan-expansion suffix if any) —
    # see app/services/database.py's get_verses_by_sam_ch.
    cur.execute('''SELECT c.number, v.number, v.id, v.text, v.sam_ch_id FROM verses v
                   JOIN chapters c ON c.id=v.chapter_id
                   JOIN books b ON b.id=c.book_id WHERE b.id=2
                   ORDER BY c.number, CAST(v.number AS INTEGER),
                            CASE WHEN instr(v.number,'-')>0
                                 THEN CAST(substr(v.number, instr(v.number,'-')+1) AS INTEGER)
                                 ELSE 0 END''')
    ours = {}
    verse_id_map = {}
    sam_ch_of = {}
    ours_by_chapter = {}
    for chnum, vnum, vid, text, sam_ch_id in cur.fetchall():
        ours[(chnum, vnum)] = text or ''
        verse_id_map[(chnum, vnum)] = vid
        sam_ch_of[vid] = sam_ch_id
        ours_by_chapter.setdefault(chnum, []).append((vnum, text or ''))

    entries = load_batches()
    entries = merge_split_verses(entries)
    canonical, expansions = reconcile(entries, ours_by_chapter)

    # sam_chapters boundaries: for each sam_ch_id, its verse id range (min/max by MT ch,v order)
    cur.execute('''SELECT sc.id, sc.number FROM sam_chapters sc
                   JOIN books b ON b.id=sc.book_id WHERE b.id=2''')
    sam_ch_number = dict(cur.fetchall())
    # ordered list of (ch,v) -> sam_ch_id for our data, in MT reading order
    our_ordered = sorted(ours.keys(), key=sort_key)
    our_sam_end_after = set()
    for i, key in enumerate(our_ordered):
        vid = verse_id_map[key]
        this_sc = sam_ch_of.get(vid)
        nxt_sc = sam_ch_of.get(verse_id_map[our_ordered[i + 1]]) if i + 1 < len(our_ordered) else None
        if this_sc != nxt_sc:
            our_sam_end_after.add(key)

    incoming_sam_end_after = {k for k, v in canonical.items() if v['sam_end']}

    chapter_mismatches = sorted(incoming_sam_end_after.symmetric_difference(our_sam_end_after), key=sort_key)

    text_diffs = []
    identical = 0
    only_incoming = []
    only_ours = []
    shared = 0
    for key in sorted(canonical.keys(), key=sort_key):
        v = canonical[key]
        if key not in ours:
            only_incoming.append(key)
            continue
        shared += 1
        a = strip_punct(v['text'])
        b = strip_punct(ours[key])
        if a == b:
            identical += 1
        else:
            text_diffs.append({'ch': key[0], 'v': key[1], 'incoming': v['text'], 'ours': ours[key],
                                'incoming_stripped': a, 'ours_stripped': b})
    for key in ours:
        if key not in canonical:
            only_ours.append(key)

    # ---- punctuation-merge proposal ----
    punct_rows, colon_add, period_add, suppressed = build_punctuation_merge(canonical, ours)

    # ---- write xlsx ----
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'סיכום'
    ws.sheet_view.rightToLeft = True
    rows = [
        ('נושא', 'המקור החדש', 'הנוסח שלנו', 'הערה'),
        ('שם/מקור', 'ספר ואלה שמות — נוסח התורה השומרונית', 'data/torah.db',
         'CamScanner 05.08.2026 01.35.pdf'),
        ('פרקים (חלוקה יהודית)', f'{len(set(k[0] for k in canonical))}', '40', ''),
        ('פסוקים משותפים', f'{shared}', '', 'מספר הפסוקים המשותפים לשתי המהדורות'),
        ('פסוקים זהים (אותיות בלבד)', f'{identical}', f'{len(text_diffs)} שונים', 'פיסוק לא נספר כאן'),
        ('פסוקים רק במקור החדש', f'{len(only_incoming)}', '', 'לא נמצא פסוק תואם אצלנו'),
        ('פסוקים רק אצלנו', '', f'{len(only_ours)}', 'לא נמצא פסוק תואם במקור החדש'),
        ('תוספות שומרוניות (סטיות ממספור רגיל)', f'{len(expansions)}', '', 'פסוקי הרחבה שומרוניים שאינם ממוספרים בחלוקה היהודית הרגילה'),
        ('אי-התאמות בחלוקה השומרונית', f'{len(chapter_mismatches)}', '', 'מקום בו סימון סוף-פרק שומרוני (.-\u200f) במקור אינו תואם את גבול sam_chapters אצלנו'),
    ]
    for r in rows:
        ws.append(r)
    for row in ws.iter_rows():
        for c in row:
            c.font = Font(name='Arial', size=11)

    ws2 = wb.create_sheet('חלוקה שומרונית - אי-התאמות')
    ws2.sheet_view.rightToLeft = True
    ws2.append(('פרק (יהודי)', 'פסוק (יהודי)', 'סוף-פרק שומרוני במקור?', 'סוף-פרק שומרוני אצלנו?'))
    for key in chapter_mismatches:
        ws2.append((key[0], key[1], 'כן' if key in incoming_sam_end_after else 'לא',
                    'כן' if key in our_sam_end_after else 'לא'))

    ws3 = wb.create_sheet('הבדלי נוסח (אותיות)')
    ws3.sheet_view.rightToLeft = True
    ws3.append(('פרק', 'פסוק', 'מהמקור החדש', 'אצלנו'))
    for d in text_diffs:
        ws3.append((d['ch'], d['v'], d['incoming'], d['ours']))

    ws4 = wb.create_sheet('הצעת מיזוג פיסוק')
    ws4.sheet_view.rightToLeft = True
    ws4.append(('פרק', 'פסוק', 'אצלנו (לפני)', 'הצעה (אחרי)', 'נקודתיים+', 'נקודה+', 'דוכא'))
    for p in punct_rows:
        ws4.append((p['ch'], p['v'], p['before'], p['after'], p['colon_add'], p['period_add'], p['suppressed']))

    ws5 = wb.create_sheet('תוספות שומרוניות')
    ws5.sheet_view.rightToLeft = True
    ws5.append(('פרק', 'סמוך לפסוק', 'מספר במקור (גולמי)', 'טקסט', 'סוף-פרק שומרוני?'))
    for x in sorted(expansions, key=lambda r: sort_key((r['chapter'], r['near_verse']))):
        ws5.append((x['chapter'], x['near_verse'], str(x['verse_num_raw']), x['text'],
                    'כן' if x['sam_end'] else 'לא'))

    ws6 = wb.create_sheet('פסוקים חסרים ועודפים')
    ws6.sheet_view.rightToLeft = True
    ws6.append(('סוג', 'פרק', 'פסוק', 'טקסט'))
    for key in sorted(only_incoming, key=sort_key):
        ws6.append(('רק במקור החדש', key[0], key[1], canonical[key]['text']))
    for key in sorted(only_ours, key=sort_key):
        ws6.append(('רק אצלנו', key[0], key[1], ours[key]))

    wb.save(OUT_XLSX)

    print('canonical verses reconciled:', len(canonical))
    print('expansions (samaritan-only / non-standard numbering):', len(expansions))
    print('shared verses:', shared, '| identical:', identical, '| differing:', len(text_diffs))
    print('only in incoming:', len(only_incoming), '| only in ours:', len(only_ours))
    print('chapter-division mismatches:', len(chapter_mismatches))
    print('punctuation-merge candidate verses:', len(punct_rows),
          '(colon+%d period+%d suppressed+%d)' % (colon_add, period_add, suppressed))
    print('saved:', OUT_XLSX)


if __name__ == '__main__':
    main()
