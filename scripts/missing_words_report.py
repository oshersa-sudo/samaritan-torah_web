# -*- coding: utf-8 -*-
"""
Per-chapter "which words are missing" report: reconciles each transcribed scan
against torah.db and lists, for every Samaritan chapter, the words the printed
source has that our text does not (and, separately, the ones we have that it
does not).

Read-only against torah.db - it reports, it never writes. Punctuation and the
matres lectionis are folded out before comparing, so the report shows genuine
missing WORDS rather than spelling variants; a spelling-variant count is kept
alongside so nothing is hidden.

Usage:
  py -3 scripts/missing_words_report.py                 # all three books
  py -3 scripts/missing_words_report.py --book 3        # Leviticus only
"""
import os
import re
import io
import sys
import argparse
import sqlite3
import difflib
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import exod_compare as X

_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH = X.DB_PATH

FIN = {'ך': 'כ', 'ם': 'מ', 'ן': 'נ',
       'ף': 'פ', 'ץ': 'צ'}


def words(s):
    """Letters only, final forms folded - punctuation must not read as a word."""
    s = ''.join(FIN.get(c, c) for c in re.sub(r'[^א-ת ]+', ' ', s or ''))
    return s.split()


def skel(w):
    """Consonantal skeleton: drops א ה ו י so מלא/חסר spellings collapse."""
    return re.sub(r'[אהוי]', '', w) or w


def run(book_id, book_he, transcript_dir, page_order, out_xlsx):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute('''SELECT c.number, v.number, v.id, v.text, v.sam_ch_id
                   FROM verses v JOIN chapters c ON c.id = v.chapter_id
                   WHERE c.book_id = ?
                   ORDER BY c.number, CAST(v.number AS INTEGER)''', (book_id,))
    ours, sam_of, by_ch = {}, {}, {}
    for chnum, vnum, vid, text, sam_ch_id in cur.fetchall():
        ours[(chnum, vnum)] = text or ''
        sam_of[(chnum, vnum)] = sam_ch_id
        by_ch.setdefault(chnum, []).append((vnum, text or ''))

    sam_num = dict(cur.execute(
        'SELECT id, number FROM sam_chapters WHERE book_id=?', (book_id,)).fetchall())
    conn.close()

    entries = X.load_batches(transcript_dir=transcript_dir, page_order=page_order)
    entries = X.merge_split_verses(entries)
    canonical, _expansions = X.reconcile(entries, by_ch)

    # per Samaritan chapter: missing / extra / spelling-only
    miss = defaultdict(list)      # sam_ch_id -> [(ch, verse, word)]
    extra = defaultdict(list)
    spell = defaultdict(list)     # same skeleton, different spelling
    dup = defaultdict(list)       # source repeats text we already have once
    verses_seen = defaultdict(set)

    def contains(seq, sub):
        """Is `sub` a contiguous run inside `seq`? Compared on the consonantal
        skeleton, because the source's second copy of a repeated formula spells
        a word or two differently (מלאים / מלואים) and an exact test would miss
        the duplication and report the whole copy as a gap."""
        s2 = [skel(w) for w in seq]
        b2 = [skel(w) for w in sub]
        n, m = len(s2), len(b2)
        if not m or m > n:
            return False
        return any(s2[i:i + m] == b2 for i in range(n - m + 1))

    # sort_key takes the (chapter, verse) key itself, not the (key, value) pair;
    # and reconcile() maps each key to a record, whose 'text' holds the reading
    for key, rec in sorted(canonical.items(), key=lambda kv: X.sort_key(kv[0])):
        if key not in ours:
            continue
        sid = sam_of.get(key)
        if sid is None:
            continue
        src_text = rec['text'] if isinstance(rec, dict) else rec
        a, b = words(ours[key]), words(src_text)
        verses_seen[sid].add(key)
        for op, i1, i2, j1, j2 in difflib.SequenceMatcher(None, a, b).get_opcodes():
            if op == 'equal':
                continue
            if op == 'replace' and (i2 - i1) == (j2 - j1):
                for x, y in zip(a[i1:i2], b[j1:j2]):
                    (spell if skel(x) == skel(y) else miss)[sid].append((key[0], key[1], y))
                    if skel(x) != skel(y):
                        extra[sid].append((key[0], key[1], x))
                continue
            # The printed source sometimes sets a repeated formula twice where we
            # carry it once (Numbers 7's offering formula does exactly this). The
            # second copy then shows up as a 20-word "gap" although our text is
            # complete, so a block that already appears verbatim in OUR verse is
            # a duplication in the source, not something missing from us.
            block = b[j1:j2]
            if len(block) >= 3 and contains(a, block):
                for y in block:
                    dup[sid].append((key[0], key[1], y))
            else:
                for y in block:
                    miss[sid].append((key[0], key[1], y))
            for x in a[i1:i2]:
                extra[sid].append((key[0], key[1], x))

    # A word counted "missing" from verse V but sitting in OUR text of another
    # verse of the same chapter is not lost text - it is the two sources
    # numbering the same words differently (the Gerizim expansion block in
    # Exodus 20:14-N is exactly this). Splitting the two is the difference
    # between a report that says "26 words missing" and the truth, which is
    # that the chapter is complete and only its verse split differs.
    def classify(sid):
        ours_words = defaultdict(int)
        for _c, _v, w in extra.get(sid, []):
            ours_words[w] += 1
        shifted, genuine = [], []
        for ch, vn, w in miss.get(sid, []):
            if ours_words.get(w, 0) > 0:
                ours_words[w] -= 1
                shifted.append((ch, vn, w))
            else:
                genuine.append((ch, vn, w))
        return genuine, shifted

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'סיכום לפי פרק'
    ws.sheet_view.rightToLeft = True
    ws.append(('פרק שומרוני', 'פסוקים שנבדקו', 'חסר באמת', 'הזזת מספור',
               'כפילות במקור', 'עודף אצלנו', 'הפרשי כתיב', 'המילים שחסרות באמת'))
    for c in ws[1]:
        c.font = Font(bold=True)

    rows = []
    gen_by_sid, shift_by_sid = {}, {}
    for sid in sorted(verses_seen, key=lambda s: int(sam_num.get(s, 0))):
        genuine, shifted = classify(sid)
        gen_by_sid[sid], shift_by_sid[sid] = genuine, shifted
        rows.append((sam_num.get(sid, '?'), len(verses_seen[sid]), len(genuine), len(shifted),
                     len(dup.get(sid, [])), len(extra.get(sid, [])), len(spell.get(sid, [])),
                     ' · '.join(w for _, _, w in genuine[:40])))
    for r in rows:
        ws.append(r)

    ws2 = wb.create_sheet('חסר באמת - פירוט')
    ws2.sheet_view.rightToLeft = True
    ws2.append(('פרק שומרוני', 'פרק (יהודי)', 'פסוק', 'המילה שבמקור וחסרה אצלנו'))
    for c in ws2[1]:
        c.font = Font(bold=True)
    for sid in sorted(gen_by_sid, key=lambda s: int(sam_num.get(s, 0))):
        for ch, vn, w in gen_by_sid[sid]:
            ws2.append((sam_num.get(sid, '?'), ch, vn, w))

    ws4 = wb.create_sheet('הזזת מספור - פירוט')
    ws4.sheet_view.rightToLeft = True
    ws4.append(('פרק שומרוני', 'פרק (יהודי)', 'פסוק', 'המילה (קיימת אצלנו בפסוק אחר בפרק)'))
    for c in ws4[1]:
        c.font = Font(bold=True)
    for sid in sorted(shift_by_sid, key=lambda s: int(sam_num.get(s, 0))):
        for ch, vn, w in shift_by_sid[sid]:
            ws4.append((sam_num.get(sid, '?'), ch, vn, w))

    ws3 = wb.create_sheet('מילים עודפות אצלנו')
    ws3.sheet_view.rightToLeft = True
    ws3.append(('פרק שומרוני', 'פרק (יהודי)', 'פסוק', 'המילה שאצלנו ואינה במקור'))
    for c in ws3[1]:
        c.font = Font(bold=True)
    for sid in sorted(extra, key=lambda s: int(sam_num.get(s, 0))):
        for ch, vn, w in extra[sid]:
            ws3.append((sam_num.get(sid, '?'), ch, vn, w))

    for w_ in (ws, ws2, ws3, ws4):
        for col, width in zip('ABCDEFGH', (14, 16, 12, 12, 12, 12, 12, 90)):
            w_.column_dimensions[col].width = width
    wb.save(out_xlsx)

    tg = sum(len(v) for v in gen_by_sid.values())
    tsh = sum(len(v) for v in shift_by_sid.values())
    te = sum(len(v) for v in extra.values())
    ts = sum(len(v) for v in spell.values())
    print('--- %s ---' % book_he)
    print('  פרקים שומרוניים שנבדקו : %d' % len(verses_seen))
    print('  פסוקים שהושוו          : %d' % sum(len(v) for v in verses_seen.values()))
    td = sum(len(v) for v in dup.values())
    print('  מילים חסרות באמת       : %d' % tg)
    print('  מזה רק הזזת מספור      : %d' % tsh)
    print('  כפילות במקור (לא חסר)  : %d' % td)
    print('  מילים עודפות אצלנו     : %d' % te)
    print('  הפרשי כתיב בלבד        : %d' % ts)
    nz = [r for r in sorted(rows, key=lambda r: -r[2]) if r[2]][:8]
    print('  הפרקים עם הכי הרבה חסר אמיתי :')
    for num, nv, g, sh, dp, e, sp, _s in nz:
        print('     פרק %-4s  חסר %3d | הזזה %3d | כפילות %3d | עודף %3d' % (num, g, sh, dp, e))
    print('  פרקים ללא שום חסר אמיתי : %d מתוך %d' % (sum(1 for r in rows if not r[2]), len(rows)))
    print('  -> %s\n' % out_xlsx)
    return {'chapters': len(verses_seen), 'genuine': tg, 'shifted': tsh,
            'extra': te, 'spelling': ts}


if __name__ == '__main__':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    ap = argparse.ArgumentParser()
    ap.add_argument('--book', type=int, help='2=שמות 3=ויקרא 4=במדבר (default: all)')
    args = ap.parse_args()

    import lev_compare as L
    import num_compare as N
    JOBS = {
        2: ('ספר שמות',  X.TRANSCRIPT_DIR, X.PAGE_ORDER, 'exodus_missing_words.xlsx'),
        3: ('ספר ויקרא', L.TRANSCRIPT_DIR, L.PAGE_ORDER, 'leviticus_missing_words.xlsx'),
        4: ('ספר במדבר', N.TRANSCRIPT_DIR, N.PAGE_ORDER, 'numbers_missing_words.xlsx'),
    }
    todo = [args.book] if args.book else sorted(JOBS)
    for bid in todo:
        he, td, po, out = JOBS[bid]
        run(bid, he, td, po, os.path.join(_ROOT, 'data', out))
