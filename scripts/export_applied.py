# -*- coding: utf-8 -*-
"""Export the list of corrections that were actually applied to the .txt files
(reconstructed from the *.aziz.bak originals) into an Excel workbook, with parsha."""
import sys
sys.path.insert(0, 'scripts')
from difflib import SequenceMatcher
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from aziz_lib import parse_txt, norm
from aziz_docx import extract as extract_docx
from compare_aziz import load_parsha, parsha_for

BAK = {'Genesis': 'data/בראשית.txt.aziz.bak', 'Exodus': 'data/שמות.txt.aziz.bak'}
BOOK_HE = {'Genesis': 'בראשית', 'Exodus': 'שמות'}
HEBSET = set(chr(c) for c in range(0x05D0, 0x05EB))


def split_affix(raw):
    idx = [i for i, c in enumerate(raw) if c in HEBSET]
    if not idx:
        return ('', raw, '')
    a, b = idx[0], idx[-1]
    return raw[:a], raw[a:b + 1], raw[b + 1:]


def core_is_clean(raw):
    _, core, _ = split_affix(raw)
    return all(c in HEBSET for c in core)


def applied_rows():
    docx = extract_docx()
    ranges = load_parsha()
    rows = []
    for book in ('Genesis', 'Exodus'):
        ref = parse_txt(BAK[book], book)
        hyp = [w for w in docx if w['book'] == book]
        sm = SequenceMatcher(None, [norm(w['word']) for w in ref],
                             [norm(w['word']) for w in hyp], autojunk=False)
        for tag, i1, i2, j1, j2 in sm.get_opcodes():
            if tag != 'replace':
                continue
            rs, hs = ref[i1:i2], hyp[j1:j2]
            if any(w['table'] for w in hs):                       continue
            if max(len(rs), len(hs)) > 12:                        continue
            if norm(''.join(w['word'] for w in hs)) == norm(''.join(w['word'] for w in rs)): continue
            if len(rs) != 1 or len(hs) != 1:                      continue
            if not core_is_clean(rs[0]['raw']):                   continue
            if (SequenceMatcher(None, norm(rs[0]['word']), norm(hs[0]['word'])).ratio() < 0.5
                    and abs(len(rs[0]['word']) - len(hs[0]['word'])) >= 2):  continue
            pre, _, suf = split_affix(rs[0]['raw'])
            if any(ch in pre + suf for ch in '[]{}<>'):           continue
            rows.append((book, parsha_for(ranges, book, rs[0]['chap'], rs[0]['verse']),
                         rs[0]['chap'], rs[0]['verse'], rs[0]['word'], hs[0]['word']))
    return rows


def main():
    rows = applied_rows()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'שינויים שבוצעו'
    ws.sheet_view.rightToLeft = True
    headers = ['ספר', 'פרשה', 'פרק', 'פסוק', 'מילה מקורית (txt)', 'מילה מתוקנת (וורד)']
    ws.append(headers)
    thin = Side(style='thin', color='BBBBBB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF', size=11)
        c.fill = PatternFill('solid', fgColor='548235')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
    g = e = 0
    for r in rows:
        if r[0] == 'Genesis':
            g += 1
        else:
            e += 1
        ws.append([BOOK_HE[r[0]], r[1], r[2], r[3], r[4], r[5]])
        for c in ws[ws.max_row]:
            c.border = border
            c.alignment = Alignment(horizontal='right', vertical='center')
    for i, wd in enumerate([9, 16, 6, 6, 24, 24], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = wd
    ws.freeze_panes = 'A2'
    out = 'data/aziz_applied_changes.xlsx'
    wb.save(out)
    print(f'rows: total={len(rows)} Genesis={g} Exodus={e}')
    print(f'Saved -> {out}')


if __name__ == '__main__':
    main()
