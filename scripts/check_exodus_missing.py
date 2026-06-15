# -*- coding: utf-8 -*-
"""
READ-ONLY check (changes nothing). Find Samaritan passages that exist in the
Exodus text source but are ABSENT from data/torah.db. Compares the DB's Exodus
verses against data/שמות.txt (word-level, whole book, so chapter-numbering
differences like the missing ch.22 header don't create false gaps) and confirms
each gap against the Aziz book (data/torah_aziz.docx). Output:
data/exodus_missing_check.xlsx.
"""
import sys, io, re, sqlite3
from difflib import SequenceMatcher
sys.path.insert(0, 'scripts')
import import_torah as IT
import docx, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

WORD = re.compile('[א-ת]+')
LET = re.compile('[א-ת]')
def words(t): return WORD.findall(t or '')
def lets(t): return ''.join(LET.findall(t or ''))


BOOK = next((a for a in sys.argv[1:] if not a.startswith('-')), 'שמות')
TRANSLIT = {'בראשית': 'genesis', 'שמות': 'exodus', 'ויקרא': 'leviticus',
            'במדבר': 'numbers', 'דברים': 'deuteronomy'}


def main():
    c = sqlite3.connect('data/torah.db'); c.row_factory = sqlite3.Row
    bk = [v for k, v in IT.BOOK_MAP.items() if v[0] == BOOK][0]
    chapters = IT.parse_book_file('data/' + bk[1])

    tw, ttag = [], []
    for cn in sorted(chapters):
        for vn in sorted(chapters[cn]):
            for w in words(chapters[cn][vn]['text']):
                tw.append(w); ttag.append((cn, vn))
    dbrows = c.execute(
        '''SELECT ch.number cn, v.number vn, v.text FROM verses v
           JOIN chapters ch ON ch.id=v.chapter_id JOIN books b ON b.id=ch.book_id
           WHERE b.name=? ORDER BY ch.number, v.number''', (BOOK,)).fetchall()
    dw, dw_chap = [], []
    for r in dbrows:
        for w in words(r['text']):
            dw.append(w); dw_chap.append(r['cn'])
    db_letters = ''.join(lets(r['text']) for r in dbrows)

    az = docx.Document('data/torah_aziz.docx')
    az_letters = lets(' '.join(p.text for p in az.paragraphs))

    sm = SequenceMatcher(None, dw, tw, autojunk=False)
    # collect TXT-side insert/replace ranges, then MERGE ones separated by only a
    # short matched run (<=4 words) — the repeated "ויבא משה ואהרן..." formula —
    # so each expansion is one passage, not split into formula + content.
    raw = [(j1, j2) for tag, i1, i2, j1, j2 in sm.get_opcodes()
           if tag in ('insert', 'replace')]
    merged = []
    for j1, j2 in raw:
        if merged and j1 - merged[-1][1] <= 4:
            merged[-1] = (merged[-1][0], j2)
        else:
            merged.append((j1, j2))

    from collections import defaultdict
    def renumbering_artifact(seg, chap):
        """True only if the passage appears as one near-contiguous block in the
        SAME Jewish chapter of the DB (i.e. content that exists but was renumbered,
        e.g. the ch.22-folding region). A contiguous match in a DIFFERENT chapter
        (the same plague-formula reused at another plague) does NOT count — that
        instance is genuinely missing here."""
        sm2 = SequenceMatcher(None, dw, seg, autojunk=False)
        off = defaultdict(list)
        for db_i, pj, size in sm2.get_matching_blocks():
            if size:
                off[db_i - pj].append((size, db_i))
        if not off:
            return False
        best_off = max(off, key=lambda o: sum(s for s, _ in off[o]))
        cov = sum(s for s, _ in off[best_off])
        if cov < 0.6 * len(seg):
            return False
        _, db_i = max(off[best_off])             # chapter of the largest matched block
        return dw_chap[db_i] == chap

    rows = []
    for j1, j2 in merged:
        seg = tw[j1:j2]
        if len(seg) < 6:
            continue
        text = ' '.join(seg)
        loc = ttag[j1]
        loc0 = loc[0]
        # the only renumbering artifact is the ch.21/22-folding region (the TXT
        # has no "Exodus 22" header, so its content shifts); everywhere else a
        # gap is a genuinely-missing Samaritan passage.
        in_db = renumbering_artifact(seg, loc0)
        in_az = (lets(' '.join(seg[:10])) in az_letters) or (lets(text) in az_letters)
        rows.append((loc[0], loc[1], len(seg), text, in_db, in_az))

    genuine = [r for r in rows if not r[4]]
    artifact = [r for r in rows if r[4]]

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'פרקים חסרים'
    ws.sheet_view.rightToLeft = True
    ws.append(['פרק (יהודי)', 'סמוך לפסוק', "מס' מילים", 'הטקסט השומרוני החסר',
               'קיים בעזיז', 'הערה'])
    for cc in ws[1]:
        cc.font = Font(bold=True, color='FFFFFF'); cc.fill = PatternFill('solid', fgColor='C00000')
        cc.alignment = Alignment(horizontal='center', wrap_text=True)
    for cn, vn, n, text, in_db, in_az in genuine:
        ws.append([cn, vn, n, text, 'כן' if in_az else 'לא נמצא מילולית',
                   'חסר ב-DB'])
    for i, w in enumerate([12, 10, 10, 90, 14, 18], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    out='data/%s_missing_check.xlsx'%TRANSLIT.get(BOOK,'book'); wb.save(out)

    print('=== %s: genuinely-missing Samaritan passages (absent from DB) ==='%BOOK)
    for cn, vn, n, text, in_db, in_az in genuine:
        print('  near %2d:%-3d (%3dw) [Aziz:%s]  %s%s' % (
            cn, vn, n, 'Y' if in_az else 'n', text[:70], '...' if len(text) > 70 else ''))
    print('\ngenuinely missing chunks: %d  (total %d words)'
          % (len(genuine), sum(r[2] for r in genuine)))
    print('alignment artifacts (present in DB elsewhere, NOT missing): %d' % len(artifact))
    for cn, vn, n, text, in_db, in_az in artifact:
        print('   [artifact] near %d:%d (%dw) %s...' % (cn, vn, n, text[:45]))
    print('\n->', out)
    c.close()


if __name__ == '__main__':
    main()
