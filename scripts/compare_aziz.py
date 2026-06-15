# -*- coding: utf-8 -*-
"""Compare torah_aziz.docx against בראשית.txt / שמות.txt word-by-word and write an
Excel table of every consonantal / word / sentence difference.

Output columns (Hebrew):
  ספר | פרשה | פרק | פסוק | סוג השינוי | גרסה מקורית (txt) | גרסה בוורד | הקשר (פסוק מלא)
"""
import sys, io, re, glob
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.path.insert(0, 'scripts')
from difflib import SequenceMatcher
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from aziz_lib import parse_txt, norm
from aziz_docx import extract as extract_docx

BOOK_HE = {'Genesis': 'בראשית', 'Exodus': 'שמות'}

# ---------------------------------------------------------------- parsha ranges
def load_parsha():
    path = glob.glob('data/portions*.xlsx')[0]
    wb = openpyxl.load_workbook(path)
    ws = wb['parasha'] if 'parasha' in wb.sheetnames else wb.active
    ranges = []  # (book, sch, sv, ech, ev, name)
    for r in ws.iter_rows(min_row=2):
        name = r[1].value or r[2].value      # Hebrew name, fallback translit
        book_raw = (r[3].value or '')
        chapters = (r[4].value or '')
        if not chapters or not str(chapters).strip():
            continue
        book = next((b for b in ('Genesis', 'Exodus', 'Leviticus', 'Numbers', 'Deuteronomy')
                     if b in book_raw), None)
        if not book:
            continue
        spec = str(chapters).strip()
        m = re.match(r'^\s*(\d+)(?::(\d+))?\s*-\s*(\d+)(?::(\d+))?\s*$', spec)
        if not m:
            continue
        sch = int(m.group(1)); sv = int(m.group(2)) if m.group(2) else 1
        ech = int(m.group(3)); ev = int(m.group(4)) if m.group(4) else 9999
        ranges.append((book, sch, sv, ech, ev, str(name).strip()))
    return ranges


def parsha_for(ranges, book, ch, verse):
    for (b, sch, sv, ech, ev, name) in ranges:
        if b != book:
            continue
        if (ch > sch or (ch == sch and verse >= sv)) and \
           (ch < ech or (ch == ech and verse <= ev)):
            return name
    return ''

# ---------------------------------------------------------------- alignment
def diff_book(book, refwords, hypwords):
    rn = [norm(w['word']) for w in refwords]
    hn = [norm(w['word']) for w in hypwords]
    sm = SequenceMatcher(None, rn, hn, autojunk=False)
    rows = []
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == 'equal':
            continue
        ref_seg = refwords[i1:i2]
        hyp_seg = hypwords[j1:j2]
        # attribution verse
        if ref_seg:
            chap, verse = ref_seg[0]['chap'], ref_seg[0]['verse']
        elif i1 > 0:
            chap, verse = refwords[i1 - 1]['chap'], refwords[i1 - 1]['verse']
        elif i2 < len(refwords):
            chap, verse = refwords[i2]['chap'], refwords[i2]['verse']
        else:
            chap, verse = 0, 0
        from_table = any(w['table'] for w in hyp_seg)
        rows.append({
            'book': book, 'chap': chap, 'verse': verse, 'tag': tag,
            'orig': ' '.join(w['word'] for w in ref_seg),
            'docx': ' '.join(w['word'] for w in hyp_seg),
            'table': from_table,
        })
    return rows

# ---------------------------------------------------------------- main
def main():
    ranges = load_parsha()
    txt = {'Genesis': parse_txt('data/בראשית.txt', 'Genesis'),
           'Exodus':  parse_txt('data/שמות.txt', 'Exodus')}
    docx = extract_docx()

    # full-verse text lookup for the context column
    verse_text = {}
    for b in txt:
        for w in txt[b]:
            verse_text.setdefault((b, w['chap'], w['verse']), []).append(w['word'])

    TAG_HE = {'replace': 'שינוי', 'delete': 'חסר בוורד', 'insert': 'עודף בוורד (חסר ב-txt)'}

    all_rows = []
    for book in ('Genesis', 'Exodus'):
        hyp = [w for w in docx if w['book'] == book]
        rows = diff_book(book, txt[book], hyp)
        print(f'{book}: {len(rows)} difference blocks')
        all_rows.extend(rows)

    # build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'הבדלים'
    ws.sheet_view.rightToLeft = True
    headers = ['ספר', 'פרשה', 'פרק', 'פסוק', 'סוג השינוי',
               'גרסה מקורית (txt)', 'גרסה בוורד (aziz)', 'הקשר - פסוק מלא מקורי', 'הערה']
    ws.append(headers)
    hdr_fill = PatternFill('solid', fgColor='4472C4')
    thin = Side(style='thin', color='BBBBBB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF', size=11)
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border

    fill_tbl = PatternFill('solid', fgColor='FFF2CC')   # yellow: table / manual review
    fill_ch22 = PatternFill('solid', fgColor='FCE4D6')  # orange: missing chapter
    n_table = n_big = n_clean = 0
    for r in all_rows:
        note = ''
        big = max(len(r['orig'].split()), len(r['docx'].split())) > 12
        if r['table']:
            note = 'מקור: טבלה בוורד – סדר מעורבב, לבדוק ידנית'
            n_table += 1
        elif big:
            note = 'בלוק גדול (כפילות [א]/[ב] או סדר שונה) – לבדוק ידנית'
            n_big += 1
        else:
            n_clean += 1
        ctx = ' '.join(verse_text.get((r['book'], r['chap'], r['verse']), []))
        row = [BOOK_HE[r['book']],
               parsha_for(ranges, r['book'], r['chap'], r['verse']),
               r['chap'], r['verse'], TAG_HE[r['tag']],
               r['orig'], r['docx'], ctx, note]
        ws.append(row)
        i = ws.max_row
        for c in ws[i]:
            c.border = border
            c.alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
        if r['table']:
            for c in ws[i]:
                c.fill = fill_tbl
        elif 'בלוק גדול' in note:
            for c in ws[i]:
                c.fill = fill_ch22

    widths = [9, 16, 6, 6, 14, 26, 26, 50, 28]
    for idx, wd in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = wd
    ws.freeze_panes = 'A2'

    # ------- summary sheet
    sm = wb.create_sheet('סיכום', 0)
    sm.sheet_view.rightToLeft = True
    g = sum(1 for r in all_rows if r['book'] == 'Genesis')
    e = sum(1 for r in all_rows if r['book'] == 'Exodus')
    info = [
        ['השוואת torah_aziz.docx מול קבצי הטקסט', ''],
        ['', ''],
        ['סה"כ בלוקי הבדל', len(all_rows)],
        ['  בראשית', g],
        ['  שמות', e],
        ['', ''],
        ['הבדלים "נקיים" (מילה/עיצור בודד) – עיקר העבודה', n_clean],
        ['בלוקים גדולים (כפילות [א]/[ב] או סדר) – צהוב כתום', n_big],
        ['מתוך הטבלאות בוורד (סדר מעורבב) – צהוב', n_table],
        ['', ''],
        ['מקרא צבעים:', ''],
        ['  לבן = הבדל רגיל לתיקון', ''],
        ['  כתום = בלוק גדול / כפילות – לבדוק ידנית', ''],
        ['  צהוב = מתוך טבלה בוורד – לבדוק ידנית', ''],
        ['', ''],
        ['הערה: אותיות סופיות (ך/כ) לא נספרות כהבדל.', ''],
        ['פרק כ"ב בשמות: המספור קופץ 21→23 אך הטקסט קיים תחת פרק כ"א.', ''],
    ]
    for row in info:
        sm.append(row)
    sm['A1'].font = Font(bold=True, size=13)
    sm.column_dimensions['A'].width = 55
    sm.column_dimensions['B'].width = 12

    out = 'data/aziz_comparison.xlsx'
    wb.save(out)
    print(f'\nTotal difference rows: {len(all_rows)}  (clean={n_clean} big={n_big} table={n_table})')
    print(f'Saved -> {out}')


if __name__ == '__main__':
    main()
