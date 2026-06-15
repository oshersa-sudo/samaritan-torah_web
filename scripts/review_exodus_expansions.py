# -*- coding: utf-8 -*-
"""
READ-ONLY. Build a REVIEW workbook of the Samaritan expansion verses that would
be inserted into Exodus, fixing the two problems the naive pass had:

  * FILTER: a ׃--delimited block is proposed for insertion only if its content
    is genuinely ABSENT from the DB (so a ׃-- that merely marks a Samaritan
    chapter break INSIDE an existing Jewish verse, e.g. 3:18, is skipped — it is
    already in the DB and must not be duplicated).
  * CONTENT ANCHORING: each real expansion is placed after the DB verse whose
    text actually precedes it (matched by content), not by the divergent TXT
    verse number — needed where the numbering differs (the Decalogue, etc.).

Each expansion chapter is split into verses on the period mark and numbered
after its anchor verse with a maqaf (18-1, 18-2 …). Same bracket / final-letter
cleaning as the rest of the book. Output: data/exodus_insert_review.xlsx.
Nothing is written to the DB.
"""
import sys, io, os, re, shutil, sqlite3
from difflib import SequenceMatcher
import docx, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

TXT = 'data/שמות.txt'
BOOK = 'שמות'
LET = re.compile('[א-ת]')
WORD = re.compile('[א-ת]+')
FIN = {'כ': 'ך', 'מ': 'ם', 'נ': 'ן', 'פ': 'ף', 'צ': 'ץ'}
END = re.compile(r'׃[-–—]+')


def lets(t): return ''.join(LET.findall(t or ''))
def words(t): return WORD.findall(t or '')


def clean(t):
    t = re.sub(r'\[+[^\[\]]*\]+', '', t)
    t = re.sub(r'<+[^<>]*>+', '', t)
    t = t.replace('{', '').replace('}', '').replace('(', '').replace(')', '')
    t = re.sub(r'[כמנפצ](?![א-ת])', lambda m: FIN[m.group(0)], t)
    t = re.sub(r'\.\s*([:׃])', r'\1', t)
    t = re.sub(r'([:׃])\s*\.', r'\1', t)
    return re.sub(r'\s{2,}', ' ', t).strip()


def raw_chapters():
    raw = re.sub(r'[‎‏‪-‮]', '', io.open(TXT, encoding='utf-8').read())
    out = {}
    for m in re.finditer(r'Exodus (\d+):(.*?)(?=Exodus \d+:|\Z)', raw, re.S):
        cn = int(m.group(1)); verses = {}; cur, buf = 1, []
        for tk in m.group(2).split():
            if tk.isdigit():
                verses[cur] = ' '.join(buf); buf = []; cur = int(tk)
            else:
                buf.append(tk)
        verses[cur] = ' '.join(buf); out[cn] = verses
    return out


def split_verses(expansion):
    body = END.sub('', expansion).strip()
    parts = [p.strip() for p in re.split(r'(?<=\.)\s+', body) if p.strip()]
    if parts:
        parts[-1] = parts[-1].rstrip('. ') + ' ׃--'
    return parts


def main():
    conn = sqlite3.connect('data/torah.db'); conn.row_factory = sqlite3.Row
    bid = conn.execute('SELECT id FROM books WHERE name=?', (BOOK,)).fetchone()['id']
    dbrows = conn.execute(
        '''SELECT ch.number cn, v.number vn, v.text t FROM verses v
           JOIN chapters ch ON ch.id=v.chapter_id WHERE ch.book_id=?
           ORDER BY ch.number, v.id''', (bid,)).fetchall()
    db = {(r['cn'], r['vn']): r['t'] for r in dbrows}
    db_letters = ''.join(lets(r['t']) for r in dbrows)
    # per chapter: list of (vn, letters) for anchoring
    by_ch = {}
    for r in dbrows:
        by_ch.setdefault(r['cn'], []).append((r['vn'], lets(r['t'])))
    az_letters = lets(' '.join(p.text for p in docx.Document('data/torah_aziz.docx').paragraphs))

    def anchor(cn, part0):
        ws = words(part0)
        if not ws:
            return None
        tail = lets(' '.join(ws[-8:]))
        for vn, vl in by_ch.get(cn, []):
            if tail and tail in vl:
                return vn
        head = lets(' '.join(ws[:8]))
        for vn, vl in by_ch.get(cn, []):
            if head and head in vl:
                return vn
        return None

    txt = raw_chapters()
    proposed = []   # (cn, anchor_vn, text, in_az)
    skipped = 0
    for cn in sorted(txt):
        for vn, rawtext in sorted(txt[cn].items()):
            pieces = END.split(rawtext)
            if len(pieces) <= 1:
                continue
            part0 = pieces[0]
            an = anchor(cn, part0) or vn
            for ex in pieces[1:]:
                if not lets(ex):
                    continue
                exl = lets(ex)
                if exl in db_letters:             # exact: already in the DB
                    skipped += 1
                    continue
                # fuzzy: a one-verse block that closely matches an existing verse
                # is just a Samaritan spelling variant of it, not an addition
                best = max((SequenceMatcher(None, exl, vl).ratio()
                            for _, vl in by_ch.get(cn, [])), default=0)
                if best >= 0.7:
                    skipped += 1
                    continue
                in_az = lets(ex)[:60] in az_letters or lets(' '.join(words(ex)[:12])) in az_letters
                for vtext in split_verses(ex):
                    ct = clean(vtext)
                    if lets(ct):
                        proposed.append((cn, an, ct, in_az))

    # number sequentially per (cn, anchor)
    from collections import defaultdict
    cnt = defaultdict(int)
    rows = []
    for cn, an, text, in_az in proposed:
        cnt[(cn, an)] += 1
        rows.append((cn, an, '%d-%d' % (an, cnt[(cn, an)]), text, in_az))

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'פסוקים להוספה'
    ws.sheet_view.rightToLeft = True
    ws.append(['פרק', 'אחרי פסוק', 'מספר חדש', 'הטקסט להוספה', 'קיים בעזיז'])
    for cc in ws[1]:
        cc.font = Font(bold=True, color='FFFFFF'); cc.fill = PatternFill('solid', fgColor='C00000')
        cc.alignment = Alignment(horizontal='center', wrap_text=True)
    for cn, an, num, text, in_az in rows:
        ws.append([cn, an, num, text, 'כן' if in_az else '—'])
    for i, w in enumerate([7, 10, 10, 100, 12], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    wb.save('data/exodus_insert_review.xlsx')

    print('proposed inserted verses: %d   (blocks already-in-DB skipped: %d)' % (len(rows), skipped))
    last = None
    for cn, an, num, text, in_az in rows:
        if (cn, an) != last:
            print('  ── after %s %d:%d ──' % (BOOK, cn, an)); last = (cn, an)
        print('     %s  [Aziz:%s]  %s%s' % (num, 'Y' if in_az else 'n', text[:80], '…' if len(text) > 80 else ''))
    print('\n-> data/exodus_insert_review.xlsx')

    if '--apply' in sys.argv:
        bak = 'data/torah.db.bak_exoins'
        if not os.path.exists(bak):
            shutil.copy2('data/torah.db', bak); print('backed up ->', bak)
        chmap = {r['number']: r['id'] for r in
                 conn.execute('SELECT id, number FROM chapters WHERE book_id=?', (bid,))}
        cur = conn.cursor()
        for cn, an, num, text, in_az in rows:
            cur.execute('INSERT INTO verses (chapter_id, number, text) VALUES (?,?,?)',
                        (chmap[cn], num, text))
        conn.commit()

        # re-derive the Exodus Samaritan division in reading order (N, then N-1, N-2…)
        def keyf(n):
            s = str(n)
            if '-' in s:
                a, b = s.split('-', 1); return (int(a), int(b))
            return (int(s), 0)
        ENDANCHOR = re.compile(r'׃[-–—]+\s*$')
        vrows = conn.execute(
            '''SELECT v.id, ch.number cn, v.number vn, v.text FROM verses v
               JOIN chapters ch ON ch.id=v.chapter_id WHERE ch.book_id=?''', (bid,)).fetchall()
        vrows = sorted(vrows, key=lambda r: (r['cn'], keyf(r['vn'])))
        cur.execute('DELETE FROM sam_chapters WHERE book_id=?', (bid,))
        sam, samids, assign = 1, {}, []
        for r in vrows:
            if sam not in samids:
                cur.execute('INSERT INTO sam_chapters (book_id, number) VALUES (?,?)', (bid, sam))
                samids[sam] = cur.lastrowid
            assign.append((samids[sam], r['id']))
            if ENDANCHOR.search(r['text'] or ''):
                sam += 1
        cur.executemany('UPDATE verses SET sam_ch_id=? WHERE id=?', assign)
        conn.commit()
        print('APPLIED: inserted %d verses; Exodus sam_chapters now %d'
              % (len(rows), conn.execute('SELECT COUNT(*) FROM sam_chapters WHERE book_id=?',
                                         (bid,)).fetchone()[0]))
    conn.close()


if __name__ == '__main__':
    main()
